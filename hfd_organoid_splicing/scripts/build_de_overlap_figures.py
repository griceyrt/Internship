#!/usr/bin/env python3
"""
Splicing vs DE gene overlap (Venn), HFD organoid splicing project
Author: Gricey

"Fig 3C equivalent" -- surface-level identity overlap only, no direction
split or dPSI-vs-log2FC correlation, since Desmond's DE files only give
gene identity + a pheatmap cluster number, no logFC values.

ID-TYPE NOTE: my splicing lists use Ensembl gene IDs, Desmond's DE files
use gene symbols only -- converted my side to symbols via
data/reference/gene_id_to_symbol.tsv before matching (symbol matching is
a bit lossier than ID matching, treat this as a surface-level check).

4 comparisons (W8/W18/W24/W42 vs WT_SD) -- matches Desmond's available
direct vs-WT_SD DE files.

Needs: pandas, matplotlib, matplotlib-venn (new dependency, not used
elsewhere in this project -- pip install matplotlib-venn if missing).

USAGE: python3 scripts/build_de_overlap_figures.py
Run from: hfd_organoid_splicing/
"""

import pandas as pd
import matplotlib.pyplot as plt
import matplotlib
from matplotlib_venn import venn2
matplotlib.use("Agg")

BASE = "."
SIG_DIR = f"{BASE}/results/suppa/significant"
DE_DIR = f"{BASE}/raw_data/Genes - RNA Seq DEGs - logFC above 1"
GENE_SYMBOLS = f"{BASE}/data/reference/gene_id_to_symbol.tsv"
FIGURES_DIR = f"{BASE}/figures"
TABLES_DIR = f"{BASE}/tables"

import os
os.makedirs(TABLES_DIR, exist_ok=True)

# (comparison label, splicing gene list file, DE gene list file)
COMPARISONS = [
    ("W8_WHFD", "significant_genes_W8_WHFD_vs_WT_SD.txt", "pw(pairwise_method)_genes_res_8w.txt"),
    ("W18_WHFD", "significant_genes_W18_WHFD_vs_WT_SD.txt", "pw_genes_res_18.txt"),
    ("W24_WHFD", "significant_genes_W24_WHFD_vs_WT_SD.txt", "pw_genes_res_24w.txt"),
    ("W42_WHFD", "significant_genes_W42_WHFD_vs_WT_SD.txt", "pw_genes_res_42w.txt"),
]

gene_symbols = pd.read_csv(GENE_SYMBOLS, sep="\t").set_index("gene_id")["gene_name"].to_dict()

overlap_rows = []
summary_rows = []

fig, axes = plt.subplots(1, 4, figsize=(18, 5))

for (label, splicing_file, de_file), ax in zip(COMPARISONS, axes):
    # Splicing genes: Ensembl IDs -> symbols
    splicing_ids = [l.strip() for l in open(f"{SIG_DIR}/{splicing_file}") if l.strip()]
    splicing_symbols = set()
    unmapped = 0
    for gid in splicing_ids:
        sym = gene_symbols.get(gid)
        if sym:
            splicing_symbols.add(sym)
        else:
            unmapped += 1

    # DE genes: already symbols
    de_df = pd.read_csv(f'{DE_DIR}/{de_file}')
    de_symbols = set(de_df["genes"].astype(str).str.strip())

    both = splicing_symbols & de_symbols
    splicing_only = splicing_symbols - de_symbols
    de_only = de_symbols - splicing_symbols

    print(f"{label}: splicing={len(splicing_symbols)} ({unmapped} unmapped IDs excluded), "
          f"DE={len(de_symbols)}, overlap={len(both)}")

    venn2(subsets=(len(splicing_only), len(de_only), len(both)),
          set_labels=("Significant\nsplicing", "Differentially\nexpressed"),
          ax=ax)
    ax.set_title(f"{label} vs WT_SD", fontsize=11, fontweight="bold")

    summary_rows.append({
        "comparison": label,
        "splicing_genes": len(splicing_symbols),
        "de_genes": len(de_symbols),
        "overlap": len(both),
    })
    for g in sorted(both):
        overlap_rows.append({"comparison": label, "gene_name": g})

fig.suptitle("Splicing-significant vs differentially-expressed genes\n"
             "(gene identity overlap only — no direction split, no logFC correlation)",
             fontsize=12, fontweight="bold")
plt.tight_layout()
png_path = f"{FIGURES_DIR}/de_splicing_overlap_venn.png"
svg_path = f"{FIGURES_DIR}/de_splicing_overlap_venn.svg"
plt.savefig(png_path, dpi=150, bbox_inches="tight")
plt.savefig(svg_path, bbox_inches="tight")
plt.close(fig)
print(f"\nSaved: {png_path}")
print(f"Saved: {svg_path}")

# xlsx companion table
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import date

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
BODY_FONT = Font(name="Arial")

def write_df_sheet(wb, sheet_name, df):
    ws = wb.create_sheet(sheet_name)
    ws.append(list(df.columns))
    for c in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for row in df.itertuples(index=False):
        ws.append(list(row))
    for r in range(2, len(df) + 2):
        for c in range(1, len(df.columns) + 1):
            ws.cell(row=r, column=c).font = BODY_FONT
    ws.freeze_panes = "A2"
    if len(df) > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"
    for i, col in enumerate(df.columns, start=1):
        max_len = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str)])
        ws.column_dimensions[get_column_letter(i)].width = min(max(max_len + 2, 10), 60)

wb = Workbook()
wb.remove(wb.active)

ws = wb.create_sheet("README", 0)
ws["A1"] = "Splicing vs differential expression — gene overlap"
ws["A1"].font = Font(name="Arial", bold=True, size=14)
ws["A3"] = "Pairs with figure: de_splicing_overlap_venn.png"
ws["A3"].font = Font(name="Arial", bold=True)
ws["A4"] = f"Generated: {date.today().isoformat()}"
ws["A4"].font = Font(name="Arial", italic=True, size=10)
notes = [
    "Surface-level overlap only: gene identity, no up/down direction split, no",
    "logFC correlation -- Desmond's DE files only provide gene symbol + a pheatmap",
    "cluster number, not logFC values.",
    "",
    "IDs matched by gene SYMBOL (our splicing lists use Ensembl IDs, converted via",
    "data/reference/gene_id_to_symbol.tsv; Desmond's DE files only have symbols).",
    "Symbol matching is inherently a bit lossier than ID matching -- treat as a",
    "surface-level check, not exact-rigor confirmation.",
    "",
    "4 comparisons only (W8/W18/W24/W42 vs WT_SD) -- matches what's available in",
    "both our splicing results and Desmond's direct vs-WT_SD RNA-seq DE files.",
]
for i, line in enumerate(notes):
    ws.cell(row=6 + i, column=1, value=line).font = Font(name="Arial")
ws.column_dimensions["A"].width = 100

write_df_sheet(wb, "Summary", pd.DataFrame(summary_rows))
write_df_sheet(wb, "Overlap genes", pd.DataFrame(overlap_rows))

xlsx_path = f"{TABLES_DIR}/de_splicing_overlap.xlsx"
wb.save(xlsx_path)
print(f"Saved: {xlsx_path}")

print("\nDone.")
