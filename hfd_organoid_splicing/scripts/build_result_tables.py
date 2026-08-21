#!/usr/bin/env python3
"""
xlsx tables paired with each figure family (HFD organoid splicing project)
Author: Gricey

Builds 4 workbooks, each paired with a figure family in figures/, scoped to
significant/summarized results only (not the raw PSI/event tables, which
run into the hundreds of thousands of rows):

  1. significant_events_by_type.xlsx           <- significant_events_by_type.png
  2. significant_gene_count_over_time.xlsx      <- significant_gene_count_over_time.png
  3. splicing_scatter_significant_events.xlsx   <- splicing_scatter_*.png (+ _bis)
  4. go_enrichment_significant_terms.xlsx       <- go_dotplot_*.png

Each workbook opens with a README sheet (thresholds, column definitions,
which figure it pairs with).

Run locally, not on the cluster -- lightweight. Needs: pandas, openpyxl.
USAGE: python3 scripts/build_result_tables.py
Run from: hfd_organoid_splicing/
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import date

BASE = "."
SIG_TABLE = f"{BASE}/results/suppa/significant/significant_events.tsv"
SIG_DIR = f"{BASE}/results/suppa/significant"
GO_DIR = f"{BASE}/results/suppa/go_enrichment"
GENE_SYMBOLS = f"{BASE}/data/reference/gene_id_to_symbol.tsv"
TABLES_DIR = f"{BASE}/tables"

import os
os.makedirs(TABLES_DIR, exist_ok=True)

COMPARISONS = ["W8_WHFD", "W18_WHFD", "W24_WHFD", "W42_WHFD"]
EVENT_TYPE_ORDER = ["SE", "MX", "RI", "A5", "A3", "AL", "AF"]

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
BODY_FONT = Font(name="Arial")
TITLE_FONT = Font(name="Arial", bold=True, size=14)
NOTE_FONT = Font(name="Arial", italic=True, size=10)

gene_symbols = pd.read_csv(GENE_SYMBOLS, sep="\t").set_index("gene_id")["gene_name"].to_dict()

def symbol_for(gene_id):
    return gene_symbols.get(gene_id, gene_id)


def style_header_row(ws, n_cols, row=1):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")


def autofit_columns(ws, df, min_width=10, max_width=60):
    for i, col in enumerate(df.columns, start=1):
        max_len = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str)])
        ws.column_dimensions[get_column_letter(i)].width = min(max(max_len + 2, min_width), max_width)


def write_df_sheet(wb, sheet_name, df):
    ws = wb.create_sheet(sheet_name)
    ws.append(list(df.columns))
    style_header_row(ws, len(df.columns))
    for row in df.itertuples(index=False):
        ws.append(list(row))
    for r in range(2, len(df) + 2):
        for c in range(1, len(df.columns) + 1):
            ws.cell(row=r, column=c).font = BODY_FONT
    ws.freeze_panes = "A2"
    if len(df) > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"
    autofit_columns(ws, df)
    return ws


def write_readme(wb, title, pairs_with, notes_lines):
    ws = wb.create_sheet("README", 0)
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A3"] = f"Pairs with figure: {pairs_with}"
    ws["A3"].font = Font(name="Arial", bold=True)
    ws["A4"] = f"Generated: {date.today().isoformat()}"
    ws["A4"].font = NOTE_FONT
    row = 6
    for line in notes_lines:
        ws.cell(row=row, column=1, value=line).font = BODY_FONT
        row += 1
    ws.column_dimensions["A"].width = 100


sig_df = pd.read_csv(SIG_TABLE, sep="\t")
sig_df["gene_name"] = sig_df["gene_id"].map(symbol_for)

# --- Workbook 1: significant_events_by_type.xlsx ---
wb1 = Workbook()
wb1.remove(wb1.active)
write_readme(
    wb1,
    "Significant splicing events per type, per timepoint",
    "significant_events_by_type.png",
    [
        "Counts of significant splicing events (|dPSI| > 0.1, p < 0.05) per event type, split",
        "by direction (up = higher PSI in HFD timepoint vs WT_SD; down = lower).",
        "",
        "Event types: SE=skipped exon, MX=mutually exclusive exons, RI=retained intron,",
        "A5/A3=alternative 5'/3' splice site, AL/AF=alternative last/first exon.",
        "",
        "One row per (comparison, event_type, direction) combination.",
    ],
)
counts = (sig_df.assign(direction=sig_df["dPSI"].apply(lambda x: "up" if x > 0 else "down"))
          .groupby(["comparison", "event_type", "direction"]).size()
          .reset_index(name="count"))
counts["comparison"] = pd.Categorical(counts["comparison"],
    categories=[f"{c}_vs_WT_SD" for c in COMPARISONS], ordered=True)
counts["event_type"] = pd.Categorical(counts["event_type"], categories=EVENT_TYPE_ORDER, ordered=True)
counts = counts.sort_values(["comparison", "event_type", "direction"])
write_df_sheet(wb1, "Event counts", counts)
wb1.save(f"{TABLES_DIR}/significant_events_by_type.xlsx")
print(f"Saved: {TABLES_DIR}/significant_events_by_type.xlsx")

# --- Workbook 2: significant_gene_count_over_time.xlsx ---
wb2 = Workbook()
wb2.remove(wb2.active)
write_readme(
    wb2,
    "Significant splicing genes over time",
    "significant_gene_count_over_time.png",
    [
        "Summary sheet: number of unique genes with >=1 significant splicing event",
        "(|dPSI| > 0.1, p < 0.05) per HFD timepoint vs WT_SD baseline.",
        "",
        "One sheet per timepoint (+ Union_All): every significant gene, ranked by",
        "n_significant_events (how many of its events were significant in that",
        "comparison) then by best_p_value (most significant event) -- genes at the",
        "top are the strongest/most robust candidates to highlight in future figures.",
    ],
)

summary_rows = []
for c in COMPARISONS:
    comp_df = sig_df[sig_df["comparison"] == f"{c}_vs_WT_SD"]
    summary_rows.append({"timepoint": c, "unique_gene_count": comp_df["gene_id"].nunique()})
summary_df = pd.DataFrame(summary_rows)
write_df_sheet(wb2, "Summary", summary_df)

for c in COMPARISONS + ["union_all_comparisons"]:
    if c == "union_all_comparisons":
        comp_df = sig_df.copy()
        sheet_name = "Union_All"
    else:
        comp_df = sig_df[sig_df["comparison"] == f"{c}_vs_WT_SD"]
        sheet_name = c
    gene_stats = (comp_df.groupby(["gene_id", "gene_name"])
                  .agg(n_significant_events=("event_id", "count"),
                       best_p_value=("p_value", "min"))
                  .reset_index()
                  .sort_values(["n_significant_events", "best_p_value"], ascending=[False, True]))
    write_df_sheet(wb2, sheet_name, gene_stats)
wb2.save(f"{TABLES_DIR}/significant_gene_count_over_time.xlsx")
print(f"Saved: {TABLES_DIR}/significant_gene_count_over_time.xlsx")

# --- Workbook 3: splicing_scatter_significant_events.xlsx ---
wb3 = Workbook()
wb3.remove(wb3.active)
write_readme(
    wb3,
    "Significant splicing events per comparison",
    "splicing_scatter_*_vs_WT_SD.png and *_bis.png (both styles, same data)",
    [
        "Every significant splicing event (|dPSI| > 0.1, p < 0.05) shown as a colored",
        "point in the PSI scatter plots, one sheet per timepoint comparison.",
        "",
        "dPSI = PSI(HFD timepoint) - PSI(WT_SD). Positive = higher PSI in HFD",
        "(green points in the figures); negative = lower PSI (red points).",
        "",
        "Sorted by p_value ascending (most significant first).",
    ],
)
for c in COMPARISONS:
    comp_df = sig_df[sig_df["comparison"] == f"{c}_vs_WT_SD"].copy()
    comp_df = comp_df[["gene_id", "gene_name", "event_type", "event_id", "dPSI", "p_value"]]
    comp_df = comp_df.sort_values("p_value")
    write_df_sheet(wb3, c, comp_df)
wb3.save(f"{TABLES_DIR}/splicing_scatter_significant_events.xlsx")
print(f"Saved: {TABLES_DIR}/splicing_scatter_significant_events.xlsx")

# --- Workbook 4: go_enrichment_significant_terms.xlsx ---
wb4 = Workbook()
wb4.remove(wb4.active)
write_readme(
    wb4,
    "GO Biological Process enrichment — significant terms",
    "go_dotplot_*.png",
    [
        "GO Biological Process terms significantly enriched (adjusted p-value < 0.05)",
        "among genes with significant splicing changes, one sheet per timepoint",
        "comparison (+ Union_All).",
        "",
        "NOTE: this is filtered to significant terms only. Each comparison tested",
        "6,000-8,500 GO terms total; only the significant ones are kept here to stay",
        "a reasonable size to browse -- full test results (incl. non-significant)",
        "remain on the cluster if ever needed.",
        "",
        "Sorted by adjusted p-value ascending. 'Genes' column lists the significant",
        "genes (by symbol) annotated to that term.",
    ],
)
for c in COMPARISONS + ["union_all_comparisons"]:
    csv_path = f"{GO_DIR}/go_enrichment_{c}_vs_WT_SD.csv" if c != "union_all_comparisons" \
        else f"{GO_DIR}/go_enrichment_union_all_comparisons.csv"
    df = pd.read_csv(csv_path)
    df = df[df["p.adjust"] < 0.05].sort_values("p.adjust")
    df = df[["ID", "Description", "Count", "GeneRatio", "pvalue", "p.adjust", "geneID"]].copy()
    df["geneID"] = df["geneID"].str.replace("/", ", ")
    df = df.rename(columns={"ID": "GO_ID", "geneID": "Genes"})
    sheet_name = "Union_All" if c == "union_all_comparisons" else c
    write_df_sheet(wb4, sheet_name, df)
wb4.save(f"{TABLES_DIR}/go_enrichment_significant_terms.xlsx")
print(f"Saved: {TABLES_DIR}/go_enrichment_significant_terms.xlsx")

print("\nDone.")
