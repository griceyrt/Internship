#!/usr/bin/env python3
"""
Sequencing/QC summary table (HFD organoid splicing project)
Author: Gricey

Per-sample read counts, read length, trimming retention, and salmon
mapping rate. Numbers are transcribed from the 02_fastp_trim.sh and
04_salmon_quant.sh job logs as hardcoded values, not re-parsed live -- if
the pipeline is ever rerun with new data, regenerate this table from the
new logs rather than trusting these values.

Needs: pandas, openpyxl. Run locally.
USAGE: python3 scripts/build_sequencing_stats_table.py
Run from: hfd_organoid_splicing/
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from datetime import date

BASE = "."
TABLES_DIR = f"{BASE}/tables"
import os
os.makedirs(TABLES_DIR, exist_ok=True)

# sample, condition_replicate, raw_reads, raw_bases, trimmed_reads, trimmed_bases, salmon_mapped_reads, salmon_mapping_pct
# Transcribed from 02_fastp_trim.sh log (15218991) and 04_salmon_quant.sh log (15219172)
DATA = [
    ("Lib1_70_25_S1",   "WT_SD_1",    34801577, 2628534679, 34062784, 2513053712, 27013158, 79.30),
    ("Lib2_70_25_S2",   "WT_SD_2",    33387433, 2521723096, 32659826, 2409888610, 25916221, 79.35),
    ("Lib3_70_25_S3",   "WT_SD_3",    33362510, 2519815699, 32522247, 2381482659, 24231755, 74.51),
    ("Lib4_70_25_S4",   "W8_WHFD_1",  33774912, 2551092287, 33051648, 2438958922, 25238056, 76.36),
    ("Lib5_70_25_S5",   "W8_WHFD_2",  29557646, 2232714153, 28894602, 2130160831, 21898632, 75.79),
    ("Lib6_70_25_S6",   "W8_WHFD_3",  33048949, 2496512113, 32270343, 2379180185, 24926160, 77.24),
    ("Lib7_70_25_S7",   "W18_WHFD_1", 33388078, 2522102220, 32623627, 2408447859, 25122456, 77.01),
    ("Lib8_70_25_S8",   "W18_WHFD_2", 37305735, 2818267843, 36509695, 2697465254, 27538807, 75.43),
    ("Lib9_70_25_S9",   "W18_WHFD_3", 34441798, 2602045209, 33681688, 2486962571, 26088480, 77.46),
    ("Lib10_70_25_S10", "W24_WHFD_1", 29977598, 2264138250, 29313636, 2159841903, 22709333, 77.47),
    ("Lib11_70_25_S11", "W24_WHFD_2", 31960857, 2414114536, 31305441, 2309821228, 24169451, 77.21),
    ("Lib12_70_25_S12", "W24_WHFD_3", 33983097, 2567405336, 33239634, 2450935151, 25829332, 77.71),
    ("Lib13_70_25_S13", "W42_WHFD_1", 29141289, 2201055664, 28448322, 2091616290, 21993005, 77.31),
    ("Lib14_70_25_S14", "W42_WHFD_2", 32153983, 2428483638, 31431497, 2312682294, 24346343, 77.46),
    ("Lib15_70_25_S15", "W42_WHFD_3", 29655192, 2240446869, 29057423, 2147163489, 22665911, 78.00),
]

rows = []
for sample, cond_rep, raw_reads, raw_bases, trim_reads, trim_bases, mapped, map_pct in DATA:
    condition, replicate = cond_rep.rsplit("_", 1)
    rows.append({
        "sample": sample,
        "condition": condition,
        "replicate": replicate,
        "raw_reads": raw_reads,
        "raw_read_length_bp": round(raw_bases / raw_reads, 1),
        "trimmed_reads": trim_reads,
        "pct_reads_retained_after_trim": round(100 * trim_reads / raw_reads, 2),
        "trimmed_mean_length_bp": round(trim_bases / trim_reads, 1),
        "salmon_mapped_reads": mapped,
        "salmon_mapping_rate_pct": map_pct,
    })

df = pd.DataFrame(rows)

# Overall summary
total_raw = df["raw_reads"].sum()
total_trim = df["trimmed_reads"].sum()
total_mapped = df["salmon_mapped_reads"].sum()
summary = pd.DataFrame([
    {"metric": "Total raw reads (all 15 samples)", "value": f"{total_raw:,}"},
    {"metric": "Total trimmed reads (all 15 samples)", "value": f"{total_trim:,}"},
    {"metric": "Overall retention after trimming", "value": f"{100*total_trim/total_raw:.2f}%"},
    {"metric": "Raw read length", "value": f"{df['raw_read_length_bp'].mean():.1f} bp (fixed-cycle Illumina)"},
    {"metric": "Trimmed mean read length", "value": f"{df['trimmed_mean_length_bp'].mean():.1f} bp (varies per read)"},
    {"metric": "Total salmon-mapped reads", "value": f"{total_mapped:,}"},
    {"metric": "Overall salmon mapping rate", "value": f"{100*total_mapped/total_trim:.2f}%"},
    {"metric": "Mapping rate range (min-max across samples)", "value": f"{df['salmon_mapping_rate_pct'].min():.1f}%-{df['salmon_mapping_rate_pct'].max():.1f}%"},
    {"metric": "Read type", "value": "Single-end, forward stranded (SF)"},
    {"metric": "Sequencer chemistry", "value": "Two-channel Illumina (NextSeq/NovaSeq-type, per poly-G trimming behavior)"},
])

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
BODY_FONT = Font(name="Arial")

def write_df_sheet(wb, sheet_name, df):
    ws = wb.create_sheet(sheet_name)
    ws.append(list(df.columns))
    for c in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for row in df.itertuples(index=False):
        ws.append(list(row))
    for r in range(2, len(df) + 2):
        for c in range(1, len(df.columns) + 1):
            ws.cell(row=r, column=c).font = BODY_FONT
    ws.freeze_panes = "A2"
    if len(df) > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"
    for i, col in enumerate(df.columns, start=1):
        max_len = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str)])
        ws.column_dimensions[get_column_letter(i)].width = min(max(max_len + 2, 10), 60)

wb = Workbook()
wb.remove(wb.active)

ws = wb.create_sheet("README", 0)
ws["A1"] = "Sequencing & QC summary — all 15 samples"
ws["A1"].font = Font(name="Arial", bold=True, size=14)
ws["A3"] = f"Generated: {date.today().isoformat()}"
ws["A3"].font = Font(name="Arial", italic=True, size=10)
notes = [
    "Raw read length = total bases / total reads (uniform, fixed-cycle Illumina run).",
    "Trimmed mean length = average post-trim (varies per read -- fastp trims each",
    "read individually for poly-G, quality, and adapter content).",
    "",
    "Source: 02_fastp_trim.sh and 04_salmon_quant.sh job logs.",
    "NOTE: numbers are transcribed from those logs, not re-parsed live -- if the",
    "pipeline is ever rerun with new/different data, regenerate this table from",
    "the new logs rather than trusting these hardcoded values.",
]
for i, line in enumerate(notes):
    ws.cell(row=5 + i, column=1, value=line).font = Font(name="Arial")
ws.column_dimensions["A"].width = 90

write_df_sheet(wb, "Overall summary", summary)
write_df_sheet(wb, "Per-sample stats", df)

xlsx_path = f"{TABLES_DIR}/sequencing_qc_summary.xlsx"
wb.save(xlsx_path)
print(f"Saved: {xlsx_path}")
print()
print(summary.to_string(index=False))
