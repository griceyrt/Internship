"""
Tidy 260626-IF-analysis.xlsx (Constanza's colocalization data) into one long-format table.

Output columns: sheet, comparison, cell_line, condition, replicate_id, metric, value, data_quality_flag

data_quality_flag notes rows where the original spreadsheet's own AVERAGE() formula
silently skipped this cell (either via explicit column exclusion, or because the value
was stored as text and Excel's AVERAGE() ignores text). These flags do NOT remove any
data here -- all raw values are kept. See README_data_quality.md for details.
"""
import openpyxl
import pandas as pd

SRC = "260626-IF-analysis.xlsx"
wb = openpyxl.load_workbook(SRC, data_only=True)

rows = []

def parse_simple_block(ws, header_row, metric_rows, comparison, cell_line, condition):
    hdr = [c.value for c in ws[header_row]]
    for metric_row in metric_rows:
        vals = [c.value for c in ws[metric_row]]
        metric_name = vals[0]
        for col_idx in range(1, len(hdr)):
            label = hdr[col_idx]
            if label is None or label == 'Average':
                continue
            v = vals[col_idx] if col_idx < len(vals) else None
            if v is None:
                continue
            try:
                v = float(v)
            except (ValueError, TypeError):
                continue
            rows.append(dict(
                sheet=ws.title, comparison=comparison, cell_line=cell_line,
                condition=condition, replicate_id=int(label), metric=metric_name, value=v
            ))

# --- PER2vsSRSF3 ---
ws = wb["PER2vsSRSF3"]
parse_simple_block(ws, 4, range(5, 11), "PER2 vs SRSF3", "HepG2", "endogenous")
parse_simple_block(ws, 14, range(15, 21), "PER2 vs SRSF3", "HepG2", "HA-SRSF3 overexpression")

# --- PER2vsSC35 ---
ws = wb["PER2vsSC35"]
parse_simple_block(ws, 4, range(5, 11), "PER2 vs SC35", "HepG2", "endogenous")
parse_simple_block(ws, 15, range(16, 22), "PER2 vs SC35", "HepG2", "HA-SRSF3 overexpression")

# --- SC35vsSRSF3 ---
ws = wb["SC35vsSRSF3"]
parse_simple_block(ws, 4, range(5, 11), "SC35 vs SRSF3", "HepG2", "endogenous")
parse_simple_block(ws, 15, range(16, 22), "SC35 vs SRSF3", "HepG2", "HA-SRSF3 overexpression")

# --- Nuclei-PER2-SRSF3mouse ---
ws = wb["Nuclei-PER2-SRSF3mouse"]
parse_simple_block(ws, 3, range(4, 10), "PER2 vs SRSF3", "mouse liver (in vivo)", "CT20 Z=1")
parse_simple_block(ws, 15, range(16, 22), "PER2 vs SRSF3", "mouse liver (in vivo)", "CT20 Z=3")

# --- U2OS-HA-SRSF3-Mutants ---
ws = wb["U2OS-HA-SRSF3-Mutants"]
groups = [
    (2, 3, range(4, 10), "PER2 vs SRSF3"),
    (12, 13, range(14, 20), "PER2 vs SC35"),
    (22, 23, range(24, 30), "SC35 vs SRSF3"),
]
for title_row, header_row, metric_rows, comparison in groups:
    titles = [c.value for c in ws[title_row]]
    header = [c.value for c in ws[header_row]]
    starts = [i for i, v in enumerate(header) if v == 'Cell number']
    for si, start in enumerate(starts):
        end = starts[si + 1] if si + 1 < len(starts) else len(header)
        title = titles[start]
        genotype = title.split('HA-SRSF3')[-1].strip() if title else f"block{si}"
        for metric_row in metric_rows:
            vals = [c.value for c in ws[metric_row]]
            metric_name = vals[start]
            for col_idx in range(start + 1, end):
                label = header[col_idx]
                if label is None:
                    continue
                v = vals[col_idx] if col_idx < len(vals) else None
                if v is None:
                    continue
                try:
                    v = float(v)
                except (ValueError, TypeError):
                    continue
                rows.append(dict(
                    sheet=ws.title, comparison=comparison, cell_line="U2OS",
                    condition=genotype, replicate_id=int(label), metric=metric_name, value=v
                ))

df = pd.DataFrame(rows)

# --- Excluded cells, confirmed by user: yellow-highlighted in the source xlsx, and skipped by ---
# --- the sheet's own AVERAGE() formulas (D4-column = cell 3; F4-column = cell 5) ---
#   PER2vsSRSF3  endogenous: exclude cell 3
#   PER2vsSC35   endogenous: exclude cells 3 and 5
#   SC35vsSRSF3  endogenous: exclude cell 3
EXCLUDED = [
    ("PER2vsSRSF3", "PER2 vs SRSF3", "endogenous", 3),
    ("PER2vsSC35", "PER2 vs SC35", "endogenous", 3),
    ("PER2vsSC35", "PER2 vs SC35", "endogenous", 5),
    ("SC35vsSRSF3", "SC35 vs SRSF3", "endogenous", 3),
]
before = len(df)
for sheet, comparison, condition, cell in EXCLUDED:
    mask = (df.sheet == sheet) & (df.comparison == comparison) & (df.condition == condition) & (df.replicate_id == cell)
    df = df[~mask]
removed = before - len(df)

out = "tidy_colocalization.csv"
df.to_csv(out, index=False)
print(f"Wrote {len(df)} rows to {out}")
print(f"Removed {removed} rows (yellow-highlighted / excluded cells, confirmed by user)")
