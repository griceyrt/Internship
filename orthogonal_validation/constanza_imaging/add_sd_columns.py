"""
Duplicate 260626-IF-analysis.xlsx and add a "SD" (standard deviation) column
immediately to the right of every "Average" column, for the requested tables:

  Tab1 PER2vsSRSF3            : both tables  (Average at AC4, H14)
  Tab2 PER2vsSC35             : both tables  (Average at AC4, H15)
  Tab3 SC35vsSRSF3            : both tables  (Average at AC4, H15)
  Tab4 Nuclei-PER2-SRSF3mouse : bottom table only (CT20 Z=3, Average at AA15)
  Tab5 U2OS-HA-SRSF3-Mutants  : untouched

Each SD formula mirrors the exact cell range of its neighboring AVERAGE formula
(same STDEV as AVERAGE, function name swapped) so it respects the same excluded
cells (e.g. cell 3 skipped in the HepG2-endogenous tables).
"""
import shutil
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import copy

SRC = "260626-IF-analysis.xlsx"
OUT = "260626-IF-analysis_with_SD.xlsx"

shutil.copy(SRC, OUT)
wb = load_workbook(OUT, data_only=False)

# (sheet, header_row, avg_col, new_col, metric_rows)
tables = [
    ("PER2vsSRSF3", 4, 29, 30, range(5, 11)),    # AC4 -> AD4
    ("PER2vsSRSF3", 14, 8, 9, range(15, 21)),    # H14 -> I14
    ("PER2vsSC35", 4, 29, 30, range(5, 11)),     # AC4 -> AD4
    ("PER2vsSC35", 15, 8, 9, range(16, 22)),     # H15 -> I15
    ("SC35vsSRSF3", 4, 29, 30, range(5, 11)),    # AC4 -> AD4
    ("SC35vsSRSF3", 15, 8, 9, range(16, 22)),    # H15 -> I15
    ("Nuclei-PER2-SRSF3mouse", 15, 27, 28, range(16, 22)),  # AA15 -> AB15 (bottom table / Z=3 only)
]

for sheet, header_row, avg_col, new_col, metric_rows in tables:
    ws = wb[sheet]
    avg_header_cell = ws.cell(row=header_row, column=avg_col)
    assert avg_header_cell.value == "Average", f"{sheet} row{header_row} col{avg_col} = {avg_header_cell.value!r}, expected 'Average'"

    header_cell = ws.cell(row=header_row, column=new_col)
    header_cell.value = "SD"
    header_cell.font = copy.copy(avg_header_cell.font)
    header_cell.fill = copy.copy(avg_header_cell.fill)
    header_cell.border = copy.copy(avg_header_cell.border)
    header_cell.alignment = copy.copy(avg_header_cell.alignment)

    avg_letter = get_column_letter(avg_col)
    new_letter = get_column_letter(new_col)
    if avg_letter in ws.column_dimensions:
        ws.column_dimensions[new_letter].width = ws.column_dimensions[avg_letter].width

    for row in metric_rows:
        avg_cell = ws.cell(row=row, column=avg_col)
        formula = avg_cell.value
        assert isinstance(formula, str) and formula.upper().startswith("=AVERAGE"), \
            f"{sheet} row{row} col{avg_col} = {formula!r}, expected an AVERAGE formula"
        sd_formula = formula.replace("AVERAGE", "STDEV", 1)

        new_cell = ws.cell(row=row, column=new_col)
        new_cell.value = sd_formula
        new_cell.number_format = avg_cell.number_format
        new_cell.font = copy.copy(avg_cell.font)

    print(f"{sheet}: added SD column {new_letter} (from Average col {avg_letter}), rows {list(metric_rows)}")

wb.save(OUT)
print(f"\nSaved {OUT}")
