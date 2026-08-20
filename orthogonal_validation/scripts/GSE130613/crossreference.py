#!/usr/bin/env python3
"""
Orthogonal Validation — Steps 12, 14, 15
Author: Gricey

Step 12: dPSI scatter plot — Illumina vs Nanopore splicing agreement
Step 14: Excel tables — DE vs splicing overlap (Illumina only)
Step 15: DE overlap table — Illumina vs Nanopore (Khushi's Nanopore DE results)
         Source: data/Table3-differential_expressiong_WTvsPerKO.xlsx (Khushi)
         Compares significant DE genes (padj < 0.05) across both platforms.
         Output: tables/step15_DE_overlap_Illumina_vs_Nanopore.xlsx

Significance thresholds:
  - Illumina splicing : |dPSI| > 0.1, p < 0.3
  - Nanopore splicing : |dPSI| > 0.1, p < 0.05
  - DE genes          : padj_gene_stageR < 0.05 (stageR stage 1, Illumina)
  - Nanopore DE       : padj_per_ko < 0.05 (Khushi's results, sheet perKO_sig)

Run from: orthogonal_validation/
"""

import os
import re
import glob
import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from adjustText import adjust_text
from matplotlib.lines import Line2D

# STYLE — DESeq2-inspired clean aesthetic
# White background, open axes (no top/right spines), subtle grid behind data.
matplotlib.rcParams.update({
    "font.family":          "sans-serif",
    "font.sans-serif":      ["Arial", "DejaVu Sans", "Helvetica", "Verdana"],
    "font.size":            11,
    "axes.spines.top":      False,
    "axes.spines.right":    False,
    "axes.spines.left":     True,
    "axes.spines.bottom":   True,
    "axes.edgecolor":       "#AAAAAA",
    "axes.linewidth":       0.8,
    "axes.grid":            True,
    "grid.color":           "#E5E5E5",
    "grid.linewidth":       0.6,
    "grid.linestyle":       "--",
    "axes.axisbelow":       True,
    "figure.facecolor":     "white",
    "axes.facecolor":       "white",
    "xtick.color":          "#444444",
    "ytick.color":          "#444444",
    "xtick.labelsize":      10,
    "ytick.labelsize":      10,
    "axes.labelcolor":      "#222222",
    "axes.titlepad":        10,
    "legend.frameon":       True,
    "legend.framealpha":    0.9,
    "legend.edgecolor":     "#DDDDDD",
})

def autofit_excel(path):
    """Auto-fit all column widths in an Excel file to their content."""
    wb = openpyxl.load_workbook(path)
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 3
    wb.save(path)

BASE_DIR      = "/Users/gricey/Desktop/Internship/orthogonal_validation"
SUPPA_DATE    = "2026-06-20"
ILLUMINA_DIR  = os.path.join(BASE_DIR, "results", "GSE130613", f"suppa_{SUPPA_DATE}", "diff")
NANOPORE_DIR  = os.path.join(BASE_DIR, "data", "nanopore_suppa")
OUT_DIR       = os.path.join(BASE_DIR, "results", "GSE130613", "figures")
TABLES_DIR    = os.path.join(BASE_DIR, "results", "GSE130613", "tables")
DESEQ_FILE    = os.path.join(BASE_DIR, "results", "GSE130613", "normalisation", "deseq2_KO_vs_WT.tsv")
GTF_PATH      = "/Users/gricey/Desktop/Internship/boundary_analysis/data/transcriptome_productivity.gtf"
os.makedirs(OUT_DIR, exist_ok=True)
os.makedirs(TABLES_DIR, exist_ok=True)

# THRESHOLDS
DPSI_THRESH   = 0.1
PVAL_ILLUMINA = 0.3
PVAL_NANOPORE = 0.05
PADJ_GENE     = 0.05   # stageR stage 1 threshold

# HELPER FUNCTIONS

def build_gene_map(gtf_path):
    """Extract gene_id -> gene_name from GTF."""
    gene_map = {}
    with open(gtf_path) as fh:
        for line in fh:
            if "\tgene\t" not in line:
                continue
            gid   = re.search(r'gene_id "([^"]+)"', line)
            gname = re.search(r'gene_name "([^"]+)"', line)
            if gid and gname:
                gene_map[gid.group(1)] = gname.group(1)
    return gene_map


def load_all_events(dpsi_dir, file_pattern):
    """Load all events from dpsi files.
    Returns dict: event_id -> {dPSI, pval, event_type}
    Sign: as-is from file (WT - KO).
    """
    all_rows = {}
    for f in sorted(glob.glob(os.path.join(dpsi_dir, file_pattern))):
        event_type = os.path.basename(f).split("_")[1] if "_" in os.path.basename(f) else "?"
        df = pd.read_csv(f, sep="\t", index_col=0)
        df.columns = ["dPSI", "pval"]
        df = df.dropna(subset=["dPSI", "pval"])
        df = df[df["pval"] > 0]
        for idx, row in df.iterrows():
            etype = idx.split(";")[1].split(":")[0] if ";" in idx else event_type
            all_rows[idx] = {"dPSI": row["dPSI"], "pval": row["pval"], "event_type": etype}
    return all_rows


def is_significant(row, dpsi_thresh, pval_thresh):
    return abs(row["dPSI"]) > dpsi_thresh and row["pval"] < pval_thresh


# LOAD DATA
print("Building gene name map from GTF...")
gene_map = build_gene_map(GTF_PATH)
print(f"  {len(gene_map)} genes loaded.")

print("Loading Illumina events...")
illumina_all = load_all_events(ILLUMINA_DIR, "diff_*_strict.dpsi.temp.0")
illumina_sig = {k for k, v in illumina_all.items()
                if is_significant(v, DPSI_THRESH, PVAL_ILLUMINA)}

print("Loading Nanopore events...")
nanopore_all = load_all_events(NANOPORE_DIR, "res_*.dpsi.temp.0")
nanopore_sig = {k for k, v in nanopore_all.items()
                if is_significant(v, DPSI_THRESH, PVAL_NANOPORE)}

# Events present in BOTH datasets (exact event ID match)
common_events = set(illumina_all.keys()) & set(nanopore_all.keys())
overlap_sig   = illumina_sig & nanopore_sig   # significant in both

print(f"\n{'='*50}")
print(f"Illumina significant events (p<{PVAL_ILLUMINA}) : {len(illumina_sig)}")
print(f"Nanopore significant events (p<{PVAL_NANOPORE}) : {len(nanopore_sig)}")
print(f"Events in common (both datasets)                 : {len(common_events)}")
print(f"Significant in both                              : {len(overlap_sig)}")
print(f"Illumina only                                    : {len(illumina_sig - nanopore_sig)}")
print(f"Nanopore only                                    : {len(nanopore_sig - illumina_sig)}")
print(f"{'='*50}\n")

# STEP 12 — Excel: overlapping significant events
if len(overlap_sig) > 0:
    rows = []
    for event_id in sorted(overlap_sig):
        gene_id    = event_id.split(";")[0]
        event_type = event_id.split(";")[1].split(":")[0]
        rows.append({
            "gene_id":       gene_id,
            "gene_name":     gene_map.get(gene_id, "unannotated"),
            "event_type":    event_type,
            "event_id":      event_id,
            "dPSI_illumina": round(-illumina_all[event_id]["dPSI"], 4),  # negate → KO-WT
            "pval_illumina": round( illumina_all[event_id]["pval"], 4),
            "dPSI_nanopore": round(-nanopore_all[event_id]["dPSI"], 4),  # negate → KO-WT
            "pval_nanopore": round( nanopore_all[event_id]["pval"], 4),
        })
    overlap_df = pd.DataFrame(rows)
    out = os.path.join(TABLES_DIR, "step12_overlap_events.xlsx")
    overlap_df.to_excel(out, index=False)
    autofit_excel(out)
    print(f"Step 12 Excel saved: {out}")

# STEP 12 — Scatter plot: Illumina dPSI vs Nanopore dPSI
# Replaces Venn diagram — shows agreement vs disagreement between platforms
# All events present in both datasets are plotted
fig, ax = plt.subplots(figsize=(10, 7))

# No gridlines — paper style
ax.grid(False)

# Build arrays for common events
x_vals, y_vals, colors, sizes, event_ids = [], [], [], [], []

for event_id in common_events:
    dpsi_ill = -illumina_all[event_id]["dPSI"]
    dpsi_nan = -nanopore_all[event_id]["dPSI"]
    sig_ill  = event_id in illumina_sig
    sig_nan  = event_id in nanopore_sig
    x_vals.append(dpsi_ill)
    y_vals.append(dpsi_nan)
    event_ids.append(event_id)
    if sig_ill and sig_nan:
        colors.append("#E69F00"); sizes.append(60)
    elif sig_ill:
        colors.append("#56B4E9"); sizes.append(30)
    elif sig_nan:
        colors.append("#009E73"); sizes.append(30)
    else:
        colors.append("lightgrey"); sizes.append(10)

x_arr = np.array(x_vals)
y_arr = np.array(y_vals)

# Background: tiny light grey dots
grey_mask = np.array([c == "lightgrey" for c in colors])
ax.scatter(x_arr[grey_mask], y_arr[grey_mask],
           c="#CCCCCC", s=6, alpha=0.35, linewidths=0, zorder=1)

# Blue and green significant dots
for c, z in [("#56B4E9", 2), ("#009E73", 3)]:
    mask = np.array([col == c for col in colors])
    ax.scatter(x_arr[mask], y_arr[mask],
               c=c, s=40, alpha=0.8, linewidths=0, zorder=z)

# Orange (both significant) — black outline
orange_mask = np.array([c == "#E69F00" for c in colors])
ax.scatter(x_arr[orange_mask], y_arr[orange_mask],
           c="#E69F00", s=55, alpha=0.95,
           edgecolors="black", linewidths=0.7, zorder=4)

# Reference lines
ax.axhline(0, color="#BBBBBB", linewidth=0.6, alpha=0.5)
ax.axvline(0, color="#BBBBBB", linewidth=0.6, alpha=0.5)
ax.plot([-1, 1], [-1, 1], color="#888888", linewidth=0.9,
        linestyle="--", alpha=0.6)

# Labels — orange events only, one per gene, |dPSI| > 0.2 on at least one axis
gene_best = {}
for i, event_id in enumerate(event_ids):
    if event_id not in overlap_sig:
        continue
    if abs(x_arr[i]) <= 0.2 and abs(y_arr[i]) <= 0.2:
        continue
    gene_id   = event_id.split(";")[0]
    gene_name = gene_map.get(gene_id, gene_id)
    if gene_name not in gene_best or abs(x_arr[i]) > abs(gene_best[gene_name][0]):
        gene_best[gene_name] = (x_arr[i], y_arr[i])

# Seed, adjust, then redraw as annotate (lines touch dot and label)
texts      = []
gene_order = []
for gene_name, (dx, dy) in gene_best.items():
    offset_x = 0.06 if dx >= 0 else -0.06
    offset_y = 0.04 if dy >= 0 else -0.04
    texts.append(ax.text(dx + offset_x, dy + offset_y, gene_name,
                         fontsize=8, color="#1a1a1a", fontstyle="italic"))
    gene_order.append((gene_name, dx, dy))

adjust_text(texts, ax=ax,
            expand=(3.5, 3.5), force_text=(2.5, 2.5),
            force_points=(1.5, 1.5), lim=700)

for txt, (gene_name, dot_x, dot_y) in zip(texts, gene_order):
    lx, ly = txt.get_position()
    txt.remove()
    ax.annotate(gene_name,
                xy=(dot_x, dot_y), xytext=(lx, ly),
                fontsize=8, color="#1a1a1a", fontstyle="italic",
                ha="center", va="center",
                arrowprops=dict(arrowstyle="-", color="#888888", lw=0.7,
                                shrinkA=4, shrinkB=4))

# Legend
from matplotlib.lines import Line2D
legend_elements = [
    Line2D([0],[0], marker='o', color='w', markerfacecolor='#E69F00',
           markeredgecolor='black', markeredgewidth=0.7,
           markersize=8, label=f'Both significant (n={len(overlap_sig)})'),
    Line2D([0],[0], marker='o', color='w', markerfacecolor='#56B4E9',
           markersize=7, label=f'Illumina only (n={len(illumina_sig - nanopore_sig)})'),
    Line2D([0],[0], marker='o', color='w', markerfacecolor='#009E73',
           markersize=7, label=f'Nanopore only (n={len(nanopore_sig - illumina_sig)})'),
    Line2D([0],[0], marker='o', color='w', markerfacecolor='#CCCCCC',
           markersize=6, label=f'Not significant (n={len(common_events) - len(illumina_sig | nanopore_sig)})'),
]
ax.legend(handles=legend_elements, fontsize=8,
          bbox_to_anchor=(1.02, 1), loc="upper left")

ax.set_xlabel("dPSI Illumina (PerDKO − WT)", fontsize=12)
ax.set_ylabel("dPSI Nanopore (PerDKO − WT)", fontsize=12)
ax.set_title(
    "Differentially spliced events — Illumina vs Nanopore (PerDKO vs WT)\n"
    f"All events tested in both platforms (n={len(common_events)}) — orange = significant in both",
    fontsize=11
)
plt.tight_layout()
out_path = os.path.join(OUT_DIR, "step12_dPSI_scatter.png")
plt.savefig(out_path, dpi=150, bbox_inches="tight")
plt.close()
print(f"Scatter plot saved: {out_path}")

# EXCEL TABLES (Step 14) — DE vs Splicing
de = pd.read_csv(DESEQ_FILE, sep="\t")
de = de.dropna(subset=["log2FoldChange", "padj_gene_stageR"])
de_gene = (de.sort_values("log2FoldChange", key=abs, ascending=False)
             .drop_duplicates(subset="gene_id")
             .set_index("gene_id"))
de_sig_genes = set(de_gene[de_gene["padj_gene_stageR"] < PADJ_GENE].index)

illumina_sig_gene_events = {}
for event_id, row in illumina_all.items():
    if not is_significant(row, DPSI_THRESH, PVAL_ILLUMINA):
        continue
    gene_id = event_id.split(";")[0]
    illumina_sig_gene_events.setdefault(gene_id, []).append({
        "event_id":   event_id,
        "event_type": event_id.split(";")[1].split(":")[0],
        "dPSI":       round(-row["dPSI"], 4),
        "pval":       round(row["pval"], 4),
    })
illumina_sig_genes = set(illumina_sig_gene_events.keys())
de_splice_overlap  = de_sig_genes & illumina_sig_genes
de_only            = de_sig_genes - illumina_sig_genes
splice_only        = illumina_sig_genes - de_sig_genes

print(f"Significant DE genes (stageR padj<{PADJ_GENE}) : {len(de_sig_genes)}")
print(f"Significant spliced genes                       : {len(illumina_sig_genes)}")
print(f"Both DE and spliced                             : {len(de_splice_overlap)}")
print(f"DE only                                         : {len(de_only)}")
print(f"Spliced only                                    : {len(splice_only)}")

de_full = de_gene.copy()

def make_splice_rows(gene_set):
    rows = []
    for gene_id in sorted(gene_set):
        for ev in illumina_sig_gene_events.get(gene_id, []):
            row = {
                "gene_id":       gene_id,
                "gene_name":     gene_map.get(gene_id, "unannotated"),
                "event_type":    ev["event_type"],
                "event_id":      ev["event_id"],
                "dPSI_illumina": ev["dPSI"],
                "pval_illumina": ev["pval"],
            }
            if gene_id in de_full.index:
                row["log2FoldChange"]    = round(de_full.loc[gene_id, "log2FoldChange"], 4)
                row["padj_gene_stageR"] = round(de_full.loc[gene_id, "padj_gene_stageR"], 6)
            rows.append(row)
    return rows


def make_de_only_rows(gene_set):
    rows = []
    for gene_id in sorted(gene_set):
        if gene_id not in de_full.index:
            continue
        rows.append({
            "gene_id":           gene_id,
            "gene_name":         gene_map.get(gene_id, "unannotated"),
            "log2FoldChange":    round(de_full.loc[gene_id, "log2FoldChange"], 4),
            "padj_gene_stageR": round(de_full.loc[gene_id, "padj_gene_stageR"], 6),
        })
    return rows


# DE only
de_only_df = pd.DataFrame(make_de_only_rows(de_only))
if len(de_only_df) > 0:
    de_only_df = pd.concat([
        de_only_df[de_only_df["gene_name"] != "unannotated"].sort_values("gene_name"),
        de_only_df[de_only_df["gene_name"] == "unannotated"].sort_values("gene_id")
    ]).reset_index(drop=True)
p = os.path.join(TABLES_DIR, "step14_DE_only.xlsx")
de_only_df.to_excel(p, index=False)
autofit_excel(p)
print(f"Saved: step14_DE_only.xlsx  (n={len(de_only)})")

# Both DE and spliced
p = os.path.join(TABLES_DIR, "step14_DE_and_spliced.xlsx")
pd.DataFrame(make_splice_rows(de_splice_overlap)).to_excel(p, index=False)
autofit_excel(p)
print(f"Saved: step14_DE_and_spliced.xlsx  (n={len(de_splice_overlap)})")

# Spliced only
p = os.path.join(TABLES_DIR, "step14_spliced_only.xlsx")
pd.DataFrame(make_splice_rows(splice_only)).to_excel(p, index=False)
autofit_excel(p)
print(f"Saved: step14_spliced_only.xlsx  (n={len(splice_only)})")

# STEP 14b — DE concordance scatter: Illumina vs Nanopore at GENE level. One
# dot per gene (transcript with highest |log2FC| represents the gene).
# Symmetric thresholds: padj<0.05 AND |log2FC|>1.5.
print("\n=== Step 14b: DE concordance scatter — Illumina vs Nanopore (gene level) ===")

KHUSHI_FILE_14 = os.path.join(BASE_DIR, "data",
                               "Table3-differential_expressiong_WTvsPerKO.xlsx")

if not os.path.exists(KHUSHI_FILE_14):
    print(f"  Skipping — Khushi's file not found at: {KHUSHI_FILE_14}")
else:
    wb14 = openpyxl.load_workbook(KHUSHI_FILE_14)

    # perKO_sig: significant genes (padj<0.05, |log2FC|>1.5, pre-filtered by Khushi)
    ws_sig14 = wb14["perKO_sig"]
    rows_sig14 = list(ws_sig14.iter_rows(values_only=True))
    header_sig14 = [str(c) if c is not None else "idx" for c in rows_sig14[0]]
    nano_sig_geneid = {}   # gene_id -> log2FC (first occurrence = best representative)
    for r in rows_sig14[1:]:
        row = dict(zip(header_sig14, r))
        gid = row.get("gene_id")
        if gid and gid not in nano_sig_geneid:
            try:
                nano_sig_geneid[gid] = float(row["log2FC_perko"])
            except (TypeError, ValueError):
                pass

    # DE_all: all genes tested (background for scatter)
    ws_all14 = wb14["DE_all"]
    rows_all14 = list(ws_all14.iter_rows(values_only=True))
    header_all14 = list(rows_all14[0])
    nano_all_geneid = {}   # gene_id -> log2FC (representative: highest |log2FC|)
    for r in rows_all14[1:]:
        row = dict(zip(header_all14, r))
        gid = row.get("gene_id")
        if not gid:
            continue
        try:
            lfc  = float(row["log2FC_perko"])
            padj = float(row["padj_per_ko"])
        except (TypeError, ValueError):
            continue
        if gid not in nano_all_geneid or abs(lfc) > abs(nano_all_geneid[gid]["log2FC"]):
            nano_all_geneid[gid] = {"log2FC": lfc, "padj": padj}

    # Our Illumina: one row per gene_id (transcript with highest |log2FC|)
    de14 = pd.read_csv(DESEQ_FILE, sep="\t").dropna(subset=["log2FoldChange", "padj_gene_stageR"])
    de14["abs_lfc"] = de14["log2FoldChange"].abs()
    de14_gene = de14.sort_values("abs_lfc", ascending=False).drop_duplicates("gene_id").set_index("gene_id")

    # Symmetric significance: padj<0.05 AND |log2FC|>1.5 on Illumina side
    ill_sig14 = set(de14_gene[
        (de14_gene["padj_gene_stageR"] < PADJ_GENE) &
        (de14_gene["abs_lfc"] > 1.5)
    ].index)

    common14   = set(de14_gene.index) & set(nano_all_geneid.keys())
    nano_sig14 = set(nano_sig_geneid.keys())
    overlap14  = ill_sig14 & nano_sig14 & common14

    print(f"  Common genes                          : {len(common14)}")
    print(f"  Illumina sig (padj<0.05, |log2FC|>1.5): {len(ill_sig14)}")
    print(f"  Nanopore sig (perKO_sig)              : {len(nano_sig14)}")
    print(f"  Overlap (both significant)            : {len(overlap14)}")

    # Build plot arrays
    genes14 = list(common14)
    x14 = np.array([float(de14_gene.loc[g, "log2FoldChange"]) for g in genes14])
    y14 = np.array([float(nano_all_geneid[g]["log2FC"]) for g in genes14])

    mask14_both = np.array([g in overlap14 for g in genes14])
    mask14_ill  = np.array([g in ill_sig14 and g not in overlap14 for g in genes14])
    mask14_nano = np.array([g in nano_sig14 and g not in overlap14 for g in genes14])
    mask14_grey = ~(mask14_both | mask14_ill | mask14_nano)

    fig, ax = plt.subplots(figsize=(13, 7))
    ax.grid(False)

    ax.scatter(x14[mask14_grey], y14[mask14_grey], c="#CCCCCC", s=6, alpha=0.35,
               linewidths=0, zorder=1, rasterized=True)
    ax.scatter(x14[mask14_ill],  y14[mask14_ill],  c="#56B4E9", s=40, alpha=0.85,
               linewidths=0, zorder=2)
    ax.scatter(x14[mask14_nano], y14[mask14_nano], c="#009E73", s=40, alpha=0.85,
               linewidths=0, zorder=3)
    ax.scatter(x14[mask14_both], y14[mask14_both], c="#E69F00", s=55, alpha=0.95,
               edgecolors="black", linewidths=0.7, zorder=4)

    lim14 = max(abs(x14).max(), abs(y14).max()) * 1.05
    ax.plot([-lim14, lim14], [-lim14, lim14], color="#888888", lw=0.9, ls="--", alpha=0.6)
    ax.axhline(0, color="#BBBBBB", lw=0.6, alpha=0.5)
    ax.axvline(0, color="#BBBBBB", lw=0.6, alpha=0.5)

    # Labels: orange dots — adjust_text to avoid overlaps, then redraw as
    # ax.annotate so lines reliably connect dot to label (same pattern as step 12)
    texts14     = []
    gene_order14 = []
    for i, g in enumerate(genes14):
        if not mask14_both[i]:
            continue
        gname = de14_gene.loc[g, "gene_name"] if "gene_name" in de14_gene.columns else gene_map.get(g, g)
        offset_x = 1.2 if x14[i] >= 0 else -1.2
        offset_y = 0.8 if y14[i] >= 0 else -0.8
        texts14.append(ax.text(x14[i] + offset_x, y14[i] + offset_y, gname,
                               fontsize=8, color="#1a1a1a", fontstyle="italic"))
        gene_order14.append((gname, x14[i], y14[i]))

    adjust_text(texts14, ax=ax,
                expand=(3.5, 3.5), force_text=(2.5, 2.5),
                force_points=(1.5, 1.5), lim=700)

    for txt, (gname, dot_x, dot_y) in zip(texts14, gene_order14):
        lx, ly = txt.get_position()
        txt.remove()
        ax.annotate(gname,
                    xy=(dot_x, dot_y), xytext=(lx, ly),
                    fontsize=8, color="#1a1a1a", fontstyle="italic",
                    ha="center", va="center",
                    arrowprops=dict(arrowstyle="-", color="#888888", lw=0.7,
                                    shrinkA=4, shrinkB=4))

    legend14 = [
        Line2D([0],[0], marker="o", color="w", markerfacecolor="#E69F00",
               markeredgecolor="black", markersize=8,
               label=f"Both significant (n={mask14_both.sum()})"),
        Line2D([0],[0], marker="o", color="w", markerfacecolor="#56B4E9",
               markersize=7, label=f"Illumina only (n={mask14_ill.sum()})"),
        Line2D([0],[0], marker="o", color="w", markerfacecolor="#009E73",
               markersize=7, label=f"Nanopore only (n={mask14_nano.sum()})"),
        Line2D([0],[0], marker="o", color="w", markerfacecolor="#CCCCCC",
               markersize=6, label=f"Not significant (n={mask14_grey.sum()})"),
    ]
    ax.legend(handles=legend14, fontsize=8, bbox_to_anchor=(1.02, 1), loc="upper left")
    ax.set_xlabel("log₂ fold change — Illumina (PerDKO / WT)", fontsize=12)
    ax.set_ylabel("log₂ fold change — Nanopore (PerDKO / WT)", fontsize=12)
    ax.set_title(
        "DE concordance: Illumina vs Nanopore (PerDKO vs WT) — gene level\n"
        f"All genes tested in both platforms (n={len(common14)}) | threshold: padj<0.05, |log₂FC|>1.5",
        fontsize=11)
    plt.tight_layout()
    p14 = os.path.join(OUT_DIR, "step14_DE_concordance_gene_level.png")
    plt.savefig(p14, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"  Saved: step14_DE_concordance_gene_level.png")


# STEP 15 — DE concordance: Illumina vs Nanopore at TRANSCRIPT level. Matches
# on transcript_id (both annotated ENSMUST and novel UUID, same custom
# transcriptome_ext.fa). Thresholds: Illumina padj_gene_stageR<0.05 &
# |log2FoldChange|>1.5; Nanopore padj_per_ko<0.05 & |log2FC_perko|>1.5.
print("\n=== Step 15: DE concordance — Illumina vs Nanopore (transcript level) ===")

LFC_THRESH_15 = 1.5   # matching Khushi's manuscript threshold

KHUSHI_FILE = os.path.join(BASE_DIR, "data",
                           "Table3-differential_expressiong_WTvsPerKO.xlsx")

if not os.path.exists(KHUSHI_FILE):
    print(f"  Skipping — Khushi's file not found at: {KHUSHI_FILE}")
else:
    wb = openpyxl.load_workbook(KHUSHI_FILE)

    # --- Load Khushi's DE_all (all transcripts, background) ---
    # Include ALL transcript IDs: both ENSMUST (annotated) and UUID (novel).
    # Both our Illumina data and Khushi's Nanopore data used the same custom
    # transcriptome (transcriptome_ext.fa), so UUID IDs are shared.
    ws_all     = wb["DE_all"]
    rows_all   = list(ws_all.iter_rows(values_only=True))
    header_all = list(rows_all[0])
    khushi_all_tx = {}   # transcript_id -> {log2FC, padj, gene_name}
    for r in rows_all[1:]:
        row = dict(zip(header_all, r))
        txid = str(row.get("transcript_id", "")).strip()
        if not txid or txid == "None":
            continue
        try:
            lfc  = float(row["log2FC_perko"])
            padj = float(row["padj_per_ko"])
        except (TypeError, ValueError):
            continue
        khushi_all_tx[txid] = {
            "log2FC": lfc,
            "padj":   padj,
            "gene_name": str(row.get("mgi_symbol", "")) or "unannotated",
        }

    # --- Load Khushi's perKO_sig (significant transcripts) ---
    # Sheet is pre-filtered: padj < 0.05 AND |log2FC| > 1.5
    # Include all IDs (annotated + novel UUID).
    ws_sig    = wb["perKO_sig"]
    rows_sig  = list(ws_sig.iter_rows(values_only=True))
    header_sig = [str(c) if c is not None else "idx" for c in rows_sig[0]]
    khushi_sig_tx = set()   # transcript_ids that are significant in Nanopore
    for r in rows_sig[1:]:
        row = dict(zip(header_sig, r))
        txid = str(row.get("transcript_id", "")).strip()
        if txid and txid != "None":
            khushi_sig_tx.add(txid)

    # --- Load our Illumina data at transcript level ---
    our_tx = pd.read_csv(DESEQ_FILE, sep="\t")
    our_tx = our_tx.dropna(subset=["log2FoldChange", "padj_gene_stageR"])
    our_tx = our_tx.set_index("transcript_id")

    # Significance: padj_gene_stageR < 0.05 AND |log2FC| > 1.5 (matching Khushi)
    our_sig_tx = set(our_tx[
        (our_tx["padj_gene_stageR"] < PADJ_GENE) &
        (our_tx["log2FoldChange"].abs() > LFC_THRESH_15)
    ].index)

    # tx_to_name: transcript_id -> gene_name (from our data)
    tx_to_name = our_tx["gene_name"].to_dict()

    # --- Common transcripts (ENSEMBL IDs in both datasets) ---
    common_tx = set(our_tx.index) & set(khushi_all_tx.keys())
    nano_sig_tx_common = khushi_sig_tx & common_tx

    print(f"  Khushi transcripts in DE_all (annot+novel)        : {len(khushi_all_tx)}")
    print(f"  Khushi significant (padj<0.05, |log2FC|>1.5)     : {len(khushi_sig_tx)} (annot+novel)")
    print(f"  Our Illumina transcripts                          : {len(our_tx)}")
    print(f"  Our significant (padj<0.05, |log2FC|>1.5)        : {len(our_sig_tx)}")
    print(f"  Common transcripts (in both datasets)             : {len(common_tx)}")
    overlap_tx = our_sig_tx & khushi_sig_tx & common_tx
    print(f"  Overlap (sig in both, annot+novel)                : {len(overlap_tx)}")

    # --- Excel overlap table ---
    rows_out = []
    for txid in sorted(overlap_tx):
        ill  = our_tx.loc[txid]
        nano = khushi_all_tx[txid]
        lfc_ill = round(float(ill["log2FoldChange"]), 4)
        lfc_nan = round(nano["log2FC"], 4)
        dir_ill = "UP" if lfc_ill > 0 else "DOWN"
        dir_nan = "UP" if lfc_nan > 0 else "DOWN"
        rows_out.append({
            "transcript_id":        txid,
            "gene_id":              ill.get("gene_id", ""),
            "gene_name":            tx_to_name.get(txid, "unannotated"),
            "log2FC_illumina":      lfc_ill,
            "padj_illumina_stageR": round(float(ill["padj_gene_stageR"]), 6),
            "direction_illumina":   dir_ill,
            "log2FC_nanopore":      lfc_nan,
            "padj_nanopore":        round(nano["padj"], 6),
            "direction_nanopore":   dir_nan,
            "consistent_direction": dir_ill == dir_nan,
        })

    overlap_df = pd.DataFrame(rows_out).sort_values("gene_name")
    out_path   = os.path.join(TABLES_DIR, "step15_DE_overlap_Illumina_vs_Nanopore.xlsx")
    overlap_df.to_excel(out_path, index=False)
    autofit_excel(out_path)
    n_consistent = overlap_df["consistent_direction"].sum() if len(overlap_df) > 0 else 0
    print(f"  Consistent direction : {n_consistent} / {len(overlap_df)}")
    print(f"  Saved: step15_DE_overlap_Illumina_vs_Nanopore.xlsx")

    # --- Build scatter arrays ---
    x_vals, y_vals, colors, sizes, tx_ids = [], [], [], [], []
    for txid in common_tx:
        x = float(our_tx.loc[txid, "log2FoldChange"])
        y = khushi_all_tx[txid]["log2FC"]
        sig_ill  = txid in our_sig_tx
        sig_nano = txid in khushi_sig_tx
        x_vals.append(x);  y_vals.append(y);  tx_ids.append(txid)
        if sig_ill and sig_nano:
            colors.append("#E69F00"); sizes.append(60)
        elif sig_ill:
            colors.append("#56B4E9"); sizes.append(30)
        elif sig_nano:
            colors.append("#009E73"); sizes.append(30)
        else:
            colors.append("lightgrey"); sizes.append(10)

    x_arr = np.array(x_vals)
    y_arr = np.array(y_vals)
    both_sig_mask = np.array([c == "#E69F00" for c in colors])

    # Counts for legend
    both_n      = int(both_sig_mask.sum())
    ill_only_n  = sum(1 for c in colors if c == "#56B4E9")
    nano_only_n = sum(1 for c in colors if c == "#009E73")
    grey_n      = sum(1 for c in colors if c == "lightgrey")

    legend_elements = [
        Line2D([0],[0], marker="o", color="w", markerfacecolor="#E69F00",
               markersize=8, label=f"Both significant (n={both_n})"),
        Line2D([0],[0], marker="o", color="w", markerfacecolor="#56B4E9",
               markersize=7, label=f"Illumina only (n={ill_only_n})"),
        Line2D([0],[0], marker="o", color="w", markerfacecolor="#009E73",
               markersize=7, label=f"Nanopore only (n={nano_only_n})"),
        Line2D([0],[0], marker="o", color="w", markerfacecolor="lightgrey",
               markersize=6, label=f"Not significant (n={grey_n})"),
    ]

    def draw_de_scatter(ax, x_arr, y_arr, colors, sizes, gids,
                        both_sig_mask, clip=None):
        """Draw the DE concordance scatter on ax.
        Style: Bharath-paper inspired — no gridlines, tiny grey background dots,
        colored significant dots, clean diagonal, labels directly on plot.
        clip: if a number, clip both axes to [-clip, clip] and show
              extreme outliers as triangles at the boundary.
        """
        # Override global gridlines for this figure — paper style has none
        ax.grid(False)

        if clip:
            x_plot = np.clip(x_arr, -clip, clip)
            y_plot = np.clip(y_arr, -clip, clip)
            clipped = (np.abs(x_arr) > clip) | (np.abs(y_arr) > clip)
        else:
            x_plot, y_plot = x_arr, y_arr
            clipped = np.zeros(len(x_arr), dtype=bool)

        # Background (not significant): very small, light grey, low alpha
        grey_mask = np.array([c == "lightgrey" for c in colors]) & ~clipped
        ax.scatter(x_plot[grey_mask], y_plot[grey_mask],
                   c="#CCCCCC", s=6, alpha=0.35, linewidths=0, zorder=1)

        # Blue and green significant dots (no outline)
        for c, z in [("#56B4E9", 2), ("#009E73", 3)]:
            mask = np.array([col == c for col in colors]) & ~clipped
            ax.scatter(x_plot[mask], y_plot[mask],
                       c=c, s=45, alpha=0.85, linewidths=0, zorder=z)
            if clip:
                mask_clip = np.array([col == c for col in colors]) & clipped
                if mask_clip.any():
                    ax.scatter(x_plot[mask_clip], y_plot[mask_clip],
                               c=c, s=55, alpha=0.85, marker="^",
                               linewidths=0, zorder=z + 1)

        # Orange dots: black outline to match figure 13 labeled-gene style
        orange_mask = np.array([c == "#E69F00" for c in colors]) & ~clipped
        ax.scatter(x_plot[orange_mask], y_plot[orange_mask],
                   c="#E69F00", s=55, alpha=0.95,
                   edgecolors="black", linewidths=0.7, zorder=4)
        if clip:
            orange_clip = np.array([c == "#E69F00" for c in colors]) & clipped
            if orange_clip.any():
                ax.scatter(x_plot[orange_clip], y_plot[orange_clip],
                           c="#E69F00", s=65, alpha=0.95, marker="^",
                           edgecolors="black", linewidths=0.7, zorder=5)

        # Diagonal y=x reference line
        lim = clip * 1.05 if clip else max(abs(x_arr).max(), abs(y_arr).max()) * 1.05
        ax.plot([-lim, lim], [-lim, lim], color="#888888", linewidth=0.9,
                linestyle="--", alpha=0.6)
        ax.axhline(0, color="#BBBBBB", linewidth=0.6, alpha=0.5)
        ax.axvline(0, color="#BBBBBB", linewidth=0.6, alpha=0.5)

        if clip:
            ax.set_xlim(-clip * 1.05, clip * 1.05)
            ax.set_ylim(-clip * 1.05, clip * 1.05)
            ax.text(0.01, 0.99, "▲ clipped (|log₂FC| > 15)",
                    transform=ax.transAxes, fontsize=7, color="#AAAAAA", va="top")

        # --- Labels: orange dots only, one label per gene name ---
        # Step 1: collect dot positions — use tx_to_name for gene name
        gene_best = {}
        for i, txid in enumerate(tx_ids):
            if not both_sig_mask[i]:
                continue
            x_real, y_real = x_arr[i], y_arr[i]
            # Both already pass |log2FC|>1.5 by construction — label all orange
            gname = tx_to_name.get(txid, txid)
            if gname == "unannotated":
                gname = txid   # fallback to transcript ID
            if gname not in gene_best or abs(x_real) > abs(gene_best[gname][0]):
                gene_best[gname] = (x_plot[i], y_plot[i])

        # Step 2: seed labels with a small directional offset, then
        # use adjust_text to find non-overlapping positions.
        # For clipped outliers sitting at the axis boundary, flip the label
        # inward so it doesn't get cut off.
        clip_tol = (clip * 0.97) if clip else None

        texts = []
        gene_order = []
        for gname, (dx, dy) in gene_best.items():
            # Detect if point is at a clipped boundary — flip label inward
            at_right  = clip_tol and dx >= clip_tol
            at_left   = clip_tol and dx <= -clip_tol
            at_top    = clip_tol and dy >= clip_tol
            at_bottom = clip_tol and dy <= -clip_tol

            if at_right:
                offset_x, ha = -2.0, "right"
            elif at_left:
                offset_x, ha = 2.0, "left"
            else:
                offset_x, ha = (1.2 if dx >= 0 else -1.2), "center"

            offset_y = -1.5 if at_top else (1.5 if at_bottom else
                        (0.8 if dy >= 0 else -0.8))

            texts.append(ax.text(dx + offset_x, dy + offset_y, gname,
                                 fontsize=8, color="#1a1a1a", fontstyle="italic",
                                 fontweight="bold" if (at_right or at_left) else "normal"))
            gene_order.append((gname, dx, dy))

        adjust_text(texts, ax=ax,
                    expand=(3.5, 3.5), force_text=(2.5, 2.5),
                    force_points=(1.5, 1.5),
                    lim=700)

        # Step 3: remove adjust_text texts and redraw as ax.annotate
        # so lines reliably touch both the label and the dot
        for txt, (gname, dot_x, dot_y) in zip(texts, gene_order):
            lx, ly = txt.get_position()
            txt.remove()
            ax.annotate(
                gname,
                xy=(dot_x, dot_y),
                xytext=(lx, ly),
                fontsize=8, color="#1a1a1a", fontstyle="italic",
                ha="center", va="center",
                arrowprops=dict(arrowstyle="-", color="#888888", lw=0.7,
                                shrinkA=4, shrinkB=4),
            )

    # --- v1: original (full range) ---
    fig, ax = plt.subplots(figsize=(13, 7))
    draw_de_scatter(ax, x_arr, y_arr, colors, sizes, tx_ids, both_sig_mask, clip=None)
    ax.legend(handles=legend_elements, fontsize=8,
              bbox_to_anchor=(1.02, 1), loc="upper left")
    ax.set_xlabel("log₂ fold change — Illumina (PerDKO / WT)", fontsize=12)
    ax.set_ylabel("log₂ fold change — Nanopore (PerDKO / WT)", fontsize=12)
    ax.set_title(
        "DE concordance: Illumina vs Nanopore (PerDKO vs WT) — transcript level\n"
        f"Transcripts in both platforms (n={len(common_tx)}) | threshold: padj<0.05, |log₂FC|>1.5",
        fontsize=11)
    plt.tight_layout()
    p = os.path.join(OUT_DIR, "step15_DE_concordance_scatter.png")
    plt.savefig(p, dpi=150, bbox_inches="tight");  plt.close()
    print(f"  Saved: step15_DE_concordance_scatter.png")

    # --- v2 bis: clipped axes ±15 ---
    fig, ax = plt.subplots(figsize=(13, 7))
    draw_de_scatter(ax, x_arr, y_arr, colors, sizes, tx_ids, both_sig_mask, clip=15)
    ax.legend(handles=legend_elements, fontsize=8,
              bbox_to_anchor=(1.02, 1), loc="upper left")
    ax.set_xlabel("log₂ fold change — Illumina (PerDKO / WT)", fontsize=12)
    ax.set_ylabel("log₂ fold change — Nanopore (PerDKO / WT)", fontsize=12)
    ax.set_title(
        "DE concordance: Illumina vs Nanopore (PerDKO vs WT) — transcript level\n"
        f"Transcripts in both platforms (n={len(common_tx)}) — axes clipped ±15 | padj<0.05, |log₂FC|>1.5",
        fontsize=11)
    plt.tight_layout()
    p = os.path.join(OUT_DIR, "step15_DE_concordance_scatter_v2.png")
    plt.savefig(p, dpi=150, bbox_inches="tight");  plt.close()
    print(f"  Saved: step15_DE_concordance_scatter_v2.png")

print("\n=== ALL DONE ===")
