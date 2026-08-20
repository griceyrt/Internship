#!/usr/bin/env python3
"""
Targeted fix: rebuilds ONLY GSE130613_STARhtseq_vs_Nanopore_DE_overlap.xlsx
with a gene_id fallback for blank gene_name cells. Reuses the same
data-loading/overlap logic as make_star_htseq_de_figures.py /
make_gse130613_publication_figures.py, skips all plotting.
Author: Gricey

Run from: orthogonal_validation/
"""
import os
import pandas as pd
import openpyxl

BASE_DIR  = os.environ.get("OV_BASE_DIR", "/Users/gricey/Desktop/Internship/orthogonal_validation")
DESEQ_NEW = os.path.join(BASE_DIR, "results", "GSE130613", "normalisation_star_htseq_original_gtf", "deseq2_star_htseq_KO_vs_WT.tsv")
NANO_FILE = os.path.join(BASE_DIR, "data", "Table3-differential_expressiong_WTvsPerKO.xlsx")
TABLE_DIR = os.path.join(BASE_DIR, "results", "GSE130613", "tables")

PADJ_THRESH = 0.05
LFC_15 = 1.5

de = pd.read_csv(DESEQ_NEW, sep="\t").dropna(subset=["log2FoldChange", "padj"])
de = de[de["padj"] > 0].set_index("gene_id")

wb = openpyxl.load_workbook(NANO_FILE)
ws_sig = wb["perKO_sig"]
rows_sig = list(ws_sig.iter_rows(values_only=True))
header_sig = [str(c) if c is not None else "idx" for c in rows_sig[0]]
nano_sig_geneid = {}
for r in rows_sig[1:]:
    row = dict(zip(header_sig, r))
    gid = row.get("gene_id")
    if gid and gid not in nano_sig_geneid:
        try:
            nano_sig_geneid[gid] = float(row["log2FC_perko"])
        except (TypeError, ValueError):
            pass

ws_all = wb["DE_all"]
rows_all = list(ws_all.iter_rows(values_only=True))
header_all = list(rows_all[0])
nano_all_geneid = {}
for r in rows_all[1:]:
    row = dict(zip(header_all, r))
    gid = row.get("gene_id")
    if not gid:
        continue
    try:
        lfc = float(row["log2FC_perko"])
        padj = float(row["padj_per_ko"])
    except (TypeError, ValueError):
        continue
    if gid not in nano_all_geneid or abs(lfc) > abs(nano_all_geneid[gid]["log2FC"]):
        nano_all_geneid[gid] = {"log2FC": lfc, "padj": padj}

ill_sig = set(de[(de["padj"] < PADJ_THRESH) & (de["log2FoldChange"].abs() > LFC_15)].index)
common = set(de.index) & set(nano_all_geneid.keys())
nano_sig = set(nano_sig_geneid.keys())
overlap = ill_sig & nano_sig & common
print(f"Overlap genes: {len(overlap)} (should match the existing 16)")

gene_map_ht = de["gene_name"].to_dict() if "gene_name" in de.columns else {}

rows_out = []
for g in sorted(overlap):
    gname_out = gene_map_ht.get(g, "")
    if pd.isna(gname_out) or gname_out == "":
        gname_out = g
    rows_out.append({
        "gene_id": g,
        "gene_name": gname_out,
        "log2FC_Illumina_STARhtseq": de.loc[g, "log2FoldChange"],
        "padj_Illumina_STARhtseq": de.loc[g, "padj"],
        "log2FC_Nanopore": nano_all_geneid[g]["log2FC"],
        "padj_Nanopore": nano_all_geneid[g]["padj"],
        "same_direction": (de.loc[g, "log2FoldChange"] > 0) == (nano_all_geneid[g]["log2FC"] > 0),
    })
overlap_df = pd.DataFrame(rows_out).sort_values("padj_Illumina_STARhtseq")
same_dir_n = overlap_df["same_direction"].sum()
print(f"Same direction: {same_dir_n}, opposite: {len(overlap_df) - same_dir_n} (should match existing 12/4)")

xlsx_path = os.path.join(TABLE_DIR, "GSE130613_STARhtseq_vs_Nanopore_DE_overlap.xlsx")
with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
    overlap_df.to_excel(writer, sheet_name="Overlap_genes", index=False)
    summary = pd.DataFrame({
        "Metric": ["Common genes tested (both platforms)", "Illumina sig (padj<0.05, |log2FC|>1.5)",
                   "Nanopore sig (perKO_sig)", "Overlap (both sig)", "Same direction", "Opposite direction"],
        "Value": [len(common), len(ill_sig), len(nano_sig), len(overlap), int(same_dir_n), int(len(overlap_df) - same_dir_n)],
    })
    summary.to_excel(writer, sheet_name="Summary", index=False)

wb_out = openpyxl.load_workbook(xlsx_path)
for ws_name in wb_out.sheetnames:
    ws_out = wb_out[ws_name]
    for cell in ws_out[1]:
        cell.font = openpyxl.styles.Font(name="Arial", bold=True)
    for row in ws_out.iter_rows(min_row=2):
        for cell in row:
            cell.font = openpyxl.styles.Font(name="Arial")
    for col_idx, col_cells in enumerate(ws_out.columns, start=1):
        ws_out.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 24
wb_out.save(xlsx_path)
print(f"Saved: {xlsx_path}")
