#!/usr/bin/env python3
"""
Final publication-ready figures for GSE130613, NEW pipeline (STAR + rMATS / htseq).
Author: Gricey

Same minimal style as the original make_final_figures.py (Kiran's requested
format): SVG, Arial, no title, no legend, bare-minimum axes. A companion
figure_legend_new_pipeline.txt explains colors and thresholds, same pattern
as the original figure_legend.txt.

Figures produced (analogous to the originals, new pipeline):
  GSE130613_rmats_dPSI_volcano_p0.3_minimal.svg        (like Illumina_events_dPSI_volcano.svg)
  GSE130613_rmats_dPSI_volcano_FDR0.05_minimal.svg     (same, stricter threshold)
  GSE130613_STARhtseq_DE_volcano_minimal.svg           (like Illumina_DE_genes_volcano.svg)
  GSE130613_STARhtseq_vs_Nanopore_DE_concordance_minimal.svg  (like Illumina_Nanopore_DE_genes_concordance.svg)

No transcript-level equivalent (Illumina_Nanopore_DE_transcript_concordance.svg
has no counterpart here): htseq-count is gene-level only.

Output: results/GSE130613/figures/figures_final_new_pipeline/
Run from: orthogonal_validation/
"""

import os
import numpy as np
import pandas as pd
import openpyxl
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from adjustText import adjust_text

# GLOBAL STYLE — identical block to make_final_figures.py
matplotlib.rcParams.update({
    "font.family":       "sans-serif",
    "font.sans-serif":   ["Arial", "Helvetica", "DejaVu Sans"],
    "font.size":         8,
    "axes.spines.top":   False,
    "axes.spines.right": False,
    "axes.spines.left":  True,
    "axes.spines.bottom":True,
    "axes.edgecolor":    "#AAAAAA",
    "axes.linewidth":    0.7,
    "axes.grid":         False,
    "figure.facecolor":  "white",
    "axes.facecolor":    "white",
    "xtick.color":       "#444444",
    "ytick.color":       "#444444",
    "xtick.labelsize":   7,
    "ytick.labelsize":   7,
    "axes.labelsize":    8,
    "axes.labelcolor":   "#222222",
    "xtick.major.size":  3,
    "ytick.major.size":  3,
    "xtick.major.width": 0.6,
    "ytick.major.width": 0.6,
})

BASE_DIR  = os.environ.get("OV_BASE_DIR", "/Users/gricey/Desktop/Internship/orthogonal_validation")
RMATS_DIR = os.path.join(BASE_DIR, "results", "GSE130613", "SRP194523_rmats_2026-07-20")
DESEQ_NEW = os.path.join(BASE_DIR, "results", "GSE130613", "normalisation_star_htseq_original_gtf", "deseq2_star_htseq_KO_vs_WT.tsv")
NANO_FILE = os.path.join(BASE_DIR, "data", "Table3-differential_expressiong_WTvsPerKO.xlsx")
OUT = os.path.join(BASE_DIR, "results", "GSE130613", "figures", "figures_final_new_pipeline")
os.makedirs(OUT, exist_ok=True)

DPSI_THRESH = 0.1
PVAL_THRESH = 0.3
FDR_THRESH  = 0.05
LFC_THRESH  = 0.5
PADJ_THRESH = 0.05
LFC_15      = 1.5

EVENT_TYPE_MAP = {"SE": "SE", "A5SS": "A5", "A3SS": "A3", "MXE": "MX", "RI": "RI"}
EVENT_COLORS = {"A3": "#E69F00", "A5": "#56B4E9", "MX": "#0072B2", "RI": "#D55E00", "SE": "#CC79A7"}

def save_svg(fig, name):
    path = os.path.join(OUT, name)
    fig.savefig(path, format="svg", bbox_inches="tight")
    plt.close(fig)
    print(f"  Saved: {name}")

# LOAD rMATS (all 5 event types, JC files)
print("Loading rMATS JC output...")
dfs = []
for rmats_type, code in EVENT_TYPE_MAP.items():
    d = pd.read_csv(os.path.join(RMATS_DIR, f"{rmats_type}.MATS.JC.txt"), sep="\t")
    d["GeneID"] = d["GeneID"].str.strip('"')
    d["geneSymbol"] = d["geneSymbol"].str.strip('"')
    d["event_type"] = code
    dfs.append(d[["event_type", "GeneID", "geneSymbol", "PValue", "FDR", "IncLevelDifference"]])
rmats = pd.concat(dfs, ignore_index=True)
rmats["dPSI"] = -rmats["IncLevelDifference"]  # WT-KO -> KO-WT, project convention

# FIGURE A — rMATS dPSI volcano, p<0.3 (minimal)
print("\n=== Figure A: rMATS dPSI volcano, p<0.3 ===")
d = rmats[rmats["PValue"] > 0].copy()
d["neg_log10_p"] = -np.log10(d["PValue"])
d["sig"] = (d["PValue"] < PVAL_THRESH) & (d["dPSI"].abs() > DPSI_THRESH)
print(f"  Significant: {d['sig'].sum()}")

fig, ax = plt.subplots(figsize=(3.5, 3.5))
ns = d[~d["sig"]]
ax.scatter(ns["dPSI"], ns["neg_log10_p"], c="#CCCCCC", s=4, alpha=0.4, linewidths=0, zorder=1, rasterized=True)
for etype, col in EVENT_COLORS.items():
    sub = d[d["sig"] & (d["event_type"] == etype)]
    if len(sub):
        ax.scatter(sub["dPSI"], sub["neg_log10_p"], c=col, s=14, alpha=0.9, linewidths=0, zorder=2, rasterized=True)
ax.axvline(DPSI_THRESH, color="#BBBBBB", ls="--", lw=0.7)
ax.axvline(-DPSI_THRESH, color="#BBBBBB", ls="--", lw=0.7)
ax.axhline(-np.log10(PVAL_THRESH), color="#BBBBBB", ls="--", lw=0.7)
ax.set_xlabel("dPSI (PerDKO − WT)")
ax.set_ylabel("−log₁₀(p)")
save_svg(fig, "GSE130613_rmats_dPSI_volcano_p0.3_minimal.svg")

# FIGURE B — rMATS dPSI volcano, FDR<0.05 (minimal)
print("\n=== Figure B: rMATS dPSI volcano, FDR<0.05 ===")
d2 = rmats[rmats["FDR"] > 0].copy()
d2["neg_log10_fdr"] = -np.log10(d2["FDR"])
d2["sig"] = (d2["FDR"] < FDR_THRESH) & (d2["dPSI"].abs() > DPSI_THRESH)
print(f"  Significant: {d2['sig'].sum()}")

fig, ax = plt.subplots(figsize=(3.5, 3.5))
ns2 = d2[~d2["sig"]]
ax.scatter(ns2["dPSI"], ns2["neg_log10_fdr"], c="#CCCCCC", s=4, alpha=0.4, linewidths=0, zorder=1, rasterized=True)
for etype, col in EVENT_COLORS.items():
    sub = d2[d2["sig"] & (d2["event_type"] == etype)]
    if len(sub):
        ax.scatter(sub["dPSI"], sub["neg_log10_fdr"], c=col, s=14, alpha=0.9, linewidths=0, zorder=2, rasterized=True)
ax.axvline(DPSI_THRESH, color="#BBBBBB", ls="--", lw=0.7)
ax.axvline(-DPSI_THRESH, color="#BBBBBB", ls="--", lw=0.7)
ax.axhline(-np.log10(FDR_THRESH), color="#BBBBBB", ls="--", lw=0.7)
ax.set_xlabel("dPSI (PerDKO − WT)")
ax.set_ylabel("−log₁₀(FDR)")
save_svg(fig, "GSE130613_rmats_dPSI_volcano_FDR0.05_minimal.svg")

# FIGURE C — STAR+htseq DE volcano (minimal)
print("\n=== Figure C: STAR+htseq DE volcano ===")
de = pd.read_csv(DESEQ_NEW, sep="\t").dropna(subset=["log2FoldChange", "padj"])
de = de[de["padj"] > 0].set_index("gene_id")
de["neg_log10_padj"] = -np.log10(de["padj"])
de["category"] = "ns"
de.loc[(de["log2FoldChange"] > LFC_THRESH) & (de["padj"] < PADJ_THRESH), "category"] = "up"
de.loc[(de["log2FoldChange"] < -LFC_THRESH) & (de["padj"] < PADJ_THRESH), "category"] = "down"
print(f"  up: {(de['category']=='up').sum()}, down: {(de['category']=='down').sum()}")
print(f"  max -log10(padj): {de['neg_log10_padj'].max():.2f} (no clipping needed, unlike the old ±28 pipeline)")

CIRCADIAN_GENES = [
    "Ren1", "Orm2", "Prtn3", "Cyp2b9", "Dbp", "Cxcl1", "Saa1", "Ighg2b",
    "Celf4", "Egr1", "Ciart", "Nr1d1", "Nr1d2", "Btg2", "Per3", "Ephx1",
    "Gstm2", "Clock", "Cyp2a5", "Arntl", "Cry1", "Kmt2d", "Gstm3",
    "Prkca", "Cyp7a1", "Fancm", "Cspg5"
]

fig, ax = plt.subplots(figsize=(4, 3.5))
ns_mask = de["category"] == "ns"
ax.scatter(de.loc[ns_mask, "log2FoldChange"], de.loc[ns_mask, "neg_log10_padj"],
           c="#CCCCCC", s=3, alpha=0.3, linewidths=0, zorder=1, rasterized=True)
up_mask = de["category"] == "up"
ax.scatter(de.loc[up_mask, "log2FoldChange"], de.loc[up_mask, "neg_log10_padj"],
           c="#D55E00", s=10, alpha=0.85, linewidths=0, zorder=2)
dn_mask = de["category"] == "down"
ax.scatter(de.loc[dn_mask, "log2FoldChange"], de.loc[dn_mask, "neg_log10_padj"],
           c="#0072B2", s=10, alpha=0.85, linewidths=0, zorder=2)
ax.axvline(LFC_THRESH, color="#BBBBBB", ls="--", lw=0.7)
ax.axvline(-LFC_THRESH, color="#BBBBBB", ls="--", lw=0.7)
ax.axhline(-np.log10(PADJ_THRESH), color="#BBBBBB", ls="--", lw=0.7)

texts, labels_out = [], []
for gname in CIRCADIAN_GENES:
    matches = de[(de["category"] != "ns") & (de.get("gene_name") == gname)] if "gene_name" in de.columns else de.iloc[0:0]
    if len(matches) == 0:
        continue
    row = matches.iloc[0]
    col = "#D55E00" if row["category"] == "up" else "#0072B2"
    ax.scatter(row["log2FoldChange"], row["neg_log10_padj"], c=col, s=22, zorder=5, edgecolors="black", linewidths=0.4)
    texts.append(ax.text(row["log2FoldChange"], row["neg_log10_padj"], gname, fontsize=6, fontstyle="italic"))
    labels_out.append((gname, row["log2FoldChange"], row["neg_log10_padj"]))

if texts:
    adjust_text(texts, ax=ax, expand=(2, 2), force_text=(1.2, 1.2), force_points=(1, 1), lim=400)
    for txt, (gname, dx, dy) in zip(texts, labels_out):
        lx, ly = txt.get_position()
        txt.remove()
        ax.annotate(gname, xy=(dx, dy), xytext=(lx, ly), fontsize=6, fontstyle="italic",
                    ha="center", va="center",
                    arrowprops=dict(arrowstyle="-", color="#888888", lw=0.5, shrinkA=3, shrinkB=3))
print(f"  Circadian genes labeled: {len(labels_out)}")

ax.set_xlabel("log₂FC (PerDKO / WT)")
ax.set_ylabel("−log₁₀(padj)")
save_svg(fig, "GSE130613_STARhtseq_DE_volcano_minimal.svg")

# FIGURE D — STAR+htseq vs Nanopore DE concordance, gene level (minimal)
print("\n=== Figure D: STAR+htseq vs Nanopore concordance ===")
wb = openpyxl.load_workbook(NANO_FILE, read_only=True, data_only=True)

ws_sig = wb["perKO_sig"]
rows_sig = list(ws_sig.iter_rows(values_only=True))
header_sig = [str(c) if c else "idx" for c in rows_sig[0]]
nano_sig_gid = {}
for r in rows_sig[1:]:
    row = dict(zip(header_sig, r))
    gid = row.get("gene_id")
    if gid and gid not in nano_sig_gid:
        try:
            nano_sig_gid[gid] = float(row["log2FC_perko"])
        except (TypeError, ValueError):
            pass

ws_all = wb["DE_all"]
rows_all = list(ws_all.iter_rows(values_only=True))
header_all = list(rows_all[0])
nano_all_gid = {}
for r in rows_all[1:]:
    row = dict(zip(header_all, r))
    gid = row.get("gene_id")
    if not gid:
        continue
    try:
        lfc = float(row["log2FC_perko"])
    except (TypeError, ValueError):
        continue
    if gid not in nano_all_gid or abs(lfc) > abs(nano_all_gid[gid]):
        nano_all_gid[gid] = lfc

ill_sig = set(de[(de["padj"] < PADJ_THRESH) & (de["log2FoldChange"].abs() > LFC_15)].index)
common = set(de.index) & set(nano_all_gid.keys())
nano_sig = set(nano_sig_gid.keys())
overlap = ill_sig & nano_sig & common
print(f"  Common: {len(common)}, overlap: {len(overlap)}")

genes = list(common)
x = np.array([float(de.loc[g, "log2FoldChange"]) for g in genes])
y = np.array([float(nano_all_gid[g]) for g in genes])
mask_both = np.array([g in overlap for g in genes])
mask_ill = np.array([g in ill_sig and g not in overlap for g in genes])
mask_nano = np.array([g in nano_sig and g not in overlap for g in genes])
mask_grey = ~(mask_both | mask_ill | mask_nano)

fig, ax = plt.subplots(figsize=(4, 4))
ax.scatter(x[mask_grey], y[mask_grey], c="#CCCCCC", s=3, alpha=0.35, linewidths=0, zorder=1, rasterized=True)
ax.scatter(x[mask_ill], y[mask_ill], c="#56B4E9", s=12, alpha=0.85, linewidths=0, zorder=2)
ax.scatter(x[mask_nano], y[mask_nano], c="#009E73", s=12, alpha=0.85, linewidths=0, zorder=3)
ax.scatter(x[mask_both], y[mask_both], c="#E69F00", s=30, alpha=0.95, edgecolors="black", linewidths=0.5, zorder=4)

lim = max(abs(x).max(), abs(y).max()) * 1.05
ax.plot([-lim, lim], [-lim, lim], color="#BBBBBB", lw=0.7, ls="--", alpha=0.6)
ax.axhline(0, color="#BBBBBB", lw=0.5)
ax.axvline(0, color="#BBBBBB", lw=0.5)

gene_map_ht = de["gene_name"].to_dict() if "gene_name" in de.columns else {}
texts, labels_out = [], []
for i, g in enumerate(genes):
    if not mask_both[i]:
        continue
    gname = gene_map_ht.get(g, g)
    if pd.isna(gname) or gname == "":
        gname = g
    ox = 1.0 if x[i] >= 0 else -1.0
    oy = 0.6 if y[i] >= 0 else -0.6
    texts.append(ax.text(x[i] + ox, y[i] + oy, gname, fontsize=6, fontstyle="italic"))
    labels_out.append((gname, x[i], y[i]))

adjust_text(texts, ax=ax, expand=(3, 3), force_text=(2, 2), force_points=(1.5, 1.5), lim=600)
for txt, (gname, dx, dy) in zip(texts, labels_out):
    lx, ly = txt.get_position()
    txt.remove()
    ax.annotate(gname, xy=(dx, dy), xytext=(lx, ly), fontsize=6, fontstyle="italic",
                ha="center", va="center",
                arrowprops=dict(arrowstyle="-", color="#888888", lw=0.5, shrinkA=3, shrinkB=3))

ax.set_xlabel("log₂FC Illumina, STAR+htseq (PerDKO / WT)")
ax.set_ylabel("log₂FC Nanopore (PerDKO / WT)")
save_svg(fig, "GSE130613_STARhtseq_vs_Nanopore_DE_concordance_minimal.svg")

# LEGEND TXT
legend_text = f"""FIGURE LEGEND — Orthogonal Validation, GSE130613 NEW PIPELINE (STAR + rMATS / htseq-count)
Generated: 2026-07-20
Dataset: Liver, CT16-20, constant darkness, n=4 per condition
Companion to the original figure_legend.txt (Salmon + SUPPA pipeline, now superseded for DE/quantification).

=== GSE130613_rmats_dPSI_volcano_p0.3_minimal.svg ===
dPSI volcano — rMATS-turbo results (PerDKO vs WT), run on STAR BAMs.
X-axis: dPSI (PerDKO − WT). Y-axis: -log10(p-value).
Threshold lines: |dPSI| = 0.1 (vertical), p = 0.3 (horizontal) -- matches the old SUPPA threshold.
Grey dots: not significant.
Coloured dots (significant, |dPSI|>0.1, p<0.3), by event type:
  SE (skipping exon)         = #CC79A7 (pink/magenta)
  MX (mutually exclusive)    = #0072B2 (dark blue)
  RI (retained intron)       = #D55E00 (orange-red)
  A3 (alternative 3' ss)     = #E69F00 (amber)
  A5 (alternative 5' ss)     = #56B4E9 (light blue)
Note: rMATS does not call AF/AL (alternative first/last exon), Kiran confirmed these are
Pol II-dependent, not true alternative-splicing events.
Total significant events: 43 (of 1,145 testable, out of a much larger annotated catalog --
95-98% of candidate junctions had too few reads to test at all, a direct consequence of
MARS-seq's 3' bias, not a tool limitation. MXE was the exception, 17% testable.)

=== GSE130613_rmats_dPSI_volcano_FDR0.05_minimal.svg ===
Same data and colors as above, stricter threshold: |dPSI|>0.1, FDR<0.05.
Y-axis: -log10(FDR) instead of raw p-value.
Total significant events: 8.

=== GSE130613_STARhtseq_DE_volcano_minimal.svg ===
DE volcano — STAR + htseq-count + DESeq2 (PerDKO vs WT), gene level. Adopted pipeline,
replaces Salmon-based quantification (which had an inflated log2FC spread, -28 to +28,
traced to Salmon's EM-based transcript assignment).
X-axis: log2 fold change (PerDKO / WT). Y-axis: -log10(padj).
Threshold lines: |log2FC| = 0.5 (vertical), padj = 0.05 (horizontal). No clipping needed,
full range is -6.02 to +8.44.
Grey dots: not significant.
Orange/red dots (#D55E00): upregulated in PerDKO (n=238).
Blue dots (#0072B2): downregulated in PerDKO (n=152).
Labels: circadian and selected genes (Kiran's list), where significant.

=== GSE130613_STARhtseq_vs_Nanopore_DE_concordance_minimal.svg ===
DE concordance scatter — Illumina (STAR+htseq) vs Nanopore, GENE level (PerDKO vs WT).
X-axis: log2FC Illumina. Y-axis: log2FC Nanopore.
Threshold (both sides): padj<0.05 AND |log2FC|>1.5.
Grey dots: present in both datasets, not significant in either.
Blue dots: significant in Illumina only (n={mask_ill.sum()}).
Green dots: significant in Nanopore only (n={mask_nano.sum()}).
Orange dots: significant in BOTH platforms (n={mask_both.sum()}) -- labeled with gene names.
Total genes in both datasets: {len(common)}.
Note: the old Salmon-based version of this figure had 27 orange (overlap) genes on a
9,850-gene background; the drop to {mask_both.sum()} here is consistent with the old
pipeline's inflated spread having pushed some genes over the |log2FC|>1.5 threshold
artificially.

No transcript-level equivalent exists for this pipeline: htseq-count only counts reads
per gene, it has no transcript-level resolution (that is the EM-based isoform assignment
this pipeline switch deliberately avoided).
"""

with open(os.path.join(OUT, "figure_legend_new_pipeline.txt"), "w") as f:
    f.write(legend_text)
print("\nSaved: figure_legend_new_pipeline.txt")
print("\n=== ALL DONE ===")
