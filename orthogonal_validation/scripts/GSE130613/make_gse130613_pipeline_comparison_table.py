#!/usr/bin/env python3
"""
GSE130613 new vs old pipeline comparison table: counts of transcripts/genes
tested and significant/DE for STAR+htseq-count (NEW, adopted) vs
Salmon+tximport+stageR (OLD, inflated log2FC spread). Raw per-row data is
loaded into the workbook so the Summary sheet's counts are live COUNTIFS
formulas (except the OLD gene-level dedup, computed in Python -- see the
note cell on the Summary sheet).
Author: Gricey

Run from: orthogonal_validation/
"""

import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

BASE_DIR = os.environ.get("OV_BASE_DIR", "/Users/gricey/Desktop/Internship/orthogonal_validation")
NEW_FILE = os.path.join(BASE_DIR, "results", "GSE130613", "normalisation_star_htseq_original_gtf", "deseq2_star_htseq_KO_vs_WT.tsv")
OLD_FILE = os.path.join(BASE_DIR, "results", "GSE130613", "normalisation", "deseq2_KO_vs_WT.tsv")
OUT_PATH = os.path.join(BASE_DIR, "results", "GSE130613", "tables", "GSE130613_pipeline_comparison_for_Khushi.xlsx")

LFC_THRESH = 0.5
PADJ_THRESH = 0.05

print("=== Loading NEW pipeline (STAR+htseq, gene-level) ===")
new_df = pd.read_csv(NEW_FILE, sep="\t")
print(f"  {len(new_df)} rows")

print("=== Loading OLD pipeline (Salmon+tximport+stageR, transcript-level) ===")
old_df = pd.read_csv(OLD_FILE, sep="\t")
print(f"  {len(old_df)} rows")

# Gene-level dedup for OLD pipeline (Python, documented in the xlsx note --
# a MAXIFS-based Excel formula would double-count exact abs(log2FC) ties)
old_reliable = old_df.dropna(subset=["log2FoldChange", "padj_gene_stageR"])
old_reliable = old_reliable[old_reliable["padj_gene_stageR"] > 0]
old_reliable = old_reliable.copy()
old_reliable["abs_lfc"] = old_reliable["log2FoldChange"].abs()
old_gene_level = old_reliable.sort_values("abs_lfc", ascending=False).drop_duplicates("gene_id")
old_gene_sig = old_gene_level[(old_gene_level["padj_gene_stageR"] < PADJ_THRESH) & (old_gene_level["abs_lfc"] > LFC_THRESH)]
old_gene_total = len(old_gene_level)
old_gene_sig_n = len(old_gene_sig)
old_gene_up = int((old_gene_sig["log2FoldChange"] > 0).sum())
old_gene_down = int((old_gene_sig["log2FoldChange"] < 0).sum())
print(f"  OLD gene-level (deduped): {old_gene_total} genes, {old_gene_sig_n} significant ({old_gene_up} up, {old_gene_down} down)")

wb = openpyxl.Workbook()

def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(1, c)
        cell.font = Font(name="Arial", bold=True)

def style_body(ws, nrows, ncols, start_row=2):
    for r in range(start_row, nrows + start_row):
        for c in range(1, ncols + 1):
            ws.cell(r, c).font = Font(name="Arial")

def autosize(ws, ncols, width=16):
    for c in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(c)].width = width

# --- NEW pipeline raw sheet (gene-level) ---
ws_new = wb.active
ws_new.title = "STARhtseq_genes_NEW"
headers_new = ["gene_id", "gene_name", "log2FoldChange", "padj", "sig", "direction"]
ws_new.append(headers_new)
n_new = len(new_df)
for i, row in new_df.reset_index(drop=True).iterrows():
    r = i + 2
    ws_new.cell(r, 1, row["gene_id"])
    ws_new.cell(r, 2, row.get("gene_name", ""))
    lfc_val = row["log2FoldChange"] if pd.notna(row["log2FoldChange"]) else None
    padj_val = row["padj"] if pd.notna(row["padj"]) else None
    ws_new.cell(r, 3, float(lfc_val) if lfc_val is not None else None)
    ws_new.cell(r, 4, float(padj_val) if padj_val is not None else None)
    ws_new.cell(r, 5, f'=IF(AND(ISNUMBER(D{r}),ISNUMBER(C{r})),AND(D{r}<{PADJ_THRESH},ABS(C{r})>{LFC_THRESH}),FALSE)')
    ws_new.cell(r, 6, f'=IF(ISNUMBER(C{r}),IF(C{r}>0,"up","down"),"")')
style_header(ws_new, len(headers_new))
style_body(ws_new, n_new, len(headers_new))
autosize(ws_new, len(headers_new))

# --- OLD pipeline raw sheet (transcript-level rows) ---
ws_old = wb.create_sheet("Salmon_transcripts_OLD")
headers_old = ["transcript_id", "gene_id", "gene_name", "log2FoldChange", "padj_gene_stageR", "sig", "direction"]
ws_old.append(headers_old)
n_old = len(old_df)
for i, row in old_df.reset_index(drop=True).iterrows():
    r = i + 2
    ws_old.cell(r, 1, row["transcript_id"])
    ws_old.cell(r, 2, row["gene_id"])
    ws_old.cell(r, 3, row.get("gene_name", ""))
    lfc_val = row["log2FoldChange"] if pd.notna(row["log2FoldChange"]) else None
    padj_val = row["padj_gene_stageR"] if pd.notna(row["padj_gene_stageR"]) else None
    ws_old.cell(r, 4, float(lfc_val) if lfc_val is not None else None)
    ws_old.cell(r, 5, float(padj_val) if padj_val is not None else None)
    ws_old.cell(r, 6, f'=IF(AND(ISNUMBER(E{r}),ISNUMBER(D{r})),AND(E{r}<{PADJ_THRESH},ABS(D{r})>{LFC_THRESH}),FALSE)')
    ws_old.cell(r, 7, f'=IF(ISNUMBER(D{r}),IF(D{r}>0,"up","down"),"")')
style_header(ws_old, len(headers_old))
style_body(ws_old, n_old, len(headers_old))
autosize(ws_old, len(headers_old))

# --- Summary sheet ---
ws_sum = wb.create_sheet("Summary", 0)  # make it the first sheet
sum_headers = ["Pipeline", "Level", "Total tested", "Reliable padj", "Significant / DE", "Up", "Down"]
ws_sum.append(sum_headers)
style_header(ws_sum, len(sum_headers))

new_last = n_new + 1
old_last = n_old + 1

# Row 2: NEW, gene-level -- fully live formulas
ws_sum.cell(2, 1, "STAR+htseq-count (NEW, adopted)")
ws_sum.cell(2, 2, "Gene")
ws_sum.cell(2, 3, f"=COUNTA(STARhtseq_genes_NEW!A2:A{new_last})")
ws_sum.cell(2, 4, f"=COUNT(STARhtseq_genes_NEW!D2:D{new_last})")
ws_sum.cell(2, 5, f"=COUNTIF(STARhtseq_genes_NEW!E2:E{new_last},TRUE)")
ws_sum.cell(2, 6, f'=COUNTIFS(STARhtseq_genes_NEW!E2:E{new_last},TRUE,STARhtseq_genes_NEW!F2:F{new_last},"up")')
ws_sum.cell(2, 7, f'=COUNTIFS(STARhtseq_genes_NEW!E2:E{new_last},TRUE,STARhtseq_genes_NEW!F2:F{new_last},"down")')

# Row 3: OLD, transcript-row level -- fully live formulas
ws_sum.cell(3, 1, "Salmon + tximport + stageR (OLD)")
ws_sum.cell(3, 2, "Transcript (row-level)")
ws_sum.cell(3, 3, f"=COUNTA(Salmon_transcripts_OLD!A2:A{old_last})")
ws_sum.cell(3, 4, f"=COUNT(Salmon_transcripts_OLD!E2:E{old_last})")
ws_sum.cell(3, 5, f"=COUNTIF(Salmon_transcripts_OLD!F2:F{old_last},TRUE)")
ws_sum.cell(3, 6, f'=COUNTIFS(Salmon_transcripts_OLD!F2:F{old_last},TRUE,Salmon_transcripts_OLD!G2:G{old_last},"up")')
ws_sum.cell(3, 7, f'=COUNTIFS(Salmon_transcripts_OLD!F2:F{old_last},TRUE,Salmon_transcripts_OLD!G2:G{old_last},"down")')

# Row 4: OLD, gene-level deduped -- computed in Python, documented (see note)
ws_sum.cell(4, 1, "Salmon + tximport + stageR (OLD)")
ws_sum.cell(4, 2, "Gene (deduped, see note)")
ws_sum.cell(4, 3, old_gene_total)
ws_sum.cell(4, 4, old_gene_total)  # already filtered to reliable rows before dedup
ws_sum.cell(4, 5, old_gene_sig_n)
ws_sum.cell(4, 6, old_gene_up)
ws_sum.cell(4, 7, old_gene_down)

style_body(ws_sum, 3, len(sum_headers))
autosize(ws_sum, len(sum_headers), width=26)

note = (
    "Thresholds throughout: padj<0.05, |log2FoldChange|>0.5. "
    "Row 4 (OLD, gene-level) is computed in Python, not a live formula: one representative "
    "transcript per gene_id (the transcript with max |log2FoldChange|), then counted -- the "
    "same method used for step13/step14 figures throughout this project. A tie-safe version of "
    "this dedup isn't expressible as a robust Excel formula, so it's reported as a fixed value here. "
    "IMPORTANT CAVEAT: the OLD (Salmon) pipeline was later found to have an inflated log2FC spread "
    "(-28 to +28) traced to Salmon's EM-based transcript quantification; STAR+htseq-count (NEW) was "
    "adopted to fix this and is the pipeline Kiran referred to as 'solid'. Recommend confirming with "
    "Kiran/Bharath which row(s) should actually go in the rebuttal before using the OLD numbers."
)
ws_sum.cell(6, 1, note)
ws_sum.cell(6, 1).font = Font(name="Arial", italic=True, size=9)
ws_sum.cell(6, 1).alignment = Alignment(wrap_text=True)
ws_sum.merge_cells(start_row=6, start_column=1, end_row=6, end_column=7)
ws_sum.row_dimensions[6].height = 90

os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
wb.save(OUT_PATH)
print(f"\nSaved: {OUT_PATH}")
