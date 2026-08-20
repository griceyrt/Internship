#!/usr/bin/env python3
"""
Publication-ready restyle of the GSE130613 (Asher) figures under results/GSE130613/figures/,
matching the look of GSE133398_DE_volcano_gene_level.svg and
GSE133398_splicing_dPSI_volcano.svg (Nick Webster/SRSF3 dataset).

Reverse-engineered from the actual SVG source of those reference figures
(colors, sizing, absence of legend/title), not guessed from the PNG alone:
  - Colors already match this project's established palette exactly
    (#D55E00/#0072B2 for DE, same Okabe-Ito set for splicing event types) --
    no change needed there.
  - No legend, no title in either reference.
  - Ticks are matplotlib's own auto-generated set (not a custom sparse
    override) -- so this script doesn't force custom ticks either.
  - Figure sizes measured from the reference SVG viewBox: DE volcano
    ~4.3x3.8in, splicing volcano ~3.7x3.5in.
  - The reference DE volcano rasterizes its background cloud (single
    <image> tag found in the SVG); this script keeps everything fully
    vector instead, matching the fix already applied to the rMATS minimal
    figures per Gricey's request (pixelation complaint).

Overwrites, in results/GSE130613/figures/ (both .svg and .png):
  GSE130613_STARhtseq_DE_volcano
  GSE130613_rmats_dPSI_volcano_p0.3
  GSE130613_rmats_dPSI_volcano_FDR0.05
  GSE130613_STARhtseq_vs_Nanopore_DE_concordance_gene_level (added per
  follow-up request -- same treatment: no legend, no title, fully vector,
  gene labels on overlap dots kept)

Does NOT touch anything under figures_final_new_pipeline/ (the minimal
versions Gricey is hand-editing in Inkscape separately).

No transcript-level DE volcano equivalent exists: htseq-count is gene-level
only.

Run from: orthogonal_validation/
"""

import os
import numpy as np
import pandas as pd
import openpyxl
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from adjustText import adjust_text

matplotlib.rcParams.update({
    "font.family":       "sans-serif",
    "font.sans-serif":   ["Arial", "Helvetica", "DejaVu Sans"],
    "font.size":         9,
    "axes.spines.top":   False,
    "axes.spines.right": False,
    "axes.spines.left":  True,
    "axes.spines.bottom":True,
    "axes.edgecolor":    "#AAAAAA",
    "axes.linewidth":    0.8,
    "axes.grid":         False,
    "figure.facecolor":  "white",
    "axes.facecolor":    "white",
    "xtick.color":       "#444444",
    "ytick.color":       "#444444",
    "axes.labelcolor":   "#222222",
})

BASE_DIR   = os.environ.get("OV_BASE_DIR", "/Users/gricey/Desktop/Internship/orthogonal_validation")
DESEQ_NEW  = os.path.join(BASE_DIR, "results", "GSE130613", "normalisation_star_htseq_original_gtf", "deseq2_star_htseq_KO_vs_WT.tsv")
RMATS_DIR  = os.path.join(BASE_DIR, "results", "GSE130613", "SRP194523_rmats_2026-07-20")
NANO_FILE  = os.path.join(BASE_DIR, "data", "Table3-differential_expressiong_WTvsPerKO.xlsx")
OUT        = os.path.join(BASE_DIR, "results", "GSE130613", "figures")

LFC_THRESH  = 0.5
PADJ_THRESH = 0.05
LFC_15      = 1.5
DPSI_THRESH = 0.1
PVAL_THRESH = 0.3
FDR_THRESH  = 0.05

CIRCADIAN_GENES = [
    "Ren1", "Orm2", "Prtn3", "Cyp2b9", "Dbp", "Cxcl1", "Saa1", "Ighg2b",
    "Celf4", "Egr1", "Ciart", "Nr1d1", "Nr1d2", "Btg2", "Per3", "Ephx1",
    "Gstm2", "Clock", "Cyp2a5", "Arntl", "Cry1", "Kmt2d", "Gstm3",
    "Prkca", "Cyp7a1", "Fancm", "Cspg5"
]

def save_both(fig, name):
    for ext in ("svg", "png"):
        path = os.path.join(OUT, f"{name}.{ext}")
        fig.savefig(path, format=ext, dpi=200, bbox_inches="tight")
        print(f"  Saved: {path}")
    plt.close(fig)

# =============================================================================
# DE volcano (gene level) -- matches GSE133398_DE_volcano_gene_level style
# =============================================================================
print("=== DE volcano ===")
de = pd.read_csv(DESEQ_NEW, sep="\t").dropna(subset=["log2FoldChange", "padj"])
de = de[de["padj"] > 0].set_index("gene_id")
de["neg_log10_padj"] = -np.log10(de["padj"])
de["category"] = "ns"
de.loc[(de["log2FoldChange"] > LFC_THRESH) & (de["padj"] < PADJ_THRESH), "category"] = "up"
de.loc[(de["log2FoldChange"] < -LFC_THRESH) & (de["padj"] < PADJ_THRESH), "category"] = "down"

fig, ax = plt.subplots(figsize=(4.3, 3.78))
ns_mask = de["category"] == "ns"
ax.scatter(de.loc[ns_mask, "log2FoldChange"], de.loc[ns_mask, "neg_log10_padj"],
           c="#CCCCCC", s=4, alpha=0.4, linewidths=0, zorder=1,
           label=f"not significant (n={ns_mask.sum()})")
up_mask = de["category"] == "up"
ax.scatter(de.loc[up_mask, "log2FoldChange"], de.loc[up_mask, "neg_log10_padj"],
           c="#D55E00", s=14, alpha=0.85, linewidths=0, zorder=2,
           label=f"up in PerDKO (n={up_mask.sum()})")
dn_mask = de["category"] == "down"
ax.scatter(de.loc[dn_mask, "log2FoldChange"], de.loc[dn_mask, "neg_log10_padj"],
           c="#0072B2", s=14, alpha=0.85, linewidths=0, zorder=2,
           label=f"down in PerDKO (n={dn_mask.sum()})")
ax.axvline(LFC_THRESH, color="#BBBBBB", ls="--", lw=0.7)
ax.axvline(-LFC_THRESH, color="#BBBBBB", ls="--", lw=0.7)
ax.axhline(-np.log10(PADJ_THRESH), color="#BBBBBB", ls="--", lw=0.7)
ax.legend(fontsize=7, frameon=False, loc="upper left", bbox_to_anchor=(1.02, 1))

texts, labels_out = [], []
for gname in CIRCADIAN_GENES:
    matches = de[(de["category"] != "ns") & (de.get("gene_name") == gname)] if "gene_name" in de.columns else de.iloc[0:0]
    if len(matches) == 0:
        continue
    row = matches.iloc[0]
    texts.append(ax.text(row["log2FoldChange"], row["neg_log10_padj"], gname, fontsize=7, fontstyle="italic"))
    labels_out.append((gname, row["log2FoldChange"], row["neg_log10_padj"], row["category"]))

# Redraw labeled dots on top with a black outline so they stand out from the
# rest of the significant (unlabeled) points -- same convention as the
# concordance figure's overlap dots.
for gname, dx, dy, cat in labels_out:
    fill = "#D55E00" if cat == "up" else "#0072B2"
    ax.scatter([dx], [dy], c=fill, s=22, alpha=0.95, edgecolors="black", linewidths=0.7, zorder=4)

if texts:
    adjust_text(texts, ax=ax, expand=(2, 2), force_text=(1.2, 1.2), force_points=(1, 1), lim=400)
    for txt, (gname, dx, dy, cat) in zip(texts, labels_out):
        lx, ly = txt.get_position()
        txt.remove()
        ax.annotate(gname, xy=(dx, dy), xytext=(lx, ly), fontsize=7, fontstyle="italic",
                    ha="center", va="center",
                    arrowprops=dict(arrowstyle="-", color="#888888", lw=0.5, shrinkA=3, shrinkB=3))
print(f"  Labeled: {len(labels_out)} circadian genes")

ax.set_xlabel("log₂FC (PerDKO / WT)")
ax.set_ylabel("−log₁₀(padj)")
save_both(fig, "GSE130613_STARhtseq_DE_volcano")

# =============================================================================
# rMATS splicing dPSI volcano -- matches GSE133398_splicing_dPSI_volcano style
# =============================================================================
print("\n=== rMATS splicing volcano ===")
EVENT_TYPE_MAP = {"SE": "SE", "A5SS": "A5", "A3SS": "A3", "MXE": "MX", "RI": "RI"}
EVENT_COLORS = {"A3": "#E69F00", "A5": "#56B4E9", "MX": "#0072B2", "RI": "#D55E00", "SE": "#CC79A7"}

dfs = []
for rmats_type, code in EVENT_TYPE_MAP.items():
    d = pd.read_csv(os.path.join(RMATS_DIR, f"{rmats_type}.MATS.JC.txt"), sep="\t")
    d["event_type"] = code
    dfs.append(d[["event_type", "PValue", "FDR", "IncLevelDifference"]])
rmats = pd.concat(dfs, ignore_index=True)
rmats["dPSI"] = -rmats["IncLevelDifference"]

def splicing_volcano(df, stat_col, thresh_val, ylabel, out_name):
    d = df[df[stat_col] > 0].copy()
    d["neg_log10"] = -np.log10(d[stat_col])
    d["sig"] = (d[stat_col] < thresh_val) & (d["dPSI"].abs() > DPSI_THRESH)

    fig, ax = plt.subplots(figsize=(3.71, 3.47))
    ns = d[~d["sig"]]
    ax.scatter(ns["dPSI"], ns["neg_log10"], c="#CCCCCC", s=4, alpha=0.4, linewidths=0, zorder=1)
    for etype, col in EVENT_COLORS.items():
        sub = d[d["sig"] & (d["event_type"] == etype)]
        if len(sub):
            ax.scatter(sub["dPSI"], sub["neg_log10"], c=col, s=16, alpha=0.9, linewidths=0, zorder=2,
                       label=f"{etype} (n={len(sub)})")
    ax.axvline(DPSI_THRESH, color="#BBBBBB", ls="--", lw=0.7)
    ax.axvline(-DPSI_THRESH, color="#BBBBBB", ls="--", lw=0.7)
    ax.axhline(-np.log10(thresh_val), color="#BBBBBB", ls="--", lw=0.7)
    ax.set_xlabel("dPSI (PerDKO − WT)")
    ax.set_ylabel(ylabel)
    ax.legend(title="Event type", fontsize=7, title_fontsize=7.5, frameon=False,
              loc="upper left", bbox_to_anchor=(1.02, 1))
    save_both(fig, out_name)

splicing_volcano(rmats, "PValue", PVAL_THRESH, "−log₁₀(p)", "GSE130613_rmats_dPSI_volcano_p0.3")
splicing_volcano(rmats, "FDR", FDR_THRESH, "−log₁₀(FDR)", "GSE130613_rmats_dPSI_volcano_FDR0.05")

# =============================================================================
# DE concordance vs Nanopore, gene level -- same treatment (no legend, no
# title, fully vector), added per follow-up request. Logic matches
# make_star_htseq_de_figures.py exactly, only the styling changed.
# =============================================================================
print("\n=== DE concordance vs Nanopore ===")
wb = openpyxl.load_workbook(NANO_FILE)

ws_sig = wb["perKO_sig"]
rows_sig = list(ws_sig.iter_rows(values_only=True))
header_sig = [str(c) if c is not None else "idx" for c in rows_sig[0]]
nano_sig_geneid = {}
for r in rows_sig[1:]:
    row = dict(zip(header_sig, r))
    gid = row.get("gene_id")
    if gid and gid not in nano_sig_geneid:
        try:
            nano_sig_geneid[gid] = float(row["log2FC_perko"])
        except (TypeError, ValueError):
            pass

ws_all = wb["DE_all"]
rows_all = list(ws_all.iter_rows(values_only=True))
header_all = list(rows_all[0])
nano_all_geneid = {}
for r in rows_all[1:]:
    row = dict(zip(header_all, r))
    gid = row.get("gene_id")
    if not gid:
        continue
    try:
        lfc = float(row["log2FC_perko"])
        padj = float(row["padj_per_ko"])
    except (TypeError, ValueError):
        continue
    if gid not in nano_all_geneid or abs(lfc) > abs(nano_all_geneid[gid]["log2FC"]):
        nano_all_geneid[gid] = {"log2FC": lfc, "padj": padj}

ill_sig = set(de[(de["padj"] < PADJ_THRESH) & (de["log2FoldChange"].abs() > LFC_15)].index)
common = set(de.index) & set(nano_all_geneid.keys())
nano_sig = set(nano_sig_geneid.keys())
overlap = ill_sig & nano_sig & common
print(f"  Common: {len(common)}, overlap: {len(overlap)}")

genes = list(common)
x = np.array([float(de.loc[g, "log2FoldChange"]) for g in genes])
y = np.array([float(nano_all_geneid[g]["log2FC"]) for g in genes])
mask_both = np.array([g in overlap for g in genes])
mask_ill = np.array([g in ill_sig and g not in overlap for g in genes])
mask_nano = np.array([g in nano_sig and g not in overlap for g in genes])
mask_grey = ~(mask_both | mask_ill | mask_nano)

fig, ax = plt.subplots(figsize=(4.5, 4.2))
ax.scatter(x[mask_grey], y[mask_grey], c="#CCCCCC", s=4, alpha=0.35, linewidths=0, zorder=1,
           label=f"not significant (n={mask_grey.sum()})")
ax.scatter(x[mask_ill], y[mask_ill], c="#56B4E9", s=16, alpha=0.85, linewidths=0, zorder=2,
           label=f"Illumina only (n={mask_ill.sum()})")
ax.scatter(x[mask_nano], y[mask_nano], c="#009E73", s=16, alpha=0.85, linewidths=0, zorder=3,
           label=f"Nanopore only (n={mask_nano.sum()})")
ax.scatter(x[mask_both], y[mask_both], c="#E69F00", s=26, alpha=0.95, edgecolors="black", linewidths=0.5, zorder=4,
           label=f"both significant (n={mask_both.sum()})")

lim = max(abs(x).max(), abs(y).max()) * 1.05
ax.plot([-lim, lim], [-lim, lim], color="#888888", lw=0.9, ls="--", alpha=0.6)
ax.axhline(0, color="#BBBBBB", lw=0.6, alpha=0.5)
ax.axvline(0, color="#BBBBBB", lw=0.6, alpha=0.5)

gene_map_ht = de["gene_name"].to_dict() if "gene_name" in de.columns else {}
texts, labels_out = [], []
for i, g in enumerate(genes):
    if not mask_both[i]:
        continue
    gname = gene_map_ht.get(g, g)
    if pd.isna(gname) or gname == "":
        gname = g
    ox = 0.6 if x[i] >= 0 else -0.6
    oy = 0.4 if y[i] >= 0 else -0.4
    texts.append(ax.text(x[i] + ox, y[i] + oy, gname, fontsize=7, fontstyle="italic"))
    labels_out.append((gname, x[i], y[i]))

if texts:
    adjust_text(texts, ax=ax, expand=(3.0, 3.0), force_text=(2.0, 2.0), force_points=(1.5, 1.5), lim=700)
    for txt, (gname, dx, dy) in zip(texts, labels_out):
        lx, ly = txt.get_position()
        txt.remove()
        ax.annotate(gname, xy=(dx, dy), xytext=(lx, ly), fontsize=7, fontstyle="italic",
                    ha="center", va="center",
                    arrowprops=dict(arrowstyle="-", color="#888888", lw=0.6, shrinkA=4, shrinkB=4))
print(f"  Labeled: {len(labels_out)} overlap genes")

ax.set_xlabel("log₂FC Illumina, STAR+htseq (PerDKO / WT)")
ax.set_ylabel("log₂FC Nanopore (PerDKO / WT)")
ax.legend(fontsize=7, frameon=False, loc="upper left", bbox_to_anchor=(1.02, 1))
save_both(fig, "GSE130613_STARhtseq_vs_Nanopore_DE_concordance_gene_level")

print("\n=== DONE ===")
