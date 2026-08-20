#!/usr/bin/env python3
"""
GSE130613 rMATS-turbo splicing results -- threshold-comparison deliverable:
volcano + counts at p<0.3 (matches old SUPPA convention) and at FDR<0.05
(stricter). Input: JC output (junction-counts-only, preferred over JCEC for
single-end data). Sign convention: rMATS b1=WT, b2=KO gives IncLevelDifference
= WT-KO, negated here to dPSI = KO-WT to match the rest of the project.
Author: Gricey

Run from: orthogonal_validation/
"""

import os
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# same style block as plot_results.py, for visual consistency across the project
matplotlib.rcParams.update({
    "font.family":          "sans-serif",
    "font.sans-serif":      ["Arial", "DejaVu Sans", "Helvetica", "Verdana"],
    "font.size":            11,
    "axes.spines.top":      False,
    "axes.spines.right":    False,
    "axes.spines.left":     True,
    "axes.spines.bottom":   True,
    "axes.edgecolor":       "#AAAAAA",
    "axes.linewidth":       0.8,
    "axes.grid":            True,
    "grid.color":           "#E5E5E5",
    "grid.linewidth":       0.6,
    "grid.linestyle":       "--",
    "axes.axisbelow":       True,
    "figure.facecolor":     "white",
    "axes.facecolor":       "white",
    "xtick.color":          "#444444",
    "ytick.color":          "#444444",
    "xtick.labelsize":      10,
    "ytick.labelsize":      10,
    "axes.labelcolor":      "#222222",
    "axes.titlepad":        10,
    "legend.frameon":       True,
    "legend.framealpha":    0.9,
    "legend.edgecolor":     "#DDDDDD",
})

BASE_DIR   = os.environ.get("OV_BASE_DIR", "/Users/gricey/Desktop/Internship/orthogonal_validation")
RMATS_DIR  = os.path.join(BASE_DIR, "results", "GSE130613", "SRP194523_rmats_2026-07-20")
FIG_DIR    = os.path.join(BASE_DIR, "results", "GSE130613", "figures")
TABLE_DIR  = os.path.join(BASE_DIR, "results", "GSE130613", "tables")
os.makedirs(FIG_DIR, exist_ok=True)
os.makedirs(TABLE_DIR, exist_ok=True)

DPSI_THRESH = 0.1
PVAL_THRESH = 0.3
FDR_THRESH  = 0.05

# rMATS event-type name -> SUPPA-style short code, reuses the same EVENT_COLORS palette
EVENT_TYPE_MAP = {
    "SE":   "SE",
    "A5SS": "A5",
    "A3SS": "A3",
    "MXE":  "MX",
    "RI":   "RI",
}
EVENT_COLORS = {
    "A3": "#E69F00",
    "A5": "#56B4E9",
    "MX": "#0072B2",
    "RI": "#D55E00",
    "SE": "#CC79A7",
}

print("=== Loading rMATS JC output ===")
dfs = []
for rmats_type, short_code in EVENT_TYPE_MAP.items():
    path = os.path.join(RMATS_DIR, f"{rmats_type}.MATS.JC.txt")
    df = pd.read_csv(path, sep="\t")
    df["GeneID"] = df["GeneID"].str.strip('"')
    df["geneSymbol"] = df["geneSymbol"].str.strip('"')
    df["event_type"] = short_code
    df["rmats_type"] = rmats_type
    dfs.append(df[["event_type", "rmats_type", "GeneID", "geneSymbol", "chr", "strand",
                    "PValue", "FDR", "IncLevelDifference"]])
    print(f"  {rmats_type}: {len(df)} testable events")

combined = pd.concat(dfs, ignore_index=True)

# Sign flip: rMATS b1(WT)-b2(KO) -> project convention KO-WT
combined["dPSI_KO_vs_WT"] = -combined["IncLevelDifference"]
combined = combined.rename(columns={"IncLevelDifference": "IncLevelDifference_WT_minus_KO"})

combined["sig_p0.3"] = (combined["PValue"] < PVAL_THRESH) & (combined["dPSI_KO_vs_WT"].abs() > DPSI_THRESH)
combined["sig_FDR0.05"] = (combined["FDR"] < FDR_THRESH) & (combined["dPSI_KO_vs_WT"].abs() > DPSI_THRESH)

print(f"\nTotal testable events (all 5 types): {len(combined)}")
print(f"Significant, p<0.3 & |dPSI|>0.1   : {combined['sig_p0.3'].sum()}")
print(f"Significant, FDR<0.05 & |dPSI|>0.1: {combined['sig_FDR0.05'].sum()}")

def make_volcano(df, pcol, thresh_val, sig_col, ylabel, title_suffix, out_name):
    fig, ax = plt.subplots(figsize=(8, 6))
    plot_df = df.copy()
    plot_df = plot_df[plot_df[pcol] > 0]  # avoid log(0)
    plot_df["neg_log10"] = -np.log10(plot_df[pcol])

    ns = plot_df[~plot_df[sig_col]]
    ax.scatter(ns["dPSI_KO_vs_WT"], ns["neg_log10"],
               color="lightgrey", alpha=0.4, s=15, linewidths=0, zorder=1)

    sig = plot_df[plot_df[sig_col]]
    for etype, color in EVENT_COLORS.items():
        sub = sig[sig["event_type"] == etype]
        if len(sub) > 0:
            ax.scatter(sub["dPSI_KO_vs_WT"], sub["neg_log10"],
                       color=color, alpha=0.9, s=40, linewidths=0,
                       label=f"{etype} (n={len(sub)})", zorder=2)

    ax.axvline(x=DPSI_THRESH, color="black", linestyle="--", linewidth=0.8, alpha=0.6)
    ax.axvline(x=-DPSI_THRESH, color="black", linestyle="--", linewidth=0.8, alpha=0.6)
    ax.axhline(y=-np.log10(thresh_val), color="black", linestyle="--", linewidth=0.8, alpha=0.6)

    ax.set_xlabel("dPSI (PerDKO − WT)", fontsize=12)
    ax.set_ylabel(ylabel, fontsize=12)
    ax.set_title(f"Differentially spliced events — rMATS, Illumina only (PerDKO vs WT, CT16-20, Liver)\n"
                 f"GSE130613 — STAR BAMs, {title_suffix}", fontsize=11)
    ax.legend(title="Event type", bbox_to_anchor=(1.02, 1), loc="upper left", fontsize=9)
    plt.tight_layout()

    for ext in ("svg", "png"):
        out_path = os.path.join(FIG_DIR, f"{out_name}.{ext}")
        plt.savefig(out_path, dpi=150, bbox_inches="tight")
        print(f"Saved: {out_path}")
    plt.close()

make_volcano(combined, "PValue", PVAL_THRESH, "sig_p0.3",
             "−log₁₀(p-value)", "p<0.3 (old SUPPA convention)",
             "GSE130613_rmats_dPSI_volcano_p0.3")

make_volcano(combined, "FDR", FDR_THRESH, "sig_FDR0.05",
             "−log₁₀(FDR)", "FDR<0.05 (stricter)",
             "GSE130613_rmats_dPSI_volcano_FDR0.05")

# live COUNTIFS summary (matches verify_final_numbers.xlsx convention)
print("\n=== Building xlsx ===")
xlsx_path = os.path.join(TABLE_DIR, "GSE130613_rmats_splicing_results.xlsx")

wb = openpyxl.Workbook()

ws_all = wb.active
ws_all.title = "All_events"
headers = ["event_type", "rmats_type", "GeneID", "geneSymbol", "chr", "strand",
           "PValue", "FDR", "dPSI_KO_vs_WT", "IncLevelDifference_WT_minus_KO",
           "sig_p0.3", "sig_FDR0.05"]
ws_all.append(headers)
for cell in ws_all[1]:
    cell.font = Font(name="Arial", bold=True)

sort_cols = ["event_type", "PValue"]
combined_sorted = combined.sort_values(sort_cols)[
    ["event_type", "rmats_type", "GeneID", "geneSymbol", "chr", "strand",
     "PValue", "FDR", "dPSI_KO_vs_WT", "IncLevelDifference_WT_minus_KO"]
].reset_index(drop=True)

n_rows = len(combined_sorted)
for i, row in combined_sorted.iterrows():
    r = i + 2  # 1-indexed, +1 for header
    ws_all.cell(r, 1, row["event_type"])
    ws_all.cell(r, 2, row["rmats_type"])
    ws_all.cell(r, 3, row["GeneID"])
    ws_all.cell(r, 4, row["geneSymbol"])
    ws_all.cell(r, 5, row["chr"])
    ws_all.cell(r, 6, row["strand"])
    ws_all.cell(r, 7, float(row["PValue"]))
    ws_all.cell(r, 8, float(row["FDR"]))
    ws_all.cell(r, 9, float(row["dPSI_KO_vs_WT"]))
    ws_all.cell(r, 10, float(row["IncLevelDifference_WT_minus_KO"]))
    ws_all.cell(r, 11, f"=AND(G{r}<0.3,ABS(I{r})>0.1)")
    ws_all.cell(r, 12, f"=AND(H{r}<0.05,ABS(I{r})>0.1)")

for col_idx in range(1, len(headers) + 1):
    ws_all.column_dimensions[get_column_letter(col_idx)].width = 16
for row in ws_all.iter_rows(min_row=2, max_row=n_rows + 1):
    for cell in row:
        cell.font = Font(name="Arial")

ws_sum = wb.create_sheet("Summary")
ws_sum.append(["Event type", "Testable events", "Sig (p<0.3, |dPSI|>0.1)", "Sig (FDR<0.05, |dPSI|>0.1)"])
for cell in ws_sum[1]:
    cell.font = Font(name="Arial", bold=True)

last_row = n_rows + 1  # last data row on All_events
event_codes = ["A3", "A5", "MX", "RI", "SE"]
for i, code in enumerate(event_codes):
    r = i + 2
    ws_sum.cell(r, 1, code)
    ws_sum.cell(r, 2, f'=COUNTIF(All_events!A2:A{last_row},"{code}")')
    ws_sum.cell(r, 3, f'=COUNTIFS(All_events!A2:A{last_row},"{code}",All_events!K2:K{last_row},TRUE)')
    ws_sum.cell(r, 4, f'=COUNTIFS(All_events!A2:A{last_row},"{code}",All_events!L2:L{last_row},TRUE)')

total_row = len(event_codes) + 2
ws_sum.cell(total_row, 1, "Total")
ws_sum.cell(total_row, 1).font = Font(name="Arial", bold=True)
for col in (2, 3, 4):
    col_letter = get_column_letter(col)
    ws_sum.cell(total_row, col, f"=SUM({col_letter}2:{col_letter}{total_row - 1})")
    ws_sum.cell(total_row, col).font = Font(name="Arial", bold=True)

for row in ws_sum.iter_rows(min_row=2, max_row=len(event_codes) + 1):
    for cell in row:
        if cell.column > 1:
            cell.font = Font(name="Arial")
for col_idx in range(1, 5):
    ws_sum.column_dimensions[get_column_letter(col_idx)].width = 24

note_row = total_row + 2
ws_sum.cell(note_row, 1,
    "Old SUPPA (Salmon-based) result at p<0.3, |dPSI|>0.1: 513 significant events, "
    "7 event types (incl. AF/AL, not covered by rMATS). rMATS b1=WT, b2=KO; "
    "dPSI_KO_vs_WT = -(rMATS IncLevelDifference) to match project convention.")
ws_sum.cell(note_row, 1).font = Font(name="Arial", italic=True, size=9)
ws_sum.cell(note_row, 1).alignment = Alignment(wrap_text=True)

wb.save(xlsx_path)
print(f"Saved: {xlsx_path}")
print("\n=== DONE ===")
