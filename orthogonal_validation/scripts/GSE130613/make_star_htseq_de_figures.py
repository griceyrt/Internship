#!/usr/bin/env python3
"""
GSE130613 STAR+htseq-count DE figures (final adopted pipeline) -- same two
figures as the original Salmon pipeline's step13 (DE volcano) and step14
(gene-level DE concordance vs Nanopore), rebuilt on STAR+htseq-count
(original GTF) instead of Salmon+tximport+stageR. No step15 equivalent:
htseq-count is gene-level only, no transcript resolution. Thresholds
unchanged: volcano padj<0.05/|log2FC|>0.5, concordance padj<0.05/|log2FC|>1.5.
Author: Gricey

Run from: orthogonal_validation/
"""

import os
import numpy as np
import pandas as pd
import openpyxl
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D
from adjustText import adjust_text

# same style block as plot_results.py / crossreference.py
matplotlib.rcParams.update({
    "font.family":          "sans-serif",
    "font.sans-serif":      ["Arial", "DejaVu Sans", "Helvetica", "Verdana"],
    "font.size":            11,
    "axes.spines.top":      False,
    "axes.spines.right":    False,
    "axes.edgecolor":       "#AAAAAA",
    "axes.linewidth":       0.8,
    "figure.facecolor":     "white",
    "axes.facecolor":       "white",
})

BASE_DIR   = os.environ.get("OV_BASE_DIR", "/Users/gricey/Desktop/Internship/orthogonal_validation")
DESEQ_FILE = os.path.join(BASE_DIR, "results", "GSE130613", "normalisation_star_htseq_original_gtf", "deseq2_star_htseq_KO_vs_WT.tsv")
NANO_FILE  = os.path.join(BASE_DIR, "data", "Table3-differential_expressiong_WTvsPerKO.xlsx")
FIG_DIR    = os.path.join(BASE_DIR, "results", "GSE130613", "figures")
TABLE_DIR  = os.path.join(BASE_DIR, "results", "GSE130613", "tables")
os.makedirs(FIG_DIR, exist_ok=True)
os.makedirs(TABLE_DIR, exist_ok=True)

CIRCADIAN_GENES = [
    "Ren1", "Orm2", "Prtn3", "Cyp2b9", "Dbp", "Cxcl1", "Saa1", "Ighg2b",
    "Celf4", "Egr1", "Ciart", "Nr1d1", "Nr1d2", "Btg2", "Per3", "Ephx1",
    "Gstm2", "Clock", "Cyp2a5", "Arntl", "Cry1", "Kmt2d", "Gstm3",
    "Prkca", "Cyp7a1", "Fancm", "Cspg5"
]

print("=== Loading STAR+htseq DE result ===")
de = pd.read_csv(DESEQ_FILE, sep="\t")
de = de.dropna(subset=["log2FoldChange", "padj"])
de = de[de["padj"] > 0].set_index("gene_id")
print(f"Genes with a reliable (non-NaN) padj: {len(de)}")
print(f"log2FC range: {de['log2FoldChange'].min():.2f} to {de['log2FoldChange'].max():.2f}")

print("\n=== Volcano (step13-style) ===")
LFC_V, PADJ_V = 0.5, 0.05

de["neg_log10_padj"] = -np.log10(de["padj"])
de["category"] = "not significant"
de.loc[(de["log2FoldChange"] >  LFC_V) & (de["padj"] < PADJ_V), "category"] = "up in PerDKO"
de.loc[(de["log2FoldChange"] < -LFC_V) & (de["padj"] < PADJ_V), "category"] = "down in PerDKO"

for cat, n in de["category"].value_counts().items():
    print(f"  {cat}: {n}")

cat_colors = {"not significant": "#CCCCCC", "up in PerDKO": "#D55E00", "down in PerDKO": "#0072B2"}

fig, ax = plt.subplots(figsize=(10, 7))
ax.grid(False)
for cat, color in cat_colors.items():
    sub = de[de["category"] == cat]
    alpha = 0.85 if cat != "not significant" else 0.3
    size = 22 if cat != "not significant" else 6
    ax.scatter(sub["log2FoldChange"], sub["neg_log10_padj"], color=color, alpha=alpha,
               s=size, linewidths=0, label=f"{cat} (n={len(sub)})",
               zorder=2 if cat != "not significant" else 1)

ax.axvline(x=LFC_V, color="#BBBBBB", linestyle="--", linewidth=0.8)
ax.axvline(x=-LFC_V, color="#BBBBBB", linestyle="--", linewidth=0.8)
ax.axhline(y=-np.log10(PADJ_V), color="#BBBBBB", linestyle="--", linewidth=0.8)

# Label Kiran's curated circadian gene list, only where significant
sig = de[de["category"] != "not significant"]
texts = []
labeled = []
for gname in CIRCADIAN_GENES:
    matches = sig[sig["gene_name"] == gname] if "gene_name" in sig.columns else sig.loc[[i for i in sig.index if False]]
    if len(matches) == 0:
        continue
    row = matches.iloc[0]
    texts.append(ax.text(row["log2FoldChange"], row["neg_log10_padj"], gname,
                          fontsize=8, color="#1a1a1a", fontstyle="italic"))
    labeled.append((gname, row["log2FoldChange"], row["neg_log10_padj"]))

if texts:
    adjust_text(texts, ax=ax, expand=(2.5, 2.5), force_text=(1.5, 1.5), force_points=(1.2, 1.2))
    for txt, (gname, dx, dy) in zip(texts, labeled):
        lx, ly = txt.get_position()
        txt.remove()
        ax.annotate(gname, xy=(dx, dy), xytext=(lx, ly), fontsize=8, color="#1a1a1a",
                    fontstyle="italic", ha="center", va="center",
                    arrowprops=dict(arrowstyle="-", color="#888888", lw=0.6, shrinkA=3, shrinkB=3))

print(f"  Circadian genes labeled (significant + present): {len(labeled)}")

ax.set_xlabel("log₂ fold change (PerDKO / WT)", fontsize=12)
ax.set_ylabel("−log₁₀(padj)", fontsize=12)
ax.set_title("DE genes — STAR + htseq-count (adopted pipeline), Illumina only\n"
             f"GSE130613, PerDKO vs WT, CT16-20, Liver | n={len(de)} genes tested", fontsize=11)
ax.legend(fontsize=9, frameon=True, framealpha=0.9)
plt.tight_layout()
for ext in ("svg", "png"):
    p = os.path.join(FIG_DIR, f"GSE130613_STARhtseq_DE_volcano.{ext}")
    plt.savefig(p, dpi=150, bbox_inches="tight")
    print(f"Saved: {p}")
plt.close()

print("\n=== Concordance vs Nanopore (step14-style) ===")
LFC_C = 1.5

wb = openpyxl.load_workbook(NANO_FILE)

ws_sig = wb["perKO_sig"]
rows_sig = list(ws_sig.iter_rows(values_only=True))
header_sig = [str(c) if c is not None else "idx" for c in rows_sig[0]]
nano_sig_geneid = {}
for r in rows_sig[1:]:
    row = dict(zip(header_sig, r))
    gid = row.get("gene_id")
    if gid and gid not in nano_sig_geneid:
        try:
            nano_sig_geneid[gid] = float(row["log2FC_perko"])
        except (TypeError, ValueError):
            pass

ws_all = wb["DE_all"]
rows_all = list(ws_all.iter_rows(values_only=True))
header_all = list(rows_all[0])
nano_all_geneid = {}
for r in rows_all[1:]:
    row = dict(zip(header_all, r))
    gid = row.get("gene_id")
    if not gid:
        continue
    try:
        lfc = float(row["log2FC_perko"])
        padj = float(row["padj_per_ko"])
    except (TypeError, ValueError):
        continue
    if gid not in nano_all_geneid or abs(lfc) > abs(nano_all_geneid[gid]["log2FC"]):
        nano_all_geneid[gid] = {"log2FC": lfc, "padj": padj}

ill_sig = set(de[(de["padj"] < PADJ_V) & (de["log2FoldChange"].abs() > LFC_C)].index)
common = set(de.index) & set(nano_all_geneid.keys())
nano_sig = set(nano_sig_geneid.keys())
overlap = ill_sig & nano_sig & common

print(f"  Common genes (both platforms tested)      : {len(common)}")
print(f"  Illumina sig (padj<0.05, |log2FC|>1.5)    : {len(ill_sig)}")
print(f"  Nanopore sig (perKO_sig)                  : {len(nano_sig)}")
print(f"  Overlap (both significant)                : {len(overlap)}")

genes = list(common)
x = np.array([float(de.loc[g, "log2FoldChange"]) for g in genes])
y = np.array([float(nano_all_geneid[g]["log2FC"]) for g in genes])
mask_both = np.array([g in overlap for g in genes])
mask_ill = np.array([g in ill_sig and g not in overlap for g in genes])
mask_nano = np.array([g in nano_sig and g not in overlap for g in genes])
mask_grey = ~(mask_both | mask_ill | mask_nano)

fig, ax = plt.subplots(figsize=(10, 8))
ax.grid(False)
ax.scatter(x[mask_grey], y[mask_grey], c="#CCCCCC", s=6, alpha=0.35, linewidths=0, zorder=1, rasterized=True)
ax.scatter(x[mask_ill], y[mask_ill], c="#56B4E9", s=40, alpha=0.85, linewidths=0, zorder=2)
ax.scatter(x[mask_nano], y[mask_nano], c="#009E73", s=40, alpha=0.85, linewidths=0, zorder=3)
ax.scatter(x[mask_both], y[mask_both], c="#E69F00", s=55, alpha=0.95, edgecolors="black", linewidths=0.7, zorder=4)

lim = max(abs(x).max(), abs(y).max()) * 1.05
ax.plot([-lim, lim], [-lim, lim], color="#888888", lw=0.9, ls="--", alpha=0.6)
ax.axhline(0, color="#BBBBBB", lw=0.6, alpha=0.5)
ax.axvline(0, color="#BBBBBB", lw=0.6, alpha=0.5)

texts = []
labels_out = []
gene_map_ht = de["gene_name"].to_dict() if "gene_name" in de.columns else {}
for i, g in enumerate(genes):
    if not mask_both[i]:
        continue
    gname = gene_map_ht.get(g, g)
    if pd.isna(gname) or gname == "":
        gname = g  # ~20% of genes in this GTF have no gene_name attribute, fall back to gene_id
    offset_x = 0.6 if x[i] >= 0 else -0.6
    offset_y = 0.4 if y[i] >= 0 else -0.4
    texts.append(ax.text(x[i] + offset_x, y[i] + offset_y, gname, fontsize=8, color="#1a1a1a", fontstyle="italic"))
    labels_out.append((gname, x[i], y[i]))

if texts:
    adjust_text(texts, ax=ax, expand=(3.0, 3.0), force_text=(2.0, 2.0), force_points=(1.5, 1.5), lim=700)
    for txt, (gname, dx, dy) in zip(texts, labels_out):
        lx, ly = txt.get_position()
        txt.remove()
        ax.annotate(gname, xy=(dx, dy), xytext=(lx, ly), fontsize=8, color="#1a1a1a",
                    fontstyle="italic", ha="center", va="center",
                    arrowprops=dict(arrowstyle="-", color="#888888", lw=0.7, shrinkA=4, shrinkB=4))

legend = [
    Line2D([0], [0], marker="o", color="w", markerfacecolor="#E69F00", markeredgecolor="black",
           markersize=8, label=f"Both significant (n={mask_both.sum()})"),
    Line2D([0], [0], marker="o", color="w", markerfacecolor="#56B4E9", markersize=7,
           label=f"Illumina only (n={mask_ill.sum()})"),
    Line2D([0], [0], marker="o", color="w", markerfacecolor="#009E73", markersize=7,
           label=f"Nanopore only (n={mask_nano.sum()})"),
    Line2D([0], [0], marker="o", color="w", markerfacecolor="#CCCCCC", markersize=6,
           label=f"Not significant (n={mask_grey.sum()})"),
]
ax.legend(handles=legend, fontsize=9, bbox_to_anchor=(1.02, 1), loc="upper left")
ax.set_xlabel("log₂ fold change — Illumina, STAR+htseq (PerDKO / WT)", fontsize=12)
ax.set_ylabel("log₂ fold change — Nanopore (PerDKO / WT)", fontsize=12)
ax.set_title("DE concordance: Illumina (STAR+htseq) vs Nanopore — gene level\n"
             f"All genes tested in both platforms (n={len(common)}) | threshold: padj<0.05, |log₂FC|>1.5",
             fontsize=11)
plt.tight_layout()
for ext in ("svg", "png"):
    p = os.path.join(FIG_DIR, f"GSE130613_STARhtseq_vs_Nanopore_DE_concordance_gene_level.{ext}")
    plt.savefig(p, dpi=150, bbox_inches="tight")
    print(f"Saved: {p}")
plt.close()

# =============================================================================
# TABLE — overlap genes, both platforms' numbers side by side
# =============================================================================
print("\n=== Building overlap table ===")

rows_out = []
for g in sorted(overlap):
    rows_out.append({
        "gene_id": g,
        "gene_name": gene_map_ht.get(g, ""),
        "log2FC_Illumina_STARhtseq": de.loc[g, "log2FoldChange"],
        "padj_Illumina_STARhtseq": de.loc[g, "padj"],
        "log2FC_Nanopore": nano_all_geneid[g]["log2FC"],
        "padj_Nanopore": nano_all_geneid[g]["padj"],
        "same_direction": (de.loc[g, "log2FoldChange"] > 0) == (nano_all_geneid[g]["log2FC"] > 0),
    })
overlap_df = pd.DataFrame(rows_out).sort_values("padj_Illumina_STARhtseq")
same_dir_n = overlap_df["same_direction"].sum()
print(f"  Overlap genes: {len(overlap_df)}, same direction: {same_dir_n}, opposite: {len(overlap_df) - same_dir_n}")
xlsx_path = os.path.join(TABLE_DIR, "GSE130613_STARhtseq_vs_Nanopore_DE_overlap.xlsx")
with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
    overlap_df.to_excel(writer, sheet_name="Overlap_genes", index=False)
    summary = pd.DataFrame({
        "Metric": ["Common genes tested (both platforms)", "Illumina sig (padj<0.05, |log2FC|>1.5)",
                   "Nanopore sig (perKO_sig)", "Overlap (both sig)", "Same direction", "Opposite direction"],
        "Value": [len(common), len(ill_sig), len(nano_sig), len(overlap), int(same_dir_n), int(len(overlap_df) - same_dir_n)],
    })
    summary.to_excel(writer, sheet_name="Summary", index=False)

from openpyxl import load_workbook
wb_out = load_workbook(xlsx_path)
for ws_name in wb_out.sheetnames:
    ws_out = wb_out[ws_name]
    for cell in ws_out[1]:
        cell.font = openpyxl.styles.Font(name="Arial", bold=True)
    for row in ws_out.iter_rows(min_row=2):
        for cell in row:
            cell.font = openpyxl.styles.Font(name="Arial")
    for col_idx, col_cells in enumerate(ws_out.columns, start=1):
        ws_out.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 22
wb_out.save(xlsx_path)
print(f"Saved: {xlsx_path}")

print("\n=== DONE ===")
