#!/usr/bin/env python3
"""
GSE133398 (SRSF3-KO hepatocytes, Nick Webster lab) — final xlsx tables for the
reviewer rebuttal. Companion to scripts/GSE133398_make_final_figures.py.

IMPORTANT — relationship between figures and tables: they are siblings, not
dependents. Neither is generated from the other; both read directly from the
same upstream pipeline outputs (DESeq2 tsvs, SUPPA diff files, Nanopore
comparison files) and apply the same thresholds independently in their own
code. See figure/table headers for the shared threshold conventions
(padj<0.05, |log2FC|>0.5 for DE; |dPSI|>0.1, p<0.05 for splicing).

Produces 3 xlsx files:
  - GSE133398_DE_results.xlsx                  (3 tabs: Option A genes, Option B genes, Option B transcripts)
  - GSE133398_vs_Nanopore_overlap_summary.xlsx (4 tabs: clock genes, DE overlap A, DE overlap B, splicing overlap)
  - GSE133398_significant_splicing_events.xlsx (1 tab: 84 significant splicing events)

Inputs (same as the figures script):
  - results/normalisation_GSE133398/deseq2_KO_vs_WT.tsv
  - results/normalisation_GSE133398_stageR/deseq2_stageR_KO_vs_WT.tsv
  - results/suppa_GSE133398_2026-07-13/diff/diff_<TYPE>_strict.dpsi.temp.0
  - data/nanopore_suppa/res_<TYPE>.dpsi.temp.0
  - data/Table3-differential_expressiong_WTvsPerKO.xlsx (perKO_sig sheet)
  - boundary_analysis/data/transcriptome_productivity.gtf

Run from: orthogonal_validation/
    python3 scripts/GSE133398_make_final_tables.py

NOTE on Nanopore gene-level dedup: a small number of genes have more than one
significant Nanopore transcript (65 of 267 rows in perKO_sig share a gene_id
with another row). This script picks the transcript with the lowest
padj_per_ko as the gene's representative value -- the same "most significant
transcript represents the gene" convention used for Option B (stageR)
elsewhere in this pipeline. This is a clean, documented choice, not a
guaranteed match to the original ad hoc table-building scripts (now lost --
they lived only in /tmp and were never saved to scripts/), which appear to
have picked inconsistently across genes in a way that couldn't be reverse
-engineered from the data alone. Affects ~8 of 56 total overlap rows -- see
the email/commit notes for the exact diff if reconciling against a
previously-sent version matters.
"""

import os, re, glob
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# =============================================================================
# PATHS (relative to this script's location)
# =============================================================================
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
INTERNSHIP = os.path.dirname(BASE)
GTF = os.path.join(INTERNSHIP, "boundary_analysis", "data", "transcriptome_productivity.gtf")
SUPPA_DIR = os.path.join(BASE, "results", "suppa_GSE133398_2026-07-13", "diff")
NAN_SUPPA_DIR = os.path.join(BASE, "data", "nanopore_suppa")
NAN_DE_PATH = os.path.join(BASE, "data", "Table3-differential_expressiong_WTvsPerKO.xlsx")
OUT = os.path.join(BASE, "results", "tables_GSE133398")
os.makedirs(OUT, exist_ok=True)

LFC_THRESH = 0.5
PADJ_THRESH = 0.05
DPSI_THRESH = 0.1
SPLICE_P_THRESH = 0.05

CLOCK_GENES = ["Per1", "Per2", "Per3", "Cry1", "Cry2", "Clock", "Arntl", "Arntl2",
               "Nr1d1", "Nr1d2", "Rora", "Rorb", "Rorc", "Dbp", "Nfil3",
               "Csnk1d", "Csnk1e", "Bhlhe40", "Bhlhe41", "Ciart"]

NAVY = "FF1B2A4A"


def build_gene_map(gtf_path):
    gmap = {}
    with open(gtf_path) as fh:
        for line in fh:
            if "\tgene\t" not in line:
                continue
            gid = re.search(r'gene_id "([^"]+)"', line)
            gname = re.search(r'gene_name "([^"]+)"', line)
            if gid and gname:
                gmap[gid.group(1)] = gname.group(1)
    return gmap


def load_suppa_dir(dirpath, pattern):
    dfs = []
    for f in sorted(glob.glob(os.path.join(dirpath, pattern))):
        base = os.path.basename(f)
        etype = re.sub(r"^(diff_|res_)", "", base)
        etype = re.sub(r"(_strict)?\.dpsi\.temp\.0$", "", etype)
        df = pd.read_csv(f, sep="\t", index_col=0, header=0)
        df.columns = ["dPSI", "pval"]
        df["event_type"] = etype
        df = df.dropna().query("pval > 0")
        dfs.append(df)
    return pd.concat(dfs)


def write_sheet(wb, title, note, headers, data_rows, col_widths=None, first=False):
    ws = wb.active if first and wb.active.title == "Sheet" else wb.create_sheet(title)
    ws.title = title
    ncols = len(headers)

    ws.cell(1, 1, note)
    ws.cell(1, 1).font = Font(name="Arial", size=10, color="FF666666")
    ws.cell(1, 1).alignment = Alignment(vertical="top", wrap_text=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.row_dimensions[1].height = 30

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(2, c, h)
        cell.font = Font(name="Arial", bold=True, color="FFFFFFFF", size=11)
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor=NAVY)

    for r, row in enumerate(data_rows, start=3):
        for c, val in enumerate(row, start=1):
            v = None if (isinstance(val, float) and pd.isna(val)) else val
            ws.cell(r, c, v)
            ws.cell(r, c).font = Font(name="Arial", size=10)

    ws.freeze_panes = "A3"
    widths = col_widths or [18] * ncols
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    return ws


print("Loading gene map...")
gene_map = build_gene_map(GTF)
gene_map_rev = {}
for gid, gname in gene_map.items():
    gene_map_rev.setdefault(gname, gid)  # first-seen mapping name -> id

# =============================================================================
# Load DE (Option A, gene-level)
# =============================================================================
print("Loading Option A (gene-level DESeq2)...")
de = pd.read_csv(os.path.join(BASE, "results", "normalisation_GSE133398", "deseq2_KO_vs_WT.tsv"),
                  sep="\t", index_col=0)
de.index.name = "gene_id"
de = de.dropna(subset=["log2FoldChange", "padj"])
de["gene_name"] = de.index.map(gene_map)
de_sig = de[(de.padj < PADJ_THRESH) & (de.log2FoldChange.abs() > LFC_THRESH)].sort_values("padj")
n_unannotated = sum(1 for i in de_sig.index if not re.match(r"^ENSMUSG", str(i)))
print(f"  {len(de_sig)} significant genes ({n_unannotated} unannotated loci)")

# =============================================================================
# Load DE (Option B, transcript-level + stageR)
# =============================================================================
print("Loading Option B (stageR)...")
deb = pd.read_csv(os.path.join(BASE, "results", "normalisation_GSE133398_stageR", "deseq2_stageR_KO_vs_WT.tsv"),
                   sep="\t")
deb_valid = deb.dropna(subset=["log2FoldChange", "padj_gene_stageR"])
deb_valid = deb_valid[deb_valid.padj_gene_stageR >= 0]

deb_b_genes_sig = (deb_valid[(deb_valid.padj_gene_stageR < PADJ_THRESH) & (deb_valid.log2FoldChange.abs() > LFC_THRESH)]
                    .assign(abslfc=lambda d: d.log2FoldChange.abs())
                    .sort_values(["padj_gene_stageR", "abslfc"], ascending=[True, False])
                    .drop_duplicates("gene_id"))
deb_b_genes_sig = deb_b_genes_sig.sort_values("padj_gene_stageR")
print(f"  {len(deb_b_genes_sig)} significant genes (stage 1)")

deb_b_tx_sig = deb.dropna(subset=["padj_tx_stageR"])
deb_b_tx_sig = deb_b_tx_sig[deb_b_tx_sig.padj_tx_stageR < PADJ_THRESH].sort_values("padj_tx_stageR")
print(f"  {len(deb_b_tx_sig)} significant transcripts (stage 2)")

# for the clock-gene table: mis_expressed_OptionB uses padj_gene_stageR<0.05 only (no lfc filter),
# matching the existing deliverable's note text
deb_b_screen = deb_valid[deb_valid.padj_gene_stageR < PADJ_THRESH].assign(abslfc=lambda d: d.log2FoldChange.abs()) \
                 .sort_values(["padj_gene_stageR", "abslfc"], ascending=[True, False]).drop_duplicates("gene_id")

# =============================================================================
# Load splicing (SRSF3-KO)
# =============================================================================
print("Loading SRSF3-KO splicing (SUPPA)...")
dpsi = load_suppa_dir(SUPPA_DIR, "diff_*_strict.dpsi.temp.0")
dpsi["dPSI"] = -dpsi["dPSI"]
dpsi_sig = dpsi[(dpsi.dPSI.abs() > DPSI_THRESH) & (dpsi.pval < SPLICE_P_THRESH)].copy()
dpsi_sig["gene_id"] = dpsi_sig.index.str.split(";").str[0]
dpsi_sig["gene_name"] = dpsi_sig["gene_id"].map(gene_map)
dpsi_sig = dpsi_sig.sort_values("pval")
from collections import Counter
type_counts = Counter(dpsi_sig["event_type"])
print(f"  {len(dpsi_sig)} significant events. Breakdown: {dict(type_counts)}")

srsf3_splice_by_gene = dpsi_sig.groupby("gene_id").agg(
    event_types=("event_type", lambda s: ", ".join(sorted(s))),
    n_events=("event_type", "size"),
).reset_index()

# =============================================================================
# Load Nanopore DE (perKO_sig sheet)
# =============================================================================
print("Loading Nanopore DE...")
wb_nan = openpyxl.load_workbook(NAN_DE_PATH, read_only=True, data_only=True)
ws_nan = wb_nan["perKO_sig"]
nan_rows = list(ws_nan.iter_rows(values_only=True))
nan_cols = list(nan_rows[0])[1:]  # drop leading unlabeled index column
nan_data = [list(r)[1:] for r in nan_rows[1:]]
nan_de_all = pd.DataFrame(nan_data, columns=nan_cols)
nan_de_all["log2FC_perko"] = pd.to_numeric(nan_de_all["log2FC_perko"], errors="coerce")
nan_de_all["padj_per_ko"] = pd.to_numeric(nan_de_all["padj_per_ko"], errors="coerce")
# gene-level representative: lowest padj_per_ko per gene_id (documented convention, see module docstring)
nan_de_gene = nan_de_all.dropna(subset=["padj_per_ko"]).sort_values("padj_per_ko").drop_duplicates("gene_id")
nan_de_gene = nan_de_gene.set_index("gene_id")
nan_de_genes = set(nan_de_gene.index)
print(f"  {len(nan_de_genes)} unique Nanopore DE genes")

# =============================================================================
# Load Nanopore splicing
# =============================================================================
print("Loading Nanopore splicing...")
nan_splice = load_suppa_dir(NAN_SUPPA_DIR, "res_*.dpsi.temp.0")
nan_splice_sig = nan_splice[(nan_splice.dPSI.abs() > DPSI_THRESH) & (nan_splice.pval < SPLICE_P_THRESH)].copy()
nan_splice_sig["gene_id"] = nan_splice_sig.index.str.split(";").str[0]
nan_splice_by_gene = nan_splice_sig.groupby("gene_id").agg(
    event_types=("event_type", lambda s: ", ".join(sorted(s))),
    n_events=("event_type", "size"),
).reset_index().set_index("gene_id")
nan_splice_genes = set(nan_splice_by_gene.index)
print(f"  {len(nan_splice_genes)} unique Nanopore splicing genes")

srsf3_de_a_genes = set(de_sig.index)
srsf3_de_b_genes = set(deb_b_genes_sig.gene_id)
srsf3_splice_genes = set(dpsi_sig.gene_id)

# =============================================================================
# FILE 1: GSE133398_DE_results.xlsx
# =============================================================================
print("\n=== Building GSE133398_DE_results.xlsx ===")
wb1 = openpyxl.Workbook()

rows_a = [(gid, row.gene_name, row.baseMean, row.log2FoldChange, row.lfcSE, row.stat, row.pvalue, row.padj)
          for gid, row in de_sig.iterrows()]
write_sheet(wb1, "Option A - genes",
            f"Note: {n_unannotated} of {len(de_sig)} gene_id values are shown as chromosome coordinates "
            f"(e.g. 1:170970000) instead of a standard ENSMUSG ID. These are unannotated loci from the "
            f"custom transcriptome.",
            ["gene_id", "gene_name", "baseMean", "log2FoldChange", "lfcSE", "stat", "pvalue", "padj"],
            rows_a, col_widths=[27, 15, 18, 19, 15, 15, 23, 23], first=True)

rows_b_genes = [(row.gene_id, row.gene_name, row.transcript_id, row.log2FoldChange, row.pvalue,
                  row.padj_gene_stageR) for _, row in deb_b_genes_sig.iterrows()]
write_sheet(wb1, "Option B - genes",
            f"Option B (stageR stage 1, gene-level screening): {len(deb_b_genes_sig)} significant genes "
            f"(padj_gene_stageR<{PADJ_THRESH}, |log2FC|>{LFC_THRESH}). One representative transcript shown "
            f"per gene (lowest padj_gene_stageR, then largest |log2FC|) — see the transcript-level table "
            f"for all stage-2-confirmed transcripts.",
            ["gene_id", "gene_name", "transcript_id", "log2FoldChange", "pvalue", "padj_gene_stageR"],
            rows_b_genes, col_widths=[20, 14, 20, 16, 20, 20])

rows_b_tx = [(row.transcript_id, row.gene_id, row.gene_name, row.log2FoldChange, row.pvalue,
              row.padj_gene_stageR, row.padj_tx_stageR) for _, row in deb_b_tx_sig.iterrows()]
write_sheet(wb1, "Option B - transcripts",
            f"Option B (stageR stage 2, transcript-level confirmation): {len(deb_b_tx_sig)} significant "
            f"transcripts (padj_tx_stageR<{PADJ_THRESH}) within genes that passed stage 1 screening.",
            ["transcript_id", "gene_id", "gene_name", "log2FoldChange", "pvalue", "padj_gene_stageR", "padj_tx_stageR"],
            rows_b_tx, col_widths=[20, 20, 14, 16, 20, 20, 20])

wb1.save(os.path.join(OUT, "GSE133398_DE_results.xlsx"))
print("  Saved.")

# =============================================================================
# FILE 2: GSE133398_vs_Nanopore_overlap_summary.xlsx
# =============================================================================
print("\n=== Building GSE133398_vs_Nanopore_overlap_summary.xlsx ===")
wb2 = openpyxl.Workbook()

# --- Clock genes ---
de_a_by_id = de.set_index(de.index) if de.index.name == "gene_id" else de
de_a_lookup = de.copy()
deb_b_lookup = deb_b_screen.set_index("gene_id")
clock_rows = []
for gname in CLOCK_GENES:
    gid = gene_map_rev.get(gname)
    a_hit = gid in srsf3_de_a_genes if gid else False
    a_lfc = a_padj = None
    if gid and gid in de_a_lookup.index:
        r = de_a_lookup.loc[gid]
        if isinstance(r, pd.DataFrame):
            r = r.iloc[0]
        if a_hit:
            a_lfc, a_padj = r.log2FoldChange, r.padj
    b_hit = gid in deb_b_lookup.index if gid else False
    b_lfc = b_padj = None
    if b_hit:
        r = deb_b_lookup.loc[gid]
        b_lfc, b_padj = r.log2FoldChange, r.padj_gene_stageR
    spliced = gid in srsf3_splice_genes if gid else False
    event_types = None
    if spliced:
        event_types = srsf3_splice_by_gene.set_index("gene_id").loc[gid, "event_types"]
    nan_de_hit = gid in nan_de_genes if gid else False
    nan_splice_hit = gid in nan_splice_genes if gid else False
    clock_rows.append((
        gname,
        "YES" if a_hit else "no", a_lfc, a_padj,
        "YES" if b_hit else "no", b_lfc, b_padj,
        "yes" if spliced else "no", event_types,
        "YES" if nan_de_hit else "no",
        "YES" if nan_splice_hit else "no",
    ))

write_sheet(wb2, "Clock genes",
            "20 canonical clock genes crossed against 4 datasets: SRSF3-KO DE Option A, SRSF3-KO DE Option "
            "B (stageR), SRSF3-KO splicing (p<0.05), and Nanopore's own DE + splicing significance. "
            "mis_expressed_OptionA/B: padj<0.05, |log2FC|>0.5 (A) or padj_gene_stageR<0.05 (B).",
            ["clock_gene", "mis_expressed_OptionA", "OptionA_log2FC", "OptionA_padj",
             "mis_expressed_OptionB", "OptionB_log2FC", "OptionB_padj_gene_stageR",
             "mis_spliced_in_SRSF3KO", "SRSF3KO_event_types",
             "also_sig_in_Nanopore_DE", "also_sig_in_Nanopore_splicing"],
            clock_rows, col_widths=[14, 18, 15, 15, 18, 15, 22, 20, 18, 20, 24], first=True)

# --- DE overlap - Option A ---
overlap_a = sorted(srsf3_de_a_genes & nan_de_genes, key=lambda g: de_a_lookup.loc[g, "padj"] if not isinstance(de_a_lookup.loc[g], pd.DataFrame) else de_a_lookup.loc[g].iloc[0].padj)
rows_ov_a = []
for gid in overlap_a:
    r = de_a_lookup.loc[gid]
    if isinstance(r, pd.DataFrame):
        r = r.iloc[0]
    nr = nan_de_gene.loc[gid]
    rows_ov_a.append((gid, r.gene_name, r.log2FoldChange, r.padj, nr.log2FC_perko, nr.padj_per_ko))
write_sheet(wb2, "DE overlap - Option A",
            f"{len(overlap_a)} genes significant in both Nanopore DE and SRSF3-KO Option A DE (gene-level).",
            ["gene_id", "gene_name", "SRSF3KO_log2FC", "SRSF3KO_padj", "Nanopore_log2FC_perko", "Nanopore_padj_perko"],
            rows_ov_a, col_widths=[20, 14, 18, 20, 22, 22])

# --- DE overlap - Option B ---
overlap_b = sorted(srsf3_de_b_genes & nan_de_genes, key=lambda g: deb_b_lookup.loc[g, "padj_gene_stageR"])
rows_ov_b = []
for gid in overlap_b:
    r = deb_b_lookup.loc[gid]
    nr = nan_de_gene.loc[gid]
    rows_ov_b.append((gid, r.gene_name, r.log2FoldChange, r.padj_gene_stageR, nr.log2FC_perko, nr.padj_per_ko))
write_sheet(wb2, "DE overlap - Option B",
            f"{len(overlap_b)} genes significant in both Nanopore DE and SRSF3-KO Option B DE (stageR gene screen).",
            ["gene_id", "gene_name", "SRSF3KO_OptionB_log2FC", "SRSF3KO_OptionB_padj_gene_stageR",
             "Nanopore_log2FC_perko", "Nanopore_padj_perko"],
            rows_ov_b, col_widths=[20, 14, 24, 30, 22, 22])

# --- Splicing overlap ---
overlap_splice = sorted(srsf3_splice_genes & nan_splice_genes)
rows_ov_s = []
srsf3_by_gene_idx = srsf3_splice_by_gene.set_index("gene_id")
for gid in overlap_splice:
    sr = srsf3_by_gene_idx.loc[gid]
    nr = nan_splice_by_gene.loc[gid]
    gname = gene_map.get(gid)
    rows_ov_s.append((gid, gname, sr.event_types, sr.n_events, nr.event_types, nr.n_events))
write_sheet(wb2, "Splicing overlap",
            f"{len(overlap_splice)} genes with a significant splicing event in both Nanopore and SRSF3-KO "
            f"(|dPSI|>{DPSI_THRESH}, p<{SPLICE_P_THRESH}). Not Option A/B specific -- splicing significance "
            f"doesn't depend on which DE approach is used.",
            ["gene_id", "gene_name", "SRSF3KO_event_types", "SRSF3KO_n_events",
             "Nanopore_event_types", "Nanopore_n_events"],
            rows_ov_s, col_widths=[20, 14, 20, 18, 20, 18])

wb2.save(os.path.join(OUT, "GSE133398_vs_Nanopore_overlap_summary.xlsx"))
print(f"  Saved. DE-A overlap={len(overlap_a)}, DE-B overlap={len(overlap_b)}, splicing overlap={len(overlap_splice)}")

# =============================================================================
# FILE 3: GSE133398_significant_splicing_events.xlsx
# =============================================================================
print("\n=== Building GSE133398_significant_splicing_events.xlsx ===")
wb3 = openpyxl.Workbook()
breakdown = ", ".join(f"{k}={v}" for k, v in type_counts.most_common())
rows_events = [(idx, row.gene_id, row.gene_name, row.event_type, row.dPSI, row.pval)
               for idx, row in dpsi_sig.iterrows()]
write_sheet(wb3, "Significant splicing events",
            f"{len(dpsi_sig)} significant splicing events (|dPSI|>{DPSI_THRESH}, p<{SPLICE_P_THRESH}) across "
            f"all 7 SUPPA event types. Breakdown: {breakdown}.",
            ["event_id", "gene_id", "gene_name", "event_type", "dPSI", "pval"],
            rows_events, col_widths=[55, 20, 14, 12, 14, 16], first=True)
wb3.save(os.path.join(OUT, "GSE133398_significant_splicing_events.xlsx"))
print("  Saved.")

print("\n=== ALL 3 TABLE FILES DONE ===")
print(f"Output folder: {OUT}")
