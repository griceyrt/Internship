#!/usr/bin/env python3
"""
GSE171975 (CT20) DE figures -- second independent Asher dataset validation.
Author: Gricey

Same treatment as GSE130613's publication figures (make_gse130613_publication_figures.py):
no title, legend outside plot upper right with counts, labeled dots get a
black outline. Volcano padj<0.05/|log2FC|>0.5, concordance vs Nanopore
padj<0.05/|log2FC|>1.5.

Run from: orthogonal_validation/
"""

import os
import numpy as np
import pandas as pd
import openpyxl
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D
from adjustText import adjust_text

matplotlib.rcParams.update({
    "font.family":       "sans-serif",
    "font.sans-serif":   ["Arial", "Helvetica", "DejaVu Sans"],
    "font.size":         9,
    "axes.spines.top":   False,
    "axes.spines.right": False,
    "axes.edgecolor":    "#AAAAAA",
    "axes.linewidth":    0.8,
    "axes.grid":         False,
    "figure.facecolor":  "white",
    "axes.facecolor":    "white",
    "xtick.color":       "#444444",
    "ytick.color":       "#444444",
    "axes.labelcolor":   "#222222",
})

BASE_DIR  = os.environ.get("OV_BASE_DIR", "/Users/gricey/Desktop/Internship/orthogonal_validation")
DESEQ_FILE = os.path.join(BASE_DIR, "results", "GSE171975", "normalisation_GSE171975_star_htseq", "deseq2_GSE171975_KO_vs_WT.tsv")
NANO_FILE  = os.path.join(BASE_DIR, "data", "Table3-differential_expressiong_WTvsPerKO.xlsx")
FIG_DIR    = os.path.join(BASE_DIR, "results", "GSE171975", "figures")
TABLE_DIR  = os.path.join(BASE_DIR, "results", "GSE171975", "tables")
os.makedirs(FIG_DIR, exist_ok=True)
os.makedirs(TABLE_DIR, exist_ok=True)

LFC_THRESH  = 0.5
PADJ_THRESH = 0.05
LFC_15      = 1.5

CIRCADIAN_GENES = [
    "Ren1", "Orm2", "Prtn3", "Cyp2b9", "Dbp", "Cxcl1", "Saa1", "Ighg2b",
    "Celf4", "Egr1", "Ciart", "Nr1d1", "Nr1d2", "Btg2", "Per3", "Ephx1",
    "Gstm2", "Clock", "Cyp2a5", "Arntl", "Cry1", "Kmt2d", "Gstm3",
    "Prkca", "Cyp7a1", "Fancm", "Cspg5"
]

def save_both(fig, name):
    for ext in ("svg", "png"):
        path = os.path.join(FIG_DIR, f"{name}.{ext}")
        fig.savefig(path, format=ext, dpi=200, bbox_inches="tight")
        print(f"  Saved: {path}")
    plt.close(fig)

print("=== Loading GSE171975 DE result ===")
de = pd.read_csv(DESEQ_FILE, sep="\t").dropna(subset=["log2FoldChange", "padj"])
de = de[de["padj"] > 0].set_index("gene_id")
print(f"Genes with reliable padj: {len(de)}")

print("\n=== DE volcano ===")
de["neg_log10_padj"] = -np.log10(de["padj"])
de["category"] = "ns"
de.loc[(de["log2FoldChange"] > LFC_THRESH) & (de["padj"] < PADJ_THRESH), "category"] = "up"
de.loc[(de["log2FoldChange"] < -LFC_THRESH) & (de["padj"] < PADJ_THRESH), "category"] = "down"
for cat, n in de["category"].value_counts().items():
    print(f"  {cat}: {n}")

fig, ax = plt.subplots(figsize=(4.3, 3.78))
ns_mask = de["category"] == "ns"
ax.scatter(de.loc[ns_mask, "log2FoldChange"], de.loc[ns_mask, "neg_log10_padj"],
           c="#CCCCCC", s=4, alpha=0.4, linewidths=0, zorder=1,
           label=f"not significant (n={ns_mask.sum()})")
up_mask = de["category"] == "up"
ax.scatter(de.loc[up_mask, "log2FoldChange"], de.loc[up_mask, "neg_log10_padj"],
           c="#D55E00", s=14, alpha=0.85, linewidths=0, zorder=2,
           label=f"up in PerDKO (n={up_mask.sum()})")
dn_mask = de["category"] == "down"
ax.scatter(de.loc[dn_mask, "log2FoldChange"], de.loc[dn_mask, "neg_log10_padj"],
           c="#0072B2", s=14, alpha=0.85, linewidths=0, zorder=2,
           label=f"down in PerDKO (n={dn_mask.sum()})")
ax.axvline(LFC_THRESH, color="#BBBBBB", ls="--", lw=0.7)
ax.axvline(-LFC_THRESH, color="#BBBBBB", ls="--", lw=0.7)
ax.axhline(-np.log10(PADJ_THRESH), color="#BBBBBB", ls="--", lw=0.7)

texts, labels_out = [], []
for gname in CIRCADIAN_GENES:
    matches = de[(de["category"] != "ns") & (de.get("gene_name") == gname)] if "gene_name" in de.columns else de.iloc[0:0]
    if len(matches) == 0:
        continue
    row = matches.iloc[0]
    texts.append(ax.text(row["log2FoldChange"], row["neg_log10_padj"], gname, fontsize=7, fontstyle="italic"))
    labels_out.append((gname, row["log2FoldChange"], row["neg_log10_padj"], row["category"]))

for gname, dx, dy, cat in labels_out:
    fill = "#D55E00" if cat == "up" else "#0072B2"
    ax.scatter([dx], [dy], c=fill, s=22, alpha=0.95, edgecolors="black", linewidths=0.7, zorder=4)

if texts:
    adjust_text(texts, ax=ax, expand=(2, 2), force_text=(1.2, 1.2), force_points=(1, 1), lim=400)
    for txt, (gname, dx, dy, cat) in zip(texts, labels_out):
        lx, ly = txt.get_position()
        txt.remove()
        ax.annotate(gname, xy=(dx, dy), xytext=(lx, ly), fontsize=7, fontstyle="italic",
                    ha="center", va="center",
                    arrowprops=dict(arrowstyle="-", color="#888888", lw=0.5, shrinkA=3, shrinkB=3))
print(f"  Circadian genes labeled: {len(labels_out)}")

ax.set_xlabel("log₂FC (PerDKO / WT)")
ax.set_ylabel("−log₁₀(padj)")
ax.legend(fontsize=7, frameon=False, loc="upper left", bbox_to_anchor=(1.02, 1))
save_both(fig, "GSE171975_STARhtseq_DE_volcano")

print("\n=== Concordance vs Nanopore ===")
wb = openpyxl.load_workbook(NANO_FILE)

ws_sig = wb["perKO_sig"]
rows_sig = list(ws_sig.iter_rows(values_only=True))
header_sig = [str(c) if c is not None else "idx" for c in rows_sig[0]]
nano_sig_geneid = {}
for r in rows_sig[1:]:
    row = dict(zip(header_sig, r))
    gid = row.get("gene_id")
    if gid and gid not in nano_sig_geneid:
        try:
            nano_sig_geneid[gid] = float(row["log2FC_perko"])
        except (TypeError, ValueError):
            pass

ws_all = wb["DE_all"]
rows_all = list(ws_all.iter_rows(values_only=True))
header_all = list(rows_all[0])
nano_all_geneid = {}
for r in rows_all[1:]:
    row = dict(zip(header_all, r))
    gid = row.get("gene_id")
    if not gid:
        continue
    try:
        lfc = float(row["log2FC_perko"])
        padj = float(row["padj_per_ko"])
    except (TypeError, ValueError):
        continue
    if gid not in nano_all_geneid or abs(lfc) > abs(nano_all_geneid[gid]["log2FC"]):
        nano_all_geneid[gid] = {"log2FC": lfc, "padj": padj}

ill_sig = set(de[(de["padj"] < PADJ_THRESH) & (de["log2FoldChange"].abs() > LFC_15)].index)
common = set(de.index) & set(nano_all_geneid.keys())
nano_sig = set(nano_sig_geneid.keys())
overlap = ill_sig & nano_sig & common
print(f"  Common genes: {len(common)}")
print(f"  Illumina sig (padj<0.05, |log2FC|>1.5): {len(ill_sig)}")
print(f"  Nanopore sig: {len(nano_sig)}")
print(f"  Overlap: {len(overlap)}")

genes = list(common)
x = np.array([float(de.loc[g, "log2FoldChange"]) for g in genes])
y = np.array([float(nano_all_geneid[g]["log2FC"]) for g in genes])
mask_both = np.array([g in overlap for g in genes])
mask_ill = np.array([g in ill_sig and g not in overlap for g in genes])
mask_nano = np.array([g in nano_sig and g not in overlap for g in genes])
mask_grey = ~(mask_both | mask_ill | mask_nano)

fig, ax = plt.subplots(figsize=(4.5, 4.2))
ax.scatter(x[mask_grey], y[mask_grey], c="#CCCCCC", s=4, alpha=0.35, linewidths=0, zorder=1,
           label=f"not significant (n={mask_grey.sum()})")
ax.scatter(x[mask_ill], y[mask_ill], c="#56B4E9", s=16, alpha=0.85, linewidths=0, zorder=2,
           label=f"Illumina only (n={mask_ill.sum()})")
ax.scatter(x[mask_nano], y[mask_nano], c="#009E73", s=16, alpha=0.85, linewidths=0, zorder=3,
           label=f"Nanopore only (n={mask_nano.sum()})")
ax.scatter(x[mask_both], y[mask_both], c="#E69F00", s=26, alpha=0.95, edgecolors="black", linewidths=0.5, zorder=4,
           label=f"both significant (n={mask_both.sum()})")

lim = max(abs(x).max(), abs(y).max()) * 1.05
ax.plot([-lim, lim], [-lim, lim], color="#888888", lw=0.9, ls="--", alpha=0.6)
ax.axhline(0, color="#BBBBBB", lw=0.6, alpha=0.5)
ax.axvline(0, color="#BBBBBB", lw=0.6, alpha=0.5)

gene_map_ht = de["gene_name"].to_dict() if "gene_name" in de.columns else {}
texts, labels_out2 = [], []
for i, g in enumerate(genes):
    if not mask_both[i]:
        continue
    gname = gene_map_ht.get(g, g)
    if pd.isna(gname) or gname == "":
        gname = g
    ox = 0.6 if x[i] >= 0 else -0.6
    oy = 0.4 if y[i] >= 0 else -0.4
    texts.append(ax.text(x[i] + ox, y[i] + oy, gname, fontsize=7, fontstyle="italic"))
    labels_out2.append((gname, x[i], y[i]))

if texts:
    adjust_text(texts, ax=ax, expand=(3.0, 3.0), force_text=(2.0, 2.0), force_points=(1.5, 1.5), lim=700)
    for txt, (gname, dx, dy) in zip(texts, labels_out2):
        lx, ly = txt.get_position()
        txt.remove()
        ax.annotate(gname, xy=(dx, dy), xytext=(lx, ly), fontsize=7, fontstyle="italic",
                    ha="center", va="center",
                    arrowprops=dict(arrowstyle="-", color="#888888", lw=0.6, shrinkA=4, shrinkB=4))
print(f"  Overlap genes labeled: {len(labels_out2)}")

ax.set_xlabel("log₂FC Illumina, STAR+htseq (PerDKO / WT)")
ax.set_ylabel("log₂FC Nanopore (PerDKO / WT)")
ax.legend(fontsize=7, frameon=False, loc="upper left", bbox_to_anchor=(1.02, 1))
save_both(fig, "GSE171975_STARhtseq_vs_Nanopore_DE_concordance_gene_level")

print("\n=== Building overlap table ===")
rows_out = []
for g in sorted(overlap):
    gname_out = gene_map_ht.get(g, "")
    if pd.isna(gname_out) or gname_out == "":
        gname_out = g  # ~20% of genes in this GTF have no gene_name attribute, fall back to gene_id
    rows_out.append({
        "gene_id": g,
        "gene_name": gname_out,
        "log2FC_Illumina_GSE171975": de.loc[g, "log2FoldChange"],
        "padj_Illumina_GSE171975": de.loc[g, "padj"],
        "log2FC_Nanopore": nano_all_geneid[g]["log2FC"],
        "padj_Nanopore": nano_all_geneid[g]["padj"],
        "same_direction": (de.loc[g, "log2FoldChange"] > 0) == (nano_all_geneid[g]["log2FC"] > 0),
    })
overlap_df = pd.DataFrame(rows_out).sort_values("padj_Illumina_GSE171975")
same_dir_n = overlap_df["same_direction"].sum() if len(overlap_df) else 0
print(f"  Overlap genes: {len(overlap_df)}, same direction: {same_dir_n}, opposite: {len(overlap_df) - same_dir_n}")

xlsx_path = os.path.join(TABLE_DIR, "GSE171975_STARhtseq_vs_Nanopore_DE_overlap.xlsx")
with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
    overlap_df.to_excel(writer, sheet_name="Overlap_genes", index=False)
    summary = pd.DataFrame({
        "Metric": ["Genes tested (>=10 reads filter)", "Genes with reliable padj",
                   "Significant DE genes (padj<0.05, |log2FC|>0.5)",
                   "Common genes tested (both platforms)",
                   "Illumina sig (padj<0.05, |log2FC|>1.5)", "Nanopore sig",
                   "Overlap (both sig)", "Same direction", "Opposite direction"],
        "Value": [None, len(de), int((de["category"] != "ns").sum()), len(common),
                  len(ill_sig), len(nano_sig), len(overlap), int(same_dir_n), int(len(overlap_df) - same_dir_n)],
    })
    summary.to_excel(writer, sheet_name="Summary", index=False)

wb_out = openpyxl.load_workbook(xlsx_path)
for ws_name in wb_out.sheetnames:
    ws_out = wb_out[ws_name]
    for cell in ws_out[1]:
        cell.font = openpyxl.styles.Font(name="Arial", bold=True)
    for row in ws_out.iter_rows(min_row=2):
        for cell in row:
            cell.font = openpyxl.styles.Font(name="Arial")
    for col_idx, col_cells in enumerate(ws_out.columns, start=1):
        ws_out.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 24
wb_out.save(xlsx_path)
print(f"Saved: {xlsx_path}")

print("\n=== DONE ===")
