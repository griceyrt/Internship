#!/usr/bin/env python3
"""
GSE171975 CT20 -- all significant DE genes, with Kiran's curated circadian
gene list highlighted in yellow for his review.
Author: Gricey

Significance: padj<0.05, |log2FC|>0.5 (same threshold as the volcano figure).

Run from: orthogonal_validation/
"""
import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = os.environ.get("OV_BASE_DIR", "/Users/gricey/Desktop/Internship/orthogonal_validation")
DESEQ_FILE = os.path.join(BASE_DIR, "results", "GSE171975", "normalisation_GSE171975_star_htseq", "deseq2_GSE171975_KO_vs_WT.tsv")
OUT_PATH = os.path.join(BASE_DIR, "results", "GSE171975", "tables", "GSE171975_significant_genes_circadian_review.xlsx")

LFC_THRESH = 0.5
PADJ_THRESH = 0.05

CIRCADIAN_GENES = [
    "Ren1", "Orm2", "Prtn3", "Cyp2b9", "Dbp", "Cxcl1", "Saa1", "Ighg2b",
    "Celf4", "Egr1", "Ciart", "Nr1d1", "Nr1d2", "Btg2", "Per3", "Ephx1",
    "Gstm2", "Clock", "Cyp2a5", "Arntl", "Cry1", "Kmt2d", "Gstm3",
    "Prkca", "Cyp7a1", "Fancm", "Cspg5"
]

de = pd.read_csv(DESEQ_FILE, sep="\t").dropna(subset=["log2FoldChange", "padj"])
de = de[de["padj"] > 0]

sig = de[(de["padj"] < PADJ_THRESH) & (de["log2FoldChange"].abs() > LFC_THRESH)].copy()
sig["gene_name_display"] = sig["gene_name"]
sig.loc[sig["gene_name_display"].isna() | (sig["gene_name_display"] == ""), "gene_name_display"] = \
    sig.loc[sig["gene_name_display"].isna() | (sig["gene_name_display"] == ""), "gene_id"]
sig["is_circadian"] = sig["gene_name"].isin(CIRCADIAN_GENES)
sig["direction"] = sig["log2FoldChange"].apply(lambda x: "up" if x > 0 else "down")
sig = sig.sort_values("padj")

print(f"Total significant genes: {len(sig)}")
print(f"On Kiran's curated circadian list: {sig['is_circadian'].sum()}")
missing = set(CIRCADIAN_GENES) - set(sig.loc[sig["is_circadian"], "gene_name"])
print(f"Circadian list genes NOT significant here: {len(missing)} -> {sorted(missing)}")

cols = ["gene_id", "gene_name_display", "log2FoldChange", "padj", "direction", "is_circadian"]
out_df = sig[cols].rename(columns={"gene_name_display": "gene_name"})

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Significant_genes"

headers = ["gene_id", "gene_name", "log2FoldChange", "padj", "direction", "on_circadian_list"]
ws.append(headers)
for cell in ws[1]:
    cell.font = Font(name="Arial", bold=True)

yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

for i, row in out_df.reset_index(drop=True).iterrows():
    r = i + 2
    ws.cell(r, 1, row["gene_id"])
    ws.cell(r, 2, row["gene_name"])
    ws.cell(r, 3, float(row["log2FoldChange"]))
    ws.cell(r, 4, float(row["padj"]))
    ws.cell(r, 5, row["direction"])
    ws.cell(r, 6, bool(row["is_circadian"]))
    row_font = Font(name="Arial")
    for c in range(1, 7):
        ws.cell(r, c).font = row_font
    if row["is_circadian"]:
        for c in range(1, 7):
            ws.cell(r, c).fill = yellow

for col_idx in range(1, 7):
    ws.column_dimensions[get_column_letter(col_idx)].width = 18

ws.freeze_panes = "A2"

# Note sheet
ws_note = wb.create_sheet("Note_for_Kiran")
note_text = (
    f"Yellow-highlighted rows are genes on the existing curated circadian/clock gene list "
    f"(used to select labels on the DE volcano figure). {sig['is_circadian'].sum()} of the "
    f"{len(sig)} significant genes here are on that list.\n\n"
    f"{len(missing)} genes from the curated list are NOT significant in this GSE171975 result: "
    f"{', '.join(sorted(missing))}.\n\n"
    "Please review and let us know if the list should be updated (genes added or removed) "
    "based on what's actually significant here."
)
ws_note.cell(1, 1, note_text)
ws_note.cell(1, 1).font = Font(name="Arial", size=10)
ws_note.cell(1, 1).alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
ws_note.column_dimensions["A"].width = 100
ws_note.row_dimensions[1].height = 200

os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
wb.save(OUT_PATH)
print(f"\nSaved: {OUT_PATH}")
