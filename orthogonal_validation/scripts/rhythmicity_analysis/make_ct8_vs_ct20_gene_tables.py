#!/usr/bin/env python3
"""
Build the Excel workbook with gene/event-level tables behind the CT8-vs-CT20
"mis-expressed" and "mis-spliced" figure panels (Fig 3A/B): "Differentially
expressed" (DESeq2 output) and "Differentially spliced" (SUPPA output).
Author: Gricey

Prerequisites:
    Rscript scripts/rebuttal_ct8_vs_ct20_de.R
    bash    scripts/rebuttal_ct8_vs_ct20_splicing.sh

Run from: orthogonal_validation/
    python3 scripts/make_ct8_vs_ct20_gene_tables.py
"""
from __future__ import annotations

import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = PROJECT_ROOT / "data"
RESULTS_DIR = PROJECT_ROOT / "results" / "rhythmicity_analysis" / "rebuttal_ct8_vs_ct20"

GTF_PATH_CANDIDATES = [
    PROJECT_ROOT.parent / "boundary_analysis" / "data" / "transcriptome_productivity.gtf",
    DATA_DIR / "transcriptome_productivity.gtf",
]

LOG2FC_CUTOFF = 1.5
FDR_CUTOFF = 0.05
DPSI_CUTOFF = 0.1
PVAL_CUTOFF = 0.05
EVENT_TYPES = ["SE", "MX", "RI", "A5", "A3", "AL", "AF"]

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill("solid", fgColor="2F5C8A")
BODY_FONT = Font(name="Arial", size=10)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def find_gtf() -> Path:
    for p in GTF_PATH_CANDIDATES:
        if p.exists():
            return p
    raise FileNotFoundError(f"No GTF found in any of: {GTF_PATH_CANDIDATES}")


def load_gene_names_and_biotypes(gtf_path: Path) -> tuple[dict, dict]:
    gene_name, tx_biotype = {}, {}
    with open(gtf_path) as f:
        for line in f:
            if line.startswith("#"):
                continue
            fields = line.rstrip("\n").split("\t")
            if len(fields) < 9 or fields[2] != "transcript":
                continue
            attrs = fields[8]
            g = re.search(r'gene_id "([^"]+)"', attrs)
            t = re.search(r'transcript_id "([^"]+)"', attrs)
            gn = re.search(r'gene_name "([^"]+)"', attrs)
            tb = re.search(r'transcript_biotype "([^"]+)"', attrs)
            if g and gn:
                gene_name[g.group(1)] = gn.group(1)
            if t:
                tx_biotype[t.group(1)] = tb.group(1) if tb else None
    return gene_name, tx_biotype


def classify_tx(transcript_id: str, gene_id: str) -> str:
    if isinstance(transcript_id, str) and transcript_id.startswith("ENS"):
        return "Annotated"
    if isinstance(gene_id, str) and "ENS" in gene_id:
        return "Novel-Annotated"
    return "Novel-Novel"


def load_de_results() -> pd.DataFrame:
    r_path = RESULTS_DIR / "deseq2_R_tx_level_ct8_vs_ct20.csv"
    py_path = RESULTS_DIR / "deseq2_tx_level_ct8_vs_ct20.csv"
    if r_path.exists():
        return pd.read_csv(r_path)
    return pd.read_csv(py_path, index_col=0).reset_index(names="transcript_id")


def build_de_sheet(gene_name: dict, tx_biotype: dict) -> pd.DataFrame:
    de = load_de_results()
    de_valid = de.dropna(subset=["padj"])
    sig = de_valid[(de_valid["log2FoldChange"].abs() > LOG2FC_CUTOFF) & (de_valid["padj"] < FDR_CUTOFF)].copy()
    sig["gene_name"] = sig["gene_id"].map(gene_name).fillna("")
    sig["class"] = sig.apply(lambda r: classify_tx(r["transcript_id"], r["gene_id"]), axis=1)
    sig["biotype"] = sig["transcript_id"].map(tx_biotype).fillna("")
    sig["coding"] = sig["biotype"].apply(lambda b: "Protein coding" if b == "protein_coding" else "Non coding")
    sig["direction"] = sig["log2FoldChange"].apply(lambda x: "Up in CT20" if x > 0 else "Down in CT20")
    sig = sig[["transcript_id", "gene_id", "gene_name", "log2FoldChange", "padj",
               "direction", "class", "biotype", "coding"]]
    return sig.sort_values("padj")


def build_splicing_sheet(gene_name: dict) -> pd.DataFrame:
    frames = []
    for et in EVENT_TYPES:
        df = pd.read_csv(RESULTS_DIR / "diff" / f"diff_{et}.dpsi.temp.0", sep="\t", header=None,
                          names=["event_id", "dpsi", "pval"], skiprows=1)
        df["event_type"] = et
        frames.append(df)
    events = pd.concat(frames, ignore_index=True)
    events["dpsi"] = pd.to_numeric(events["dpsi"], errors="coerce")
    events["pval"] = pd.to_numeric(events["pval"], errors="coerce")
    events = events.dropna(subset=["dpsi", "pval"])
    events["gene_id"] = events["event_id"].str.split(";").str[0]
    sig = events[(events["dpsi"].abs() > DPSI_CUTOFF) & (events["pval"] < PVAL_CUTOFF)].copy()
    sig["gene_name"] = sig["gene_id"].map(gene_name).fillna("")
    # dPSI = CT8 - CT20 (see rebuttal_ct8_vs_ct20_splicing.sh); negative = higher in CT20
    sig["direction"] = sig["dpsi"].apply(lambda x: "Up in CT20 (higher PSI)" if x < 0 else "Down in CT20 (lower PSI)")
    sig = sig[["event_id", "gene_id", "gene_name", "event_type", "dpsi", "pval", "direction"]]
    sig = sig.rename(columns={"dpsi": "dPSI", "pval": "p_value"})
    return sig.sort_values("p_value")


def write_sheet(wb: Workbook, name: str, df: pd.DataFrame, title: str, notes: str):
    ws = wb.create_sheet(name)
    ws["A1"] = title
    ws["A1"].font = Font(name="Arial", bold=True, size=13)
    ws["A2"] = notes
    ws["A2"].font = Font(name="Arial", italic=True, size=9, color="555555")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(df.columns))

    start_row = 4
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=start_row, column=j, value=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER

    for i, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.font = BODY_FONT
            c.border = BORDER
            colname = df.columns[j - 1]
            if colname in ("log2FoldChange", "dPSI"):
                c.number_format = "0.00"
            elif colname in ("padj", "p_value"):
                c.number_format = "0.000E+00"

    for j, col in enumerate(df.columns, start=1):
        maxlen = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str)])
        ws.column_dimensions[get_column_letter(j)].width = min(max(maxlen + 3, 10), 45)

    ws.freeze_panes = f"A{start_row + 1}"
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(df.columns))}{start_row + len(df)}"


def main():
    gtf_path = find_gtf()
    print("Using GTF:", gtf_path)
    gene_name, tx_biotype = load_gene_names_and_biotypes(gtf_path)

    de_sheet = build_de_sheet(gene_name, tx_biotype)
    print(f"Differentially expressed: {len(de_sheet)} transcripts, {de_sheet['gene_id'].nunique()} genes")

    splicing_sheet = build_splicing_sheet(gene_name)
    print(f"Differentially spliced: {len(splicing_sheet)} events, {splicing_sheet['gene_id'].nunique()} genes")

    wb = Workbook()
    wb.remove(wb.active)

    write_sheet(
        wb, "Differentially expressed", de_sheet,
        "Differentially expressed transcripts, CT8 vs CT20",
        "Thresholds: padj < 0.05 & |log2FoldChange| > 1.5. Transcript-level DESeq2 (R), "
        "contrast = condition CT20 vs CT8, matching fig4_1_2024_12_04.R's method. "
        f"{len(de_sheet)} significant transcripts, {de_sheet['gene_id'].nunique()} unique genes.",
    )
    write_sheet(
        wb, "Differentially spliced", splicing_sheet,
        "Differentially spliced events, CT8 vs CT20",
        "Thresholds: |dPSI| > 0.1 & p < 0.05. SUPPA diffSplice (empirical method), all 7 event types "
        "(SE/MX/RI/A5/A3/AL/AF), matching the method used to build the original WT-vs-PerKO comparison. "
        f"{len(splicing_sheet)} significant events, {splicing_sheet['gene_id'].nunique()} unique genes.",
    )

    out_path = RESULTS_DIR / "CT8_vs_CT20_misexpressed_misspliced.xlsx"
    wb.save(out_path)
    print("Saved:", out_path)


if __name__ == "__main__":
    main()
