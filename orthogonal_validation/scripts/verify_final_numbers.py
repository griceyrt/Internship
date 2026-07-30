"""
Verifies the final numbers reported to Kiran/Bharath:
  - STAR+htseq-count (original GTF): total sig genes, up/down split
  - Overlap with Nanopore: how many genes significant in both, and whether
    they agree or disagree on direction (up/up, down/down, or mismatched)

Produces both console output AND an Excel workbook (verify_final_numbers.xlsx)
with three sheets:
  - Summary: key counts, computed via formulas (COUNTIFS/SUMPRODUCT) referencing
    the raw data sheets, so it recalculates if the underlying data changes
  - All_STAR_htseq_genes: full gene-level DE result table with a significance flag
  - Nanopore_overlap: the genes significant in both platforms, with a direction
    concordance column

Run from: orthogonal_validation/
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

LFC_THRESH = 0.5
PADJ_THRESH = 0.05
OUT_XLSX = "results/marsseq_sanity_check/verify_final_numbers.xlsx"


def gene_level_dedup(df, lfc_col, padj_col, id_col):
    df = df.dropna(subset=[lfc_col, padj_col])
    df = df[df[padj_col] > 0]
    df = (df.sort_values(lfc_col, key=abs, ascending=False)
            .drop_duplicates(subset=id_col)
            .set_index(id_col))
    return df


# =============================================================================
# Load STAR+htseq (original GTF) -- the final result
# =============================================================================
star = pd.read_csv("results/normalisation_star_htseq_original_gtf/deseq2_star_htseq_KO_vs_WT.tsv", sep="\t")
star = star.dropna(subset=["log2FoldChange", "padj"]).set_index("gene_id")
star["significant"] = (star["padj"] < PADJ_THRESH) & (star["log2FoldChange"].abs() > LFC_THRESH)
star["direction"] = star["log2FoldChange"].apply(lambda x: "up" if x > 0 else "down")

total_genes = len(star)
sig = star[star["significant"]]
n_sig = len(sig)
n_up = (sig["direction"] == "up").sum()
n_down = (sig["direction"] == "down").sum()

print("=== STAR+htseq (original GTF) ===")
print(f"Total genes tested: {total_genes}")
print(f"Significant (padj<0.05, |log2FC|>0.5): {n_sig} ({n_up} up, {n_down} down)")

# =============================================================================
# Load Nanopore, merge, compute overlap + direction concordance
# =============================================================================
nano_raw = pd.read_excel("data/Table3-differential_expressiong_WTvsPerKO.xlsx", sheet_name="DE_all")
nano_gene = gene_level_dedup(nano_raw, "log2FC_perko", "padj_per_ko", "gene_id")

merged = star.join(nano_gene, how="inner", lsuffix="_ill", rsuffix="_nano")
merged["nano_significant"] = (merged["padj_per_ko"] < PADJ_THRESH) & (merged["log2FC_perko"].abs() > LFC_THRESH)
both = merged[merged["significant"] & merged["nano_significant"]].copy()

both["ill_direction"] = both["log2FoldChange"].apply(lambda x: "up" if x > 0 else "down")
both["nano_direction"] = both["log2FC_perko"].apply(lambda x: "up" if x > 0 else "down")
both["concordant"] = both["ill_direction"] == both["nano_direction"]

n_overlap = len(both)
n_up_up = ((both["ill_direction"] == "up") & (both["nano_direction"] == "up")).sum()
n_down_down = ((both["ill_direction"] == "down") & (both["nano_direction"] == "down")).sum()
n_mismatch = (~both["concordant"]).sum()

print("\n=== Overlap with Nanopore (both significant) ===")
print(f"Total overlap: {n_overlap}")
print(f"  Up in both: {n_up_up}")
print(f"  Down in both: {n_down_down}")
print(f"  Mismatched direction: {n_mismatch}")

# =============================================================================
# Build Excel workbook
# =============================================================================
wb = openpyxl.Workbook()
FONT = "Arial"
header_font = Font(name=FONT, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
title_font = Font(name=FONT, bold=True, size=14)
label_font = Font(name=FONT, bold=True)
normal_font = Font(name=FONT)

# --- Sheet: All_STAR_htseq_genes (raw data, written first so Summary can reference it) ---
ws_genes = wb.active
ws_genes.title = "All_STAR_htseq_genes"
gene_cols = ["gene_id", "gene_name", "log2FoldChange", "padj", "direction", "significant"]
ws_genes.append(gene_cols)
for c in range(1, len(gene_cols) + 1):
    cell = ws_genes.cell(row=1, column=c)
    cell.font = header_font
    cell.fill = header_fill
star_out = star.reset_index()[["gene_id", "gene_name", "log2FoldChange", "padj", "direction", "significant"]]
for row in star_out.itertuples(index=False):
    ws_genes.append(list(row))
for c in range(1, len(gene_cols) + 1):
    ws_genes.column_dimensions[get_column_letter(c)].width = 18

# --- Sheet: Nanopore_overlap (raw data) ---
ws_overlap = wb.create_sheet("Nanopore_overlap")
overlap_cols = ["gene_id", "gene_name", "log2FoldChange_illumina", "padj_illumina",
                "log2FC_perko_nanopore", "padj_per_ko_nanopore",
                "ill_direction", "nano_direction", "concordant"]
ws_overlap.append(overlap_cols)
for c in range(1, len(overlap_cols) + 1):
    cell = ws_overlap.cell(row=1, column=c)
    cell.font = header_font
    cell.fill = header_fill
both_out = both.reset_index()[["gene_id", "gene_name", "log2FoldChange", "padj",
                                 "log2FC_perko", "padj_per_ko", "ill_direction", "nano_direction"]]
for row in both_out.itertuples(index=False):
    ws_overlap.append(list(row))
# "concordant" as a real formula per row (col G = ill_direction, col H = nano_direction),
# instead of a pre-computed value -- keeps every row uniform and genuinely recalculating
for r in range(2, len(both_out) + 2):
    ws_overlap.cell(row=r, column=9, value=f"=G{r}=H{r}")
for c in range(1, len(overlap_cols) + 1):
    ws_overlap.column_dimensions[get_column_letter(c)].width = 20

# --- Sheet: Summary (formulas referencing the two raw sheets above) ---
ws_sum = wb.create_sheet("Summary", 0)  # insert as first sheet
ws_sum["A1"] = "GSE130613 -- final numbers verification"
ws_sum["A1"].font = title_font
ws_sum["A2"] = "STAR + htseq-count (original GTF), vs Nanopore. Formulas below recompute from the raw data sheets."
ws_sum["A2"].font = Font(name=FONT, italic=True, size=9)

n_gene_rows = len(star_out)
n_overlap_rows = len(both_out)

rows = [
    ("", ""),
    ("STAR+htseq-count (original GTF)", ""),
    ("Total genes tested", f"=COUNTA(All_STAR_htseq_genes!A2:A{n_gene_rows+1})"),
    ("Significant genes (padj<0.05, |log2FC|>0.5)", f'=COUNTIF(All_STAR_htseq_genes!F2:F{n_gene_rows+1},TRUE)'),
    ("  Up in PerDKO", f'=COUNTIFS(All_STAR_htseq_genes!F2:F{n_gene_rows+1},TRUE,All_STAR_htseq_genes!E2:E{n_gene_rows+1},"up")'),
    ("  Down in PerDKO", f'=COUNTIFS(All_STAR_htseq_genes!F2:F{n_gene_rows+1},TRUE,All_STAR_htseq_genes!E2:E{n_gene_rows+1},"down")'),
    ("", ""),
    ("Overlap with Nanopore (significant in both)", ""),
    ("Total overlap", f"=COUNTA(Nanopore_overlap!A2:A{n_overlap_rows+1})"),
    ("  Up in both (concordant)", f'=COUNTIFS(Nanopore_overlap!G2:G{n_overlap_rows+1},"up",Nanopore_overlap!H2:H{n_overlap_rows+1},"up")'),
    ("  Down in both (concordant)", f'=COUNTIFS(Nanopore_overlap!G2:G{n_overlap_rows+1},"down",Nanopore_overlap!H2:H{n_overlap_rows+1},"down")'),
    ("  Mismatched direction", f'=COUNTIF(Nanopore_overlap!I2:I{n_overlap_rows+1},FALSE)'),
]
r = 3
for label, formula in rows:
    ws_sum.cell(row=r, column=1, value=label).font = label_font if formula != "" or label.endswith(")") else normal_font
    if formula:
        ws_sum.cell(row=r, column=2, value=formula).font = normal_font
    r += 1

ws_sum.column_dimensions["A"].width = 45
ws_sum.column_dimensions["B"].width = 15

wb.save(OUT_XLSX)
print(f"\nSaved: {OUT_XLSX}")
